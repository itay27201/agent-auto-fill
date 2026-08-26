"""
insurance_diff.py
-----------------
Compares a set of "input" insurance spec documents against a set of "reference" documents
using Claude via AWS Bedrock, then writes the findings to a colour-coded Excel workbook.

Usage
-----
    from insurance_diff import analyze_insurance_docs

    analyze_insurance_docs(
        input_files=["tender.pdf"],
        reference_files=["ref_fire.pdf", "ref_liability.pdf"],
        output_xlsx="gaps.xlsx",
        aws_region="us-east-1",
        model="eu.anthropic.claude-sonnet-4-6",
    )

AWS Credentials
---------------
    No API key needed. Credentials are resolved automatically from:
      - AWS_ACCESS_KEY_ID / AWS_SECRET_ACCESS_KEY environment variables
      - ~/.aws/credentials file
      - IAM role (EC2 / Lambda / ECS)

Requirements
------------
    pip install anthropic[bedrock] pypdf openpyxl python-docx olefile
"""

from __future__ import annotations

import configparser
import json
import os
import re
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Optional, Sequence

import boto3
from botocore.config import Config
from botocore.exceptions import ClientError
from anthropic import AnthropicBedrock
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Bedrock output token limit per reference comparison call (35–50 Hebrew findings need headroom)
DEFAULT_MAX_TOKENS = 64_000
# Cheap/fast model used only for JSON repair fallback — no domain knowledge required
_JSON_REPAIR_MODEL = os.environ.get(
    "JSON_REPAIR_MODEL",
    "eu.anthropic.claude-haiku-4-5-20251001-v1:0",
)
DEFAULT_INPUT_MAX_CHARS = 100_000
DEFAULT_REF_MAX_CHARS = 80_000
REF_BUNDLE_SINGLE_MAX_CHARS = 150_000
REF_BUNDLE_BATCH_MAX_CHARS = 120_000
MAX_REFERENCE_FILES_BUNDLED = 20
REF_FILE_HEADER = "\n\n=== REFERENCE FILE: {name} ===\n"
TRUNCATION_MARKER = "[... קוצר בשל אורך ...]"
DEFAULT_S3_PREFIX = "doc-comparison"
REFERENCE_S3_PREFIX = "doc-comparison-reference"
INPUT_S3_PREFIX = "doc-comparison-input"
DOC_EXTENSIONS = frozenset({"pdf", "doc", "docx"})
DEFAULT_AWS_REGION = "eu-central-1"
CHUNKS_MIN_TEXT_CHARS = 1000
REF_CORPUS_SUBDIR = "ref-corpus"
INPUT_SPLIT_PARTS = 10
INPUT_SPLIT_MAX_WORKERS = 10
INPUT_SPLIT_LLM_CALLS = INPUT_SPLIT_PARTS  # one combined pass per window
SPARSE_PAGE_MIN_FINDINGS = 3  # unused for retry; page retry triggers only at 0 findings
SPARSE_PAGE_RETRY_MAX_WORKERS = int(os.environ.get("SPARSE_PAGE_RETRY_MAX_WORKERS", "5"))
try:
    import bedrock_models as _bedrock_models_cfg

    _NOVA_MODEL_DEFAULT = (
        os.environ.get("BEDROCK_MODEL_ID_NOVA")
        or _bedrock_models_cfg.CONFIG.get("model_id_nova")
        or "eu.amazon.nova-pro-v1:0"
    )
except Exception:
    _NOVA_MODEL_DEFAULT = os.environ.get(
        "BEDROCK_MODEL_ID_NOVA", "eu.amazon.nova-pro-v1:0"
    )


import threading

# Near top of analyze_insurance_docs, or at module level:
_cpu_count = os.cpu_count() or 1
try:
    _affinity = len(os.sched_getaffinity(0))   # Linux/Lambda: CPUs the process can actually use
except AttributeError:
    _affinity = _cpu_count                      # Windows fallback

print(
    f"[env] cpu_count={_cpu_count} "
    f"affinity={_affinity} "
    f"INPUT_SPLIT_MAX_WORKERS={INPUT_SPLIT_MAX_WORKERS} "
    f"SPARSE_PAGE_RETRY_MAX_WORKERS={SPARSE_PAGE_RETRY_MAX_WORKERS} "
    f"active_threads={threading.active_count()}"
)
SPARSE_PAGE_RETRY_MODEL = os.environ.get("SPARSE_PAGE_RETRY_MODEL", _NOVA_MODEL_DEFAULT)
SPARSE_PAGE_RETRY_REGION = os.environ.get("BEDROCK_REGION", "eu-west-1")
# Nova Pro Converse API rejects maxTokens >= 10000 (ValidationException).
NOVA_MAX_OUTPUT_TOKENS = 9_999
SPARSE_PAGE_RETRY_MAX_TOKENS = min(
    int(os.environ.get("SPARSE_PAGE_RETRY_MAX_TOKENS", str(NOVA_MAX_OUTPUT_TOKENS))),
    NOVA_MAX_OUTPUT_TOKENS,
)
NOVA_READ_TIMEOUT_SEC = int(os.environ.get("NOVA_READ_TIMEOUT_SEC", "60"))
NOVA_CONNECT_TIMEOUT_SEC = int(os.environ.get("NOVA_CONNECT_TIMEOUT_SEC", "10"))
SPARSE_CHUNK_RETRY_TOP_K_REF = int(
    os.environ.get("SPARSE_CHUNK_RETRY_TOP_K_REF", "15")
)
SPARSE_CHUNK_RETRY_MAX_CHUNK_TEXT = int(
    os.environ.get("SPARSE_CHUNK_RETRY_MAX_CHUNK_TEXT", "6000")
)
SPARSE_CHUNK_RETRY_MAX_INPUT_TEXT = int(
    os.environ.get("SPARSE_CHUNK_RETRY_MAX_INPUT_TEXT", "6000")
)
_NOVA_BEDROCK_CLIENT: Optional[object] = None

S3_CLIENT = boto3.client("s3")
LAMBDA_CLIENT = boto3.client("lambda")
DYNAMODB_RESOURCE = boto3.resource("dynamodb")

PROMPTS_TABLE = os.environ.get("PROMPTS_TABLE", "phoenix-custom-prompts")
CUSTOM_PROMPT_HEADER = "\n\n## כללים מותאמים-אישית (מבוססי משוב)\n"


def _extract_text_from_response(response) -> str:
    for block in response.content:
        if getattr(block, "type", None) == "text":
            return (block.text or "").strip()
    raise ValueError("No text block in model response")


# ---------------------------------------------------------------------------
# Document text extraction
# ---------------------------------------------------------------------------

def _file_ext(path: str | Path) -> str:
    return Path(path).suffix.lower().lstrip(".")


def _extract_pdf_text(path: str | Path) -> list[tuple[int, str]]:
    """Return [(page_number, text), …] for every page in the PDF."""
    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        pages.append((i, text))
    return pages


def _extract_word_pages(path: str | Path) -> list[tuple[int, str]]:
    """Return pseudo-pages from a .doc or .docx file (split on form-feed if present)."""
    from s3_fucntions import read_doc_from_path

    text = read_doc_from_path(path)
    segments = text.split("\x0c") if "\x0c" in text else [text]
    pages = []
    for i, segment in enumerate(segments, 1):
        segment = segment.strip()
        if segment:
            pages.append((i, segment))
    return pages or [(1, "")]


def _extract_document_pages(path: str | Path) -> list[tuple[int, str]]:
    """Return [(page_number, text), …] for PDF or Word (.doc/.docx)."""
    ext = _file_ext(path)
    if ext == "pdf":
        return _extract_pdf_text(path)
    if ext in ("doc", "docx"):
        return _extract_word_pages(path)
    raise ValueError(
        f"Unsupported file type '.{ext}' ({Path(path).name}). "
        "Supported: pdf, doc, docx"
    )


def _pages_to_str(pages: list[tuple[int, str]], max_chars: int = 80_000) -> str:
    """Flatten page list to a single string with page markers, capped at max_chars."""
    parts = []
    total = 0
    for pnum, text in pages:
        chunk = f"\n--- עמוד {pnum} ---\n{text}"
        if total + len(chunk) > max_chars:
            parts.append(f"\n{TRUNCATION_MARKER}")
            break
        parts.append(chunk)
        total += len(chunk)
    return "".join(parts)


def _split_indices_into_groups(
    n_items: int,
    n_groups: int = INPUT_SPLIT_PARTS,
) -> list[tuple[int, int]]:
    """Return n_groups half-open slice ranges [(start, end), ...], evenly distributed."""
    if n_items <= 0:
        return [(0, 0)] * n_groups
    base, extra = divmod(n_items, n_groups)
    ranges: list[tuple[int, int]] = []
    idx = 0
    for i in range(n_groups):
        size = base + (1 if i < extra else 0)
        ranges.append((idx, idx + size))
        idx += size
    return ranges


def _chunk_sort_key(c: dict) -> int:
    try:
        return int(c.get("chunk_index", 0))
    except (TypeError, ValueError):
        return 0


def _chunk_index_at(chunks: list[dict], idx: int) -> int:
    try:
        return int(chunks[idx].get("chunk_index", idx))
    except (TypeError, ValueError):
        return idx


def _chunk_page_number(c: dict) -> Optional[int]:
    try:
        n = int(c.get("page_number") or c.get("page") or 0)
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _focus_context_preamble(
    focus_start: int,
    focus_end_inclusive: int,
    context_before: Optional[int],
    context_after: Optional[int],
    *,
    focus_pages: Optional[list[int]] = None,
) -> str:
    lines = ["## Focus Directive"]
    if focus_pages:
        pages_str = ", ".join(str(p) for p in focus_pages)
        lines.append(
            f"Focus on pages: {pages_str} "
            f"(chunks {focus_start}–{focus_end_inclusive} inclusive)."
        )
    else:
        lines.append(
            f"Focus on chunks {focus_start}–{focus_end_inclusive} (inclusive)."
        )
    ctx: list[str] = []
    if context_before is not None:
        ctx.append(f"chunk {context_before}")
    if context_after is not None:
        ctx.append(f"chunk {context_after}")
    if ctx:
        lines.append(
            f"{' and '.join(ctx)} {'are' if len(ctx) > 1 else 'is'} provided for context only "
            f"— use only to resolve clauses split across chunk boundaries; do not re-report findings whose focus is entirely outside the focus range."
        )
    lines.append(
        "For every finding: input_page MUST be the page number from the nearest `--- עמוד N ---` marker "
        "above the quoted INPUT text (not a REFERENCE page). "
        "Applies to מכתב פניה / מכרז items too — use the page where the quote appears."
    )
    lines.append(
        "Analyze every document layer in the focus pages: מכתב פניה, policy spec, תנאים כלליים."
    )
    lines.append(
        "Per-page rule (hard): For every `--- עמוד N ---` marker in the focus range, "
        "produce at least one finding with input_page = N. "
        "Do not skip a page even if it appears routine or header-only."
    )
    lines.append(
        "Per-section rule (hard): For every section heading, table row, extension clause, "
        "or general condition in a focus page — explicitly classify it as missing / difference / match "
        "before advancing to the next section."
    )
    return "\n".join(lines) + "\n\n"


def _build_chunk_overlap_windows(
    chunks: list[dict],
    *,
    n_groups: int = INPUT_SPLIT_PARTS,
    max_chars: int = DEFAULT_INPUT_MAX_CHARS,
) -> list[dict]:
    """Build overlapping input windows from a chunk list (prev + group + next).

    Each window dict has: text, focus_start, focus_end (inclusive chunk_index values).
    """
    ordered = sorted(chunks, key=_chunk_sort_key)
    n = len(ordered)
    groups = _split_indices_into_groups(n, n_groups)
    windows: list[dict] = []

    for start, end in groups:
        if start >= end:
            windows.append({"text": "", "focus_start": None, "focus_end": None})
            continue
        win_lo = max(0, start - 1)
        win_hi = min(n, end + 1)
        subset = ordered[win_lo:win_hi]
        body = _pages_to_str(_chunks_to_pages(subset), max_chars=max_chars)
        fs = _chunk_index_at(ordered, start)
        fe = _chunk_index_at(ordered, end - 1)
        ctx_before = _chunk_index_at(ordered, start - 1) if start > 0 else None
        ctx_after = _chunk_index_at(ordered, end) if end < n else None
        focus_pages = sorted({
            p for p in (_chunk_page_number(c) for c in ordered[start:end])
            if p is not None
        })
        preamble = _focus_context_preamble(
            fs, fe, ctx_before, ctx_after, focus_pages=focus_pages or None
        )
        windows.append({
            "text": preamble + body,
            "focus_start": fs,
            "focus_end": fe,
        })
    return windows


def _build_page_overlap_windows(
    text: str,
    *,
    n_groups: int = INPUT_SPLIT_PARTS,
) -> list[dict]:
    """PDF fallback: split page-marked text into n_groups with ±1 page overlap."""
    empty = {"text": "", "focus_start": None, "focus_end": None}
    if not text.strip():
        return [dict(empty) for _ in range(n_groups)]

    segments = re.split(r"(?=\n--- עמוד )", text)
    pages = [s for s in segments if s.strip()]
    if not pages:
        return [{"text": text, "focus_start": None, "focus_end": None}] + [
            dict(empty) for _ in range(n_groups - 1)
        ]

    groups = _split_indices_into_groups(len(pages), n_groups)
    windows: list[dict] = []
    for start, end in groups:
        if start >= end:
            windows.append(dict(empty))
            continue
        win_lo = max(0, start - 1)
        win_hi = min(len(pages), end + 1)
        windows.append({
            "text": "".join(pages[win_lo:win_hi]),
            "focus_start": None,
            "focus_end": None,
        })
    return windows


def _build_input_overlap_windows(
    input_text: str,
    input_chunks: Optional[list[dict]] = None,
    *,
    max_chars: int = DEFAULT_INPUT_MAX_CHARS,
) -> list[dict]:
    if input_chunks:
        return _build_chunk_overlap_windows(input_chunks, max_chars=max_chars)
    return _build_page_overlap_windows(input_text)


def _tag_findings_with_focus(findings: list[dict], win: dict) -> list[dict]:
    fs, fe = win.get("focus_start"), win.get("focus_end")
    if fs is None:
        return findings
    return [
        {**f, "_focus_chunk_start": fs, "_focus_chunk_end": fe}
        for f in findings
    ]


def _parse_input_page(page_raw) -> Optional[int]:
    """Parse input_page from findings — bare int, Hebrew label, or N/A."""
    if page_raw is None:
        return None
    s = str(page_raw).strip()
    if not s or s.upper() == "N/A":
        return None
    try:
        return int(s)
    except (TypeError, ValueError):
        pass
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def _parse_page_from_quote(text: str) -> Optional[int]:
    """Extract PDF page number from Hebrew page markers inside a quote."""
    if not text:
        return None
    for pat in (r"עמוד\s*(\d+)", r"עמ['\u2019]\s*(\d+)", r"דף\s*(\d+)"):
        m = re.search(pat, text)
        if m:
            return int(m.group(1))
    return None


def _group_chunks_by_page(chunks: list[dict]) -> dict[int, list[dict]]:
    """Group input chunks by page_number, sorted by chunk_index within each page."""
    by_page: dict[int, list[dict]] = {}
    for c in chunks:
        page_num = c.get("page_number") or c.get("page") or 1
        try:
            page_num = int(page_num)
        except (TypeError, ValueError):
            page_num = 1
        by_page.setdefault(page_num, []).append(c)
    for page_num in by_page:
        by_page[page_num].sort(key=_chunk_sort_key)
    return by_page


def _sparse_page_retry_preamble(
    page_num: int,
    *,
    overlap_before_page: Optional[int] = None,
    overlap_after_page: Optional[int] = None,
    overlap_before_chunk: Optional[int] = None,
    overlap_after_chunk: Optional[int] = None,
) -> str:
    lines = ["## הנחיית מיקוד — ניסיון חוזר לעמוד ללא ממצאים"]
    lines.append(f"התמקד בעמוד {page_num} בלבד — כל ה-chunks בעמוד זה.")
    ctx_parts: list[str] = []
    if overlap_before_page is not None:
        suffix = f" (chunk {overlap_before_chunk})" if overlap_before_chunk is not None else ""
        ctx_parts.append(f"chunk אחרון מעמוד {overlap_before_page}{suffix}")
    if overlap_after_page is not None:
        suffix = f" (chunk {overlap_after_chunk})" if overlap_after_chunk is not None else ""
        ctx_parts.append(f"chunk ראשון מעמוד {overlap_after_page}{suffix}")
    if ctx_parts:
        lines.append(
            f"{' ו-'.join(ctx_parts)} ניתנו להקשר בלבד — "
            "לטיפול בסעיפים שנחתכו בין עמודים; אל תדווח מחדש על סעיפים ששייכים לעמוד אחר."
        )
    lines.append(
        f"לכל ממצא: input_page חייב להיות \"{page_num}\" "
        f"(מספר העמוד מהסימון --- עמוד {page_num} --- בטקסט המצוטט)."
    )
    lines.append(
        "נתח את כל שכבות המסמך בעמוד המיקוד: מכתב פניה, מפרט פוליסות, תנאים כלליים."
    )
    return "\n".join(lines) + "\n\n"


def _build_sparse_page_retry_text(
    page_num: int,
    pages_by_num: dict[int, list[dict]],
    *,
    max_chars: int = DEFAULT_INPUT_MAX_CHARS,
) -> tuple[str, dict]:
    """Build INPUT for zero-finding page retry: focus page chunks + neighbor overlap."""
    focus = pages_by_num.get(page_num, [])
    context_before: list[dict] = []
    context_after: list[dict] = []
    overlap_before_page: Optional[int] = None
    overlap_after_page: Optional[int] = None
    overlap_before_chunk: Optional[int] = None
    overlap_after_chunk: Optional[int] = None

    prev_page = pages_by_num.get(page_num - 1)
    if prev_page:
        context_before = [prev_page[-1]]
        overlap_before_page = page_num - 1
        overlap_before_chunk = _chunk_sort_key(prev_page[-1])

    next_page = pages_by_num.get(page_num + 1)
    if next_page:
        context_after = [next_page[0]]
        overlap_after_page = page_num + 1
        overlap_after_chunk = _chunk_sort_key(next_page[0])

    subset = context_before + focus + context_after
    body = (
        _pages_to_str(_chunks_to_pages(subset), max_chars=max_chars) if subset else ""
    )
    preamble = _sparse_page_retry_preamble(
        page_num,
        overlap_before_page=overlap_before_page,
        overlap_after_page=overlap_after_page,
        overlap_before_chunk=overlap_before_chunk,
        overlap_after_chunk=overlap_after_chunk,
    )
    meta = {
        "focus_chunk_count": len(focus),
        "overlap_before_page": overlap_before_page,
        "overlap_after_page": overlap_after_page,
        "overlap_chunk_indices": [
            i
            for i in (overlap_before_chunk, overlap_after_chunk)
            if i is not None
        ],
        "total_chunks_sent": len(subset),
        "char_count": len(preamble) + len(body),
    }
    return preamble + body, meta


def _build_sparse_chunk_retry_subset(
    page_num: int,
    chunk_idx: int,
    page_chunks: list[dict],
    pages_by_num: dict[int, list[dict]],
    *,
    max_chars: int = DEFAULT_INPUT_MAX_CHARS,
) -> tuple[str, dict]:
    """Build INPUT text for one focus chunk on a sparse page (+ ±1 chunk overlap)."""
    focus_chunk = page_chunks[chunk_idx]
    focus_idx = _chunk_sort_key(focus_chunk)

    context_before: list[dict] = []
    context_after: list[dict] = []
    ctx_before_idx: Optional[int] = None
    ctx_after_idx: Optional[int] = None
    overlap_before_page: Optional[int] = None
    overlap_after_page: Optional[int] = None

    if chunk_idx > 0:
        context_before = [page_chunks[chunk_idx - 1]]
        ctx_before_idx = _chunk_sort_key(context_before[0])
    else:
        prev_page = pages_by_num.get(page_num - 1)
        if prev_page:
            context_before = [prev_page[-1]]
            ctx_before_idx = _chunk_sort_key(context_before[0])
            overlap_before_page = page_num - 1

    if chunk_idx < len(page_chunks) - 1:
        context_after = [page_chunks[chunk_idx + 1]]
        ctx_after_idx = _chunk_sort_key(context_after[0])
    else:
        next_page = pages_by_num.get(page_num + 1)
        if next_page:
            context_after = [next_page[0]]
            ctx_after_idx = _chunk_sort_key(context_after[0])
            overlap_after_page = page_num + 1

    subset = context_before + [focus_chunk] + context_after
    body = (
        _pages_to_str(_chunks_to_pages(subset), max_chars=max_chars) if subset else ""
    )
    focus_pages = sorted({
        p for p in (_chunk_page_number(focus_chunk), page_num) if p is not None
    })
    preamble = _focus_context_preamble(
        focus_idx,
        focus_idx,
        ctx_before_idx,
        ctx_after_idx,
        focus_pages=focus_pages or None,
    )
    preamble += (
        f"לכל ממצא: input_page חייב להיות \"{page_num}\" "
        f"(מספר העמוד מהסימון --- עמוד {page_num} --- בטקסט המצוטט).\n\n"
    )
    meta = {
        "page_number": page_num,
        "chunk_index": focus_idx,
        "chunk_position": chunk_idx + 1,
        "chunks_on_page": len(page_chunks),
        "overlap_before_page": overlap_before_page,
        "overlap_after_page": overlap_after_page,
        "overlap_chunk_indices": [
            i for i in (ctx_before_idx, ctx_after_idx) if i is not None
        ],
        "total_chunks_sent": len(subset),
        "char_count": len(preamble) + len(body),
    }
    return preamble + body, meta


def _query_ref_chunks_for_input_chunk(
    chunk: dict,
    reference_batch_id: str,
    top_k: int,
    *,
    verbose: bool = False,
) -> list[dict]:
    """Query Pinecone for reference chunks most similar to an input chunk."""
    import pinecone_vector_database

    text = _clean_chunk_text(chunk.get("text") or "")[:SPARSE_CHUNK_RETRY_MAX_INPUT_TEXT]
    if not text.strip():
        return []
    try:
        matches = pinecone_vector_database.query_pinecone_matches(
            query=text,
            specific_file_filter={
                "andAll": [{"equals": {"key": "batch_id", "value": reference_batch_id}}]
            },
            numberOfResults=top_k,
        )
        return matches or []
    except Exception as exc:
        if verbose:
            print(f"  [sparse-page-retry] Pinecone query failed: {exc}")
        return []


def _build_nova_chunk_user_message(
    input_text: str,
    ref_matches: list[dict],
    *,
    ref_filename: str,
    ref_text: str = "",
    ref_preamble: str = "",
    use_full_ref_fallback: bool = False,
) -> str:
    """Build Nova user message: Pinecone ref chunks, or full ref_text fallback."""
    if use_full_ref_fallback or not ref_matches:
        ref_block = f"## REFERENCE ({ref_filename})\n"
        if ref_preamble:
            ref_block += ref_preamble.strip() + "\n"
        ref_block += ref_text
        input_block = f"## INPUT (מסמך המכרז)\n{input_text}"
        return f"{ref_block}\n\n{input_block}"

    lines = [
        f"## INPUT (מסמך המכרז)\n{input_text}",
        "",
        f"## Most relevant reference chunks (Pinecone top-{len(ref_matches)}):",
    ]
    for i, m in enumerate(ref_matches, 1):
        ref_body = _clean_chunk_text(m.get("text") or "")[:SPARSE_CHUNK_RETRY_MAX_CHUNK_TEXT]
        ref_page = m.get("page_number") or m.get("page") or "?"
        ref_fname = m.get("file_name") or m.get("source") or "reference"
        lines.append(f"\n[{i}] {ref_fname} | עמ' {ref_page}")
        lines.append(ref_body)
    lines += [
        "",
        "Classify every insurance requirement in the INPUT focus chunk above.",
        "Return only the JSON object as specified in the instructions.",
    ]
    return "\n".join(lines)


def _count_findings_by_page(
    findings: list[dict],
    page_numbers: list[int] | set[int],
) -> dict[int, int]:
    """Count findings per page_number; unknown pages start at 0."""
    counts = {int(p): 0 for p in page_numbers}
    for f in findings:
        page_num = _parse_input_page(f.get("input_page"))
        if page_num is None or page_num not in counts:
            quote = f.get("input_quote") or f.get("input_text") or ""
            fallback = _parse_page_from_quote(quote)
            if fallback is not None and fallback in counts:
                page_num = fallback
        if page_num is not None and page_num in counts:
            counts[page_num] += 1
    return counts


# ---------------------------------------------------------------------------
# Reference corpus sizing and bundling
# ---------------------------------------------------------------------------

def _measure_reference_corpus(
    reference_files: Sequence[str | Path],
) -> list[dict]:
    """Extract per-file stats; cache pages for bundle building."""
    measured: list[dict] = []
    for fpath in sorted(reference_files, key=lambda p: Path(p).name):
        path = Path(fpath)
        pages = _extract_document_pages(path)
        measured.append(
            {
                "path": path,
                "name": path.name,
                "chars": sum(len(t) for _, t in pages),
                "pages": len(pages),
                "pages_data": pages,
            }
        )
    return measured


def _choose_reference_strategy(
    measurements: list[dict],
    forced: Optional[str] = None,
    *,
    single_max: int = REF_BUNDLE_SINGLE_MAX_CHARS,
    batch_max: int = REF_BUNDLE_BATCH_MAX_CHARS,
) -> str:
    """auto | single | dual | per_file — see bundled multi-ref plan."""
    if forced and forced != "auto":
        if forced not in ("single", "dual", "per_file"):
            raise ValueError(f"Invalid reference_strategy: {forced}")
        return forced

    n = len(measurements)
    if n == 0:
        raise ValueError("No reference files to compare")
    total = sum(m["chars"] for m in measurements)

    if n == 1:
        return "single"
    if n > MAX_REFERENCE_FILES_BUNDLED:
        return "per_file"
    if total <= single_max:
        return "single"
    if total <= 2 * batch_max:
        return "dual"
    return "per_file"


def _normalize_classification(value: str) -> str:
    v = (value or "").strip().lower()
    aliases = {
        "התאמה": "match",
        "הבדל": "difference",
        "חסר": "missing",
        "match": "match",
        "difference": "difference",
        "missing": "missing",
    }
    return aliases.get(v, "difference")


def _build_reference_bundle(
    measurements: list[dict],
    max_chars: int,
    *,
    verbose: bool = True,
) -> tuple[str, list[str]]:
    """Concatenate reference texts with file headers up to max_chars."""
    parts: list[str] = []
    total = 0
    included: list[str] = []

    for m in measurements:
        header = REF_FILE_HEADER.format(name=m["name"])
        if total + len(header) >= max_chars:
            if verbose:
                print(f"  [bundle] omitted (budget exhausted): {m['name']}")
            parts.append(f"\n{TRUNCATION_MARKER}")
            break
        remaining = max_chars - total - len(header)
        body = _pages_to_str(m["pages_data"], max_chars=remaining)
        chunk = header + body
        parts.append(chunk)
        total += len(chunk)
        included.append(m["name"])

    return "".join(parts), included


def _split_reference_batches(
    measurements: list[dict],
    budget: int,
) -> tuple[list[dict], list[dict]]:
    """Greedy two-bin split by character count (largest files first)."""
    ordered = sorted(measurements, key=lambda m: m["chars"], reverse=True)
    batch_a: list[dict] = []
    batch_b: list[dict] = []
    sum_a = 0
    sum_b = 0

    for m in ordered:
        if sum_a <= sum_b:
            batch_a.append(m)
            sum_a += m["chars"]
        else:
            batch_b.append(m)
            sum_b += m["chars"]

    return batch_a, batch_b


def _finding_match_key(d: dict) -> tuple[str, str]:
    section = (d.get("section") or "").strip().lower()
    quote = (d.get("input_quote") or d.get("input_text") or "")[:100].strip().lower()
    return section, quote


def _normalize_quote_prefix(q: str) -> str:
    """Strip structured-quote labels and section-number punctuation for fuzzy comparison.

    Different passes may format the same clause as:
      "ציטוט:\nסעיף 5: מוסכם כי..."
      "ציטוט:\nסעיף 5 — מוסכם כי..."
    The colon vs. em-dash difference causes the raw [:40] prefix check to fail.
    This normalizer removes those formatting artefacts so the substantive text
    can be compared regardless of which pass produced the finding.
    """
    # Remove "ציטוט:\n" or "ציטוט: " preamble
    q = re.sub(r"ציטוט:\s*", "", q)
    # Remove section-number prefix: "סעיף X:" / "סעיף X —" / "סעיף X–" etc.
    q = re.sub(r"^סעיף\s+[\d.\u05d0-\u05ea]+\s*[:\-\u2013\u2014]\s*", "", q.strip())
    # Collapse all whitespace
    q = re.sub(r"\s+", " ", q).strip()
    return q[:50]


# Common words that carry no discriminating signal in section names.
_SECTION_STOP = frozenset({
    "ביטוח", "סעיף", "פרק", "הרחבת", "הרחבה", "כיסוי", "גבול", "סכום",
    "לפי", "של", "עם", "ל", "ב", "מ", "ו",
    "of", "the", "and", "for", "in", "coverage", "clause", "section",
})


def _section_significant_words(d: dict) -> frozenset:
    """Return significant (non-stop) words from the section name, for fuzzy overlap.

    When the section uses the common "Chapter - SubClause" format (e.g.
    "ביטוח אחריות מקצועית - כיסוי לאחר הפסקת פעילות") only the sub-clause
    part (after the last " - ") is used.  Using the last segment avoids treating
    every clause inside the same chapter (e.g. "פרק א' - ביטוח אש מורחב, סעיף N")
    as overlapping just because they share chapter-level words like "אש" + "מורחב".
    """
    section = (d.get("section") or "").strip().lower()
    if " - " in section:
        section = section.rsplit(" - ", 1)[-1]
    words = re.findall(r"[\u05d0-\u05ea]{2,}|[a-z]{3,}|\d{5,}", section)
    return frozenset(w for w in words if w not in _SECTION_STOP)


def _findings_overlap(a: dict, b: dict) -> bool:
    s1, q1 = _finding_match_key(a)
    s2, q2 = _finding_match_key(b)
    # Exact section match
    if s1 and s1 == s2:
        return True
    # Input-quote containment or shared prefix (raw)
    if q1 and q2 and (q1 in q2 or q2 in q1 or q1[:40] == q2[:40]):
        return True
    # Normalized prefix match — catches formatting differences ("5:" vs "5 —")
    # between passes that label the same clause differently
    np1 = _normalize_quote_prefix(q1)
    np2 = _normalize_quote_prefix(q2)
    if np1 and np2 and np1 == np2:
        return True
    # Word-set overlap on section names: ≥2 significant words in common → same clause
    sw1 = _section_significant_words(a)
    sw2 = _section_significant_words(b)
    if len(sw1) >= 2 and len(sw2) >= 2 and len(sw1 & sw2) >= 2:
        return True
    return False


def _dedupe_findings_overlap(a: dict, b: dict) -> bool:
    """Stricter overlap check used only for cross-window dedup.

    Omits the word-set section-name overlap (check 4 in _findings_overlap) to
    avoid collapsing findings that share topic words but refer to genuinely
    distinct clauses (e.g. a general waiver clause vs. a vendor-specific one).
    Only exact section name and quote-based signals are used.
    """
    s1, q1 = _finding_match_key(a)
    s2, q2 = _finding_match_key(b)
    if s1 and s1 == s2:
        return True
    if q1 and q2 and (q1 in q2 or q2 in q1 or q1[:40] == q2[:40]):
        return True
    np1 = _normalize_quote_prefix(q1)
    np2 = _normalize_quote_prefix(q2)
    if np1 and np2 and np1 == np2:
        return True
    return False


def _extract_section_numbers_from_finding(d: dict) -> frozenset[str]:
    """Extract סעיף numbers from section names and quote fields."""
    parts = [
        d.get("section") or "",
        d.get("input_text") or d.get("input_quote") or "",
        d.get("reference_text") or d.get("ref_quote") or "",
    ]
    text = " ".join(parts)
    return frozenset(re.findall(r"סעיף\s*(\d+(?:\.\d+)*)", text))


def _input_pages_for_finding(d: dict) -> frozenset[int]:
    """Parse input_page into a set of integers (handles ranges like '51-52')."""
    page = d.get("input_page") or ""
    pages: set[int] = set()
    for part in str(page).replace("–", "-").split("-"):
        part = part.strip()
        if part.isdigit():
            pages.add(int(part))
    return frozenset(pages)


def _quotes_substantively_overlap(a: dict, b: dict) -> bool:
    """True when input quotes refer to the same clause text."""
    _, q1 = _finding_match_key(a)
    _, q2 = _finding_match_key(b)
    if not q1 or not q2:
        return False
    if q1 in q2 or q2 in q1 or q1[:40] == q2[:40]:
        return True
    np1 = _normalize_quote_prefix(q1)
    np2 = _normalize_quote_prefix(q2)
    return bool(np1 and np2 and np1 == np2)


def _findings_conflict_overlap(a: dict, b: dict) -> bool:
    """Stricter overlap for cross-classification conflict grouping (not window dedupe).

    Groups only on exact section match, substantive quote match, or shared סעיף
    number — never on chapter-level section word overlap alone.
    """
    s1, q1 = _finding_match_key(a)
    s2, q2 = _finding_match_key(b)
    if s1 and s1 == s2:
        return True
    if _quotes_substantively_overlap(a, b):
        return True
    nums_a = _extract_section_numbers_from_finding(a)
    nums_b = _extract_section_numbers_from_finding(b)
    if nums_a and nums_b and (nums_a & nums_b):
        return True
    return False


def _severity_rank(d: dict) -> int:
    sev = d.get("severity", "נמוכה")
    if sev in ("high", "גבוהה"):
        return 0
    if sev in ("medium", "בינונית"):
        return 1
    return 2


def _merge_batched_findings(
    findings_a: list[dict],
    findings_b: list[dict],
    *,
    verbose: bool = True,
) -> list[dict]:
    """Merge dual-batch results; drop false missing when other batch found a match."""
    kept_a: list[dict] = []
    for fa in findings_a:
        if _normalize_classification(fa.get("classification", "")) != "missing":
            kept_a.append(fa)
            continue
        if any(
            _findings_overlap(fa, fb)
            and _normalize_classification(fb.get("classification", "")) in ("match", "difference")
            for fb in findings_b
        ):
            continue
        kept_a.append(fa)

    merged = kept_a + list(findings_b)

    deduped: list[dict] = []
    for f in merged:
        dup_idx = None
        for i, g in enumerate(deduped):
            if (
                _normalize_classification(f.get("classification", ""))
                == _normalize_classification(g.get("classification", ""))
                and _findings_overlap(f, g)
            ):
                dup_idx = i
                break
        if dup_idx is None:
            deduped.append(f)
        elif _severity_rank(f) < _severity_rank(deduped[dup_idx]):
            deduped[dup_idx] = f

    if verbose and len(merged) != len(deduped):
        print(f"  [merge] deduped {len(merged)} -> {len(deduped)} findings")

    if verbose and len(deduped) > 55:
        print(f"  [WARN] {len(deduped)} findings after merge (target ~35-50)")

    return deduped


def _dedupe_findings(
    findings: list[dict], *, verbose: bool = True
) -> tuple[list[dict], list[dict]]:
    """Drop duplicate findings (same classification + overlapping section/quote).

    Returns (kept, dropped) where dropped items include _dropped_reason metadata.
    """
    deduped: list[dict] = []
    dropped: list[dict] = []
    for f in findings:
        dup_idx = None
        for i, g in enumerate(deduped):
            if (
                _normalize_classification(f.get("classification", ""))
                == _normalize_classification(g.get("classification", ""))
                and _dedupe_findings_overlap(f, g)
            ):
                dup_idx = i
                break
        if dup_idx is None:
            deduped.append(f)
        else:
            winner = deduped[dup_idx]
            loser = f
            if _severity_rank(f) < _severity_rank(winner):
                loser = winner
                deduped[dup_idx] = f
                winner = f
            dropped.append({
                **loser,
                "_dropped_reason": "overlap dedupe",
                "_dedupe_winner_section": winner.get("section", ""),
                "_dedupe_winner_classification": _normalize_classification(
                    winner.get("classification", "")
                ),
            })

    if verbose and len(findings) != len(deduped):
        print(
            f"  [dedupe] {len(findings)} -> {len(deduped)} findings "
            f"(dropped={len(dropped)})"
        )
    return deduped, dropped


def _findings_count_summary(findings: list[dict]) -> str:
    """Compact breakdown string for logging, e.g. 'total=5 diff=2 miss=1 match=2'."""
    if not findings:
        return "total=0"
    counts = {"difference": 0, "missing": 0, "match": 0}
    for f in findings:
        cls = _normalize_classification(f.get("classification", ""))
        if cls in counts:
            counts[cls] += 1
    parts = [f"total={len(findings)}"]
    for key, short in (("difference", "diff"), ("missing", "miss"), ("match", "match")):
        if counts[key]:
            parts.append(f"{short}={counts[key]}")
    return " ".join(parts)


def _window_focus_label(win: dict) -> str:
    fs, fe = win.get("focus_start"), win.get("focus_end")
    if fs is not None and fe is not None:
        return f"chunks {fs}-{fe}"
    return "pages"


def _ensure_ref_file_on_findings(findings: list[dict], default_file: str = "") -> None:
    for d in findings:
        if not d.get("ref_file"):
            d["ref_file"] = default_file


# ---------------------------------------------------------------------------
# Claude call
# ---------------------------------------------------------------------------
_JSON_STRING_ESCAPING_RULES = """
## JSON string escaping — CRITICAL

Inside every JSON string value, escape ASCII double-quotes as backslash-quote (write הנ\\"ל not הנ"ל). Hebrew gershayim in abbreviations (ע\\"י, דו\\"ח, תשמ\\"א, מצ\\"ב) must use escaped quotes — bare " characters break JSON.
Do NOT add extra backslashes anywhere else — only use \\ before " for gershayim. Organization names and parentheses: write literal text (no \\\\ padding).
Use JSON literals true, false, and null (not Python True/False/None).
"""

SYSTEM_PROMPT = """You are an expert Israeli insurance policy analyst.

## Primary Rule: Comparison Direction
INPUT  = the tender/procurement document — the requirements the client MANDATES. This is the source of truth.
REFERENCE = the proposed insurance policy — being tested against INPUT requirements.
Every finding must originate from a requirement in INPUT. Never generate findings that originate from REFERENCE only.
Coverage present in REFERENCE but not required by INPUT — ignore completely.

---

## Step A: Internal Mapping (not output)

Before writing a single finding, build an internal checklist of every verifiable clause in INPUT:
- Every chapter (property, liability, equipment, mechanical breakdown, appendices…)
- Every amount / limit / deductible / period row in every table
- Every detailed extension with section number
- Every general condition with contractual impact (cancellation, subrogation, jurisdiction, indemnity)

Do NOT skip a chapter because it looks "standard" — even defaults need verification.

## Step B: Compare

For each item in the Step A checklist — find its counterpart in REFERENCE and classify using the decision tree below.

---

## Classification Decision Tree — follow in exact order

1. Is the topic mentioned at all in REFERENCE (even in different wording)?
   No  → **missing**
   Yes → continue to question 2

2. Are the amount / limit / period / scope identical AND wording does not change meaning AND both documents use the same legal structure (both establish coverage as the default, or both establish non-coverage as the default)?
   Yes → **match**
   No  → **difference**

**Legal structure rule:** Two clauses covering the same topic are NOT a match if one establishes coverage as the default position (affirmative grant — the clause answers "what is covered") and the other establishes non-coverage as the default position (exclusion-with-exception — the clause answers "what is excluded, with a carve-out"). The insured's legal position — who bears the burden of proof in a dispute — is opposite under these two structures. This applies to any clause type in any policy domain. Always classify as `difference`.

**Iron Rule 1:** If REFERENCE covers the topic but with less / more / differently — use `difference`, never `missing`.
**Iron Rule 2:** A `difference` finding with empty `ref_quote` is an error. Either quote the reference line (even "נקבע ברשימה" / "ראה לוח הפוליסה"), or reclassify as `missing`.
**Iron Rule 3:** A `match` finding with empty `ref_quote` is a contradiction — reclassify as `missing` or `difference`.
**Schedule-deferral rule:** If INPUT states an explicit value/amount but REFERENCE says "to be set in schedule/appendix" (נקבע ברשימה / ראה לוח) — this is `difference`, NOT `match`. Quote the schedule-deferral line as `ref_quote`.
**Special case:** If INPUT explicitly excludes coverage ("X is not required") and REFERENCE includes it — this is `difference`, not `missing`.

---

## Severity Criteria — follow exactly

גבוהה  (high):   gap > 500,000 ₪ | coverage completely absent | different cancellation/subrogation/indemnity clause | INPUT explicitly requires an exclusion clause that is entirely absent from REFERENCE (applies to any exclusion type — cyber, pandemic, sanctions, war, nuclear, or any other — what matters is that INPUT mandates it and REFERENCE omits it)
בינונית (medium): gap 50,000–500,000 ₪ | wording change with interpretive impact | extension limited vs. required
נמוכה  (low):    gap < 50,000 ₪ | wording difference with no financial impact | match is always low

---

## Quantity Target

Report ALL meaningful findings — no minimum floor, no maximum cap.
Do NOT report titles, preambles, or purely definitional clauses.

---

## Field Format — every field mandatory

**input_quote** — always required, never empty:
❌ "40,000 ₪"
✅ "ביטוח אש מורחב — השתתפות עצמית, כל נזק אחר: 40,000 ₪ למקרה"
Rule: every reader must understand (a) which chapter, (b) which clause, (c) the exact value.
For table rows: include the row header + full value including %, minimum, maximum.

**ref_quote**:
- `missing` with no counterpart found → may be empty.
- `difference` → MANDATORY. Quote what you found in REFERENCE, even if it only says "נקבע ברשימה".
  If you have nothing to quote from REFERENCE → reclassify as `missing`.
- `match` → MANDATORY. A match without a ref_quote is invalid — reclassify as `missing`.
- Same format rule as input_quote: chapter + clause + exact value.

**ref_page** — required for every `difference` and `match`; empty only for `missing` with no quote.
**ref_file** — required for every `difference` and `match`; empty only for `missing` with no quote.

**diff_type**: שינוי סכום | הוספת כיסוי | צמצום כיסוי | שינוי ניסוח | כיסוי חסר | שינוי אחוז | אחר | התאמה

**confidence**:
high   — direct, exact quote from both documents; clear numerical comparison
medium — similar but not identical clause; page unclear; possible interpretation
low    — inferred from context; no direct quote found; missing that may be hidden

**reason**: up to 60 words — exactly what changed. For `difference`: "INPUT requires X, REFERENCE provides Y".
**severity**: per the criteria above — not by intuition.
**is_material**: false only for low + match. Every difference/missing at medium or above = true.

---

## Self-Check Before Output

Before returning JSON, verify:
□ Every chapter in INPUT is represented by at least one finding
□ No two findings cover the exact same clause
□ Every `difference`: input_quote + ref_quote + ref_page + ref_file all non-empty; if any empty → reclassify as missing
□ Every `match`: ref_quote non-empty; if empty → reclassify as missing or difference
□ Every `missing`: input_quote full; ref_quote empty or explained
□ Total findings ≤ 60; match findings ≤ 60
□ Every severity=גבוהה meets the defined criteria

---
""" + _JSON_STRING_ESCAPING_RULES + """
## Output — JSON only, no additional text

{
  "findings": [
    {
      "classification": "difference",
      "section": "פרק א' - פריצה ושוד",
      "input_quote": "פרק א' — פריצה ושוד: 1,000,000 ₪ לאירוע, 2,000,000 ₪ לתקופה",
      "input_page": "עמ' 11",
      "ref_quote": "הרחבות לפרק א' — פריצה ושוד, סעיף 4.ג: 500,000 ₪ לאירוע",
      "ref_page": "עמ' 9",
      "diff_type": "שינוי סכום",
      "severity": "גבוהה",
      "is_material": true,
      "confidence": "high",
      "reason": "INPUT requires 1M ₪ per event and 2M ₪ per period. REFERENCE provides 500K ₪ per event only.",
      "note": ""
    },
    {
      "classification": "match",
      "section": "תנאים כלליים - ויתור על תחלוף",
      "input_quote": "עמוד 13: הפוליסה כוללת סעיף ויתור על זכות התחלוף משוכרים ו/או דיירים של הנ\\"ל",
      "input_page": "עמ' 13",
      "ref_quote": "סעיף ויתור על תחלוף משוכרים ו/או דיירים של הנ\\"ל",
      "ref_page": "עמ' 12",
      "diff_type": "התאמה",
      "severity": "נמוכה",
      "is_material": false,
      "confidence": "high",
      "reason": "נוסח ויתור התחלוף תואם — כולל גרשיים מסומנים כהנ\\"ל.",
      "note": ""
    }
  ]
}

"""


SYSTEM_PROMPT_MISSING = """You are an expert Israeli insurance policy analyst.

## Task: Find MISSING items only

INPUT  = the tender/procurement document — the requirements the client MANDATES. Source of truth.
REFERENCE = the proposed insurance policy — being tested against INPUT requirements.

**Core rule — memorise this:**
A finding is `missing` if and only if INPUT states a requirement AND REFERENCE contains **no clause body** of any kind that addresses it — not a value, not a condition, not an exclusion-exception. The clause is simply not there.

Your ONLY task in this pass: scan every clause in INPUT and report those that are entirely absent from REFERENCE.

Every finding MUST map to explicit INPUT text (a real requirement — coverage, limit, condition, extension, or general term). Never report REFERENCE-only topics. Never report missing when REFERENCE contains any clause body that addresses the same substantive meaning. Do NOT search REFERENCE by section number — search by topic and context.

---

## What counts as "absent" vs "present" — CRITICAL

**ABSENT (→ report as missing):**
- REFERENCE has no text anywhere relating to this clause
- REFERENCE has a **chapter title or table-of-contents entry** for the topic but no clause body with a value, condition, or rule beneath it
  - Example: REFERENCE contains the heading "הרחבת שיקום קרקע" in a chapter list, but the clause body with amounts/conditions is missing → **this is missing**, not difference
- REFERENCE defers to a schedule ("נקבע ברשימה", "to be set", "see appendix") with no actual value anywhere
- REFERENCE mentions the topic only as a general word in a list (e.g. "כיסויים: אש, רעידת אדמה, שיקום קרקע") with no clause body specifying limits or conditions

**PRESENT (→ skip; difference pass owns it):**
- REFERENCE contains an actual clause body for this requirement — even with a different value, narrower scope, or alternate framing
- REFERENCE addresses this requirement inside an exclusion-with-exception clause (a clause that establishes non-coverage as the default and then carves out exceptions)
- REFERENCE has any specific value, limit, deductible, or condition for this exact clause

**The decisive question:** Does a clause *body* exist in REFERENCE for this requirement — not just a title, heading, or list entry?
→ YES → skip (difference pass)
→ NO  → report as missing

**DO NOT skip an item just because the general chapter appears in REFERENCE.** A chapter heading is not coverage.

---

**Exclusion-with-exception pattern — CRITICAL:**
If REFERENCE addresses a topic by stating that it is excluded from coverage and then carving out an exception for certain cases — the topic IS specifically addressed in REFERENCE. A carve-out from an exclusion is still a specific, actionable clause body. Do NOT classify as missing.
- Reclassify as `difference` if the legal structure, framing, wording, or conditions differ from INPUT. (When INPUT requires coverage affirmatively and REFERENCE provides it only as a carve-out from an exclusion, the legal structures differ — always `difference`, never `match`.)
- Reclassify as `match` only if both the legal structure AND the coverage content are fully equivalent.
Items skipped here due to exclusion-exception **must be caught by the difference pass**. Do not assume they are matches.

---

## Step A: Internal Mapping (not in output)

Build an internal checklist of every verifiable clause in INPUT:
- Every chapter (property, liability, equipment, mechanical breakdown, appendices…)
- Every amount / limit / deductible / period row in every table
- Every detailed extension with section number
- Every general condition with contractual impact (cancellation, subrogation, jurisdiction, indemnity)

Do NOT skip a chapter because it looks "standard" — even defaults need verification.

## Step B: Binary gate — apply to every item from Step A

**Gate: Does REFERENCE contain a specific, actionable clause body for this requirement?**
→ YES (any value, condition, exclusion-exception, or alternate framing in a clause body) → **SKIP**; difference pass owns framing gaps
→ NO  (completely absent, or only a chapter title / heading / list-word present with no clause body) → **report as missing**

Search REFERENCE by **meaning and topic** — not by matching section/chapter numbers. Different section numbers or chapter titles between documents are irrelevant — only whether the INPUT requirement's meaning exists in REFERENCE text.

---

## Quantity target — scan completeness (absent-only)

Your task is NOT to report fewer missing items — it is to miss NONE.

After building your Step A checklist, verify EVERY substantive INPUT requirement:
- Search REFERENCE by **meaning and topic** — not by matching section/chapter numbers
- If REFERENCE has **no clause body** anywhere that addresses this INPUT requirement → you MUST report it as missing
- If REFERENCE has **any** clause body on the same topic (value, condition, exclusion-exception, deferral line) → SKIP; difference pass owns it

**Do not stop early** after reporting the obvious high-severity gaps. Continue through all INPUT content:
- Every table row and amount/limit/deductible
- Every extension, appendix, and coverage type — by what it requires, not by its section label
- Every general condition (cancellation, subrogation, jurisdiction, indemnity)

**Hard rule — every finding must satisfy:**
- The requirement originates from explicit INPUT text (never from REFERENCE-only topics)
- The substantive meaning/context of that INPUT requirement is absent from REFERENCE (heading-only or list-word-only in REFERENCE still counts as absent)
- Your reason must state absence explicitly: "לא נמצא סעיף גוף", "נעדר לחלוטין", "אין כיסוי ל[נושא]", "מופיע בכותרת בלבד"

**Self-check before output:** For each INPUT requirement whose meaning is not addressed anywhere in REFERENCE — is it in your JSON? If not, add it now.

---

## Self-Check Before Writing Each Finding — CRITICAL

Before writing a `missing` finding, re-read your own `reason` text and apply ALL of the following tests:

1. **Specific value/page/clause test:** If your reason mentions a specific REFERENCE value, page number, or clause body (e.g. "REFERENCE provides 100,000 ₪", "REFERENCE page 5 states...") → **MUST NOT be missing**. Reclassify as `difference` or skip.

2. **Clause-body existence test:** If your reason states that REFERENCE contains an actual clause body addressing this topic — not just a chapter title or heading — (e.g. "REFERENCE has a clause that...", "REFERENCE covers this in section...", "מופיע בסעיף", "קיים כיסוי") → **MUST NOT be missing**. Reclassify as `difference`.
   - **Exception:** words like "מוזכר בכותרת" / "מופיע ברשימה בלבד" / "mentioned only in heading" confirm absence of a clause body → remains `missing`.

3. **Exclusion-exception test:** If REFERENCE contains the topic inside an exclusion-with-exception clause (a clause that establishes non-coverage as the default and then carves out the case INPUT requires) → **MUST NOT be missing**. Reclassify as `difference`.

4. **Wording-difference test:** If your reason contains any of the following phrases — "לא באותו נוסח" / "לא באותו הניסוח" / "בניסוח שונה" / "שונה בניסוחו" / "בנוסח שונה" / "different wording" / "different phrasing" / "not the same wording" — the topic IS present in REFERENCE, only phrased differently. **SKIP this item — do not report as missing.** The difference pass handles wording divergence.

**Only write `missing` if your reason explicitly states the clause body is completely absent** — e.g. "לא נמצא סעיף", "נעדר לחלוטין", "אין סעיף גוף", "not found", "completely silent", "מופיע בכותרת בלבד ללא ערכים".

5. **Sub-condition test:** If your reason states REFERENCE **has** a clause body (`"קיים סעיף"`, `"קיים כיסוי"`, `"מופיע בסעיף"`) but a **specific sub-condition** is absent (`"נעדר תנאי"`, `"ללא תנאי"`, `"ללא תנאי ראשוניות"`) → **MUST NOT be missing**. Skip; difference pass owns sub-condition gaps.

---

## Severity Criteria

גבוהה  (high):   clause completely absent from REFERENCE | missing cancellation/subrogation/indemnity/exclusion clause | no value provided where INPUT requires one
בינונית (medium): extension absent or only vaguely mentioned | condition deferred to schedule without a value | important condition absent with interpretive impact
נמוכה  (low):    minor clause absent with no financial impact

---

## Field Format
**input_quote** — always required, never empty. You MUST provide the ENTIRE relevant paragraph or complete section block from the document, not just the single line containing the value. Use the structured format:
  ציטוט:
  <הטקסט המלא, השלם והמדויק של כל הפסקה או הסעיף הרלוונטי מהמסמך — אין לקצר, לחתוך או להשמיט משפטים סמוכים השייכים לאותו סעיף/פסקה>
  הקשר:
  <מידע נוסף או הסבר קצר על מיקום הסעיף, תנאי סף נלווים או מבנה הטבלה במידת הצורך. השמט לחלוטין כשהציטוט עצמו מובן מאליו וכולל את כל ההקשר.>
**input_page** — required.
**ref_quote** — always empty string "" for missing findings.
**ref_page** — always empty string "".
**ref_file** — always empty string "".
**classification** — always "missing" for every finding in this pass.
**diff_type** — always "כיסוי חסר".
**confidence**:
  high   = thoroughly searched REFERENCE; certain no clause body exists for this requirement
  medium = clause body may exist under a different name, in a bundle file not shown, or in an appendix not visible
  low    = uncertain; the clause may be hidden or implicit
**reason**: עד 60 מילים, **בעברית בלבד** — מה דורש ה-INPUT ומדוע ה-REFERENCE אינו עונה על הדרישה (אין סעיף גוף, הפנייה לנספח, נעדר לחלוטין, מופיע בכותרת בלבד וכו').
**severity**: per criteria above — not by intuition.
**is_material**: true for medium/high severity; false for low.

---
""" + _JSON_STRING_ESCAPING_RULES + """
## Output rules — CRITICAL

- Output **only** the raw JSON object. No preamble, no explanation, no markdown, no prose before or after.
- Do NOT wrap in ```json fences.
- The very first character of your response must be `{` and the very last must be `}`.
- Do NOT add any text after the closing `}`.

## Output examples

**Example 1 — clause completely absent:**
{
  "findings": [
    {
      "classification": "missing",
      "section": "ביטוח אש מורחב - שיקום קרקע",
      "input_quote": "ציטוט:\nהרחבות פרק א' — שיקום קרקע: 1,500,000 ₪ לאירוע, 3,000,000 ₪ לתקופה",
      "input_page": "עמ' 11",
      "ref_quote": "",
      "ref_page": "",
      "diff_type": "כיסוי חסר",
      "severity": "גבוהה",
      "is_material": true,
      "confidence": "high",
      "reason": "ה-INPUT דורש כיסוי לשיקום קרקע בסך 1.5M ₪ לאירוע ו-3M ₪ לתקופה. לא נמצא סעיף גוף מקביל ב-REFERENCE — נעדר לחלוטין.",
      "note": ""
    }
  ]
}

**Example 2 — chapter title present in REFERENCE but clause body absent (still missing):**
{
  "findings": [
    {
      "classification": "missing",
      "section": "ביטוח אש מורחב - נזקי רעידת אדמה",
      "input_quote": "ציטוט:\nהרחבות פרק א' — רעידת אדמה: 5,000,000 ₪ לאירוע",
      "input_page": "עמ' 12",
      "ref_quote": "",
      "ref_page": "",
      "diff_type": "כיסוי חסר",
      "severity": "גבוהה",
      "is_material": true,
      "confidence": "high",
      "reason": "ה-INPUT דורש כיסוי רעידת אדמה בסך 5M ₪. 'רעידת אדמה' מופיע בכותרת פרק ב-REFERENCE בלבד, ללא סעיף גוף עם ערכים או תנאים — נעדר לחלוטין.",
      "note": ""
    }
  ]
}

**Example 3 — parent clause exists but sub-condition absent (NOT missing — skip entirely):**
INPUT requires sudden contamination coverage subject to third-party excess ("בתנאי כי אינו מכוסה בביטוח צד שלישי"). REFERENCE has a contamination clause but without that condition → **do NOT report as missing**. The difference pass owns this case. If you catch yourself writing a reason like "קיים סעיף … אך ללא תנאי …" — stop and skip.

"""

SYSTEM_PROMPT_DIFFERENCE = """You are an expert Israeli insurance policy analyst.

## Task: Find DIFFERENCE items only

INPUT  = the tender/procurement document — the requirements the client MANDATES. Source of truth.
REFERENCE = the proposed insurance policy — being tested against INPUT requirements.

Your ONLY task in this pass: scan INPUT and report clauses where REFERENCE addresses the same INPUT requirement but the **value, scope, conditions, or legal framing** differs — including non-numeric extensions (coverage inclusions, exclusions).

**The test is at the CLAUSE level, not the TOPIC level:**
- REFERENCE has a specific value/condition for this clause, and it differs from INPUT → **report as difference**
- REFERENCE says "נקבע ברשימה" / "to be set in schedule/appendix" while INPUT has a specific value → **report as difference** (the schedule-deferral line is your ref_quote)
- REFERENCE has no specific value for this clause at all → **skip** (missing pass handles it)
- REFERENCE matches INPUT exactly (same legal structure AND same content) → **skip** (match pass handles it)
  NOTE: identical subject matter with different legal structures is NOT a match — report as difference.

**ref_quote** is required for every difference finding. If the best available quote is approximate or paraphrased, include it and set confidence to "low". Do NOT skip a finding solely because the quote is imperfect.

---

## Step 0: Structural Framing Gate — MANDATORY, runs before all other checks

### 0a. Identical Terminology Warning — runs BEFORE Step 0b

**The more similar the terminology between INPUT and REFERENCE, the MORE carefully you must verify structural framing — not less.**
Shared vocabulary across opposite structures is the single most common missed-difference pattern in this task. High keyword overlap creates a false sense of equivalence that suppresses legitimate difference findings. When you notice strong terminological similarity, treat it as a trigger to slow down and apply the structural check with extra care — not as evidence of equivalence.

### 0b. Mandatory Structure Labeling

Before comparing any clause content, you MUST assign an explicit structural label to each document's treatment of the topic. You are not permitted to proceed to content comparison until both labels are assigned.

Assign one of these two labels to INPUT and separately to REFERENCE:

- **[affirmative-grant]** — The clause answers "what IS covered." Coverage is the default. In a dispute, the insurer bears the burden of proving an exception applies.
- **[exclusion-with-exception]** — The clause answers "what is NOT covered, with a carve-out." Non-coverage is the default. In a dispute, the insured bears the burden of proving the exception applies.

**If the two labels differ → this is ALWAYS a difference.**
→ Report as difference, `diff_type = "שינוי ניסוח"`, `severity = "בינונית"`.
The insured's legal position — who bears the burden of proof in a dispute — is opposite under these two structures. Identical subject matter and identical terminology do not make them legally equivalent. Do NOT allow keyword overlap or semantic similarity to override this rule.

### 0c. Linguistic Pattern Recognition for Structure Detection

Certain sentence-opening patterns are definitive signals of exclusion-with-exception structure, regardless of subject matter or how similar the content is to INPUT:

**Exclusion-with-exception signals (Hebrew):**
- "הפוליסה אינה מכסה... למעט..."
- "מוסכם ומוצהר כי אין כיסוי... אלא אם..."
- "הכיסוי אינו חל... למעט במקרה של..."
- "אינו מבוטח... פרט ל..."
- "לא יחריג מקרה ביטוח... בתנאי / אלא אם / ו/או..."
- "לא יפטר את המבטח... למעט..."
- Any opening that declares non-coverage and then carves back a subset
- Any clause that negates exclusion ("לא יחריג", "לא יפטר") then carves back coverage with trailing conditions

**Affirmative-grant signals (Hebrew):**
- "הביטוח חל על..."
- "מכוסה על פי פוליסה זו..."
- "הפוליסה מכסה..."
- "לא יושפע" / "לא ייפגע" / "הכיסוי ימשיך לחול" / "לא יושפע משינויים"
- Any opening that declares coverage or protection as the operative rule without carving back from an exclusion

**Framing distinction — do NOT treat both sides as [affirmative-grant] when keywords overlap:**
- INPUT: "לא יושפע משינויים..." (unconditional protection) → [affirmative-grant]
- REFERENCE: "לא יחריג מקרה ביטוח בשל..." with trailing "בתנאי שנעשה..." / "אלא אם" → [exclusion-with-exception]
High keyword overlap between such pairs is a warning to apply this check — not evidence of equivalence.

If REFERENCE opens with an exclusion-with-exception pattern and INPUT does not → labels differ → report as difference, `diff_type = "שינוי ניסוח"`, `severity = "בינונית"`.
If INPUT opens with an exclusion-with-exception pattern and REFERENCE does not → labels differ → report as difference, `diff_type = "שינוי ניסוח"`, `severity = "בינונית"`.

---

## Step 0.5: Counterpart Clause Identity Gate — MANDATORY

Before searching REFERENCE or comparing content, verify you have the **same sub-clause** — not merely the same chapter or shared keywords.

### 0.5a. Extract INPUT clause identity

For each INPUT item, identify before any REFERENCE search:
- **Named heading** (e.g. "אחריות צולבת", "ביטול הביטוח", "אמצעים להקלת הסיכון", "ביטוח ראשוני")
- **Section number** if present (e.g. 13.1, 4.2, 13.5)
- **Clause function**: definition / extension-amendment / condition / cancellation / primary-insurance / payment-priority / subrogation-waiver / claims-handling / risk-mitigation

### 0.5b. REFERENCE search order (strict priority)

1. Clause with **same or equivalent heading/name** in the same insurance chapter
2. Clause with **same section number + same topic name** (even if page differs)
3. Semantic match — **only if** steps 1–2 found nothing
4. If only a related-but-different clause exists → **skip** (do not emit difference; missing pass may handle absence)

### 0.5c. Extension / tender-amendment rule

- When INPUT is a numbered amendment ("בסוף הסעיף...", "יתווספו המילים...", "נוסח הסעיף הוחלף...") → the REFERENCE counterpart is the **base clause with that exact name/number**, not a different extension or related base-policy paragraph
- When INPUT names a specific extension (אחריות צולבת, אחריות שילוחית, טיפול בתביעות) → compare only to REFERENCE text under **that same extension heading**, never to a broader "הרחבות" or general liability paragraph

### 0.5d. Keyword-bridge trap — CRITICAL

When INPUT and REFERENCE share words (קדימות, ביטול, אחריות, עובדים…) but address **different legal questions**, do NOT emit a difference — even if both appear under the same chapter heading (e.g. "קדימות הביטוח").

Before comparing content, ask: "Does REFERENCE answer the **same question** INPUT asks?" If no → skip.

Do NOT infer a gap by importing conditions from an unrelated REFERENCE clause (e.g. do not treat a שעבוד exception from סעיף 4 קדימות as a "סייג" on a ביטוח ראשוני requirement).

### 0.5e. Clause-function mismatch = skip

These pairs are **never** valid counterparts — skip even when keywords overlap:
- **ביטוח ראשוני / primary insurance / no contribution** ≠ **סעיף קדימות / payment priority among claimants** (INPUT: who pays first among insurers; REFERENCE: order among insured parties entitled to indemnity)
- **הגדרת עובד חברה** ≠ **ויתור על זכות התחלוף**
- **אמצעים להקלת הסיכון** ≠ **אי קיום תנאי הפוליסות**
- **ביטול הביטוח** (60/10 days notice) ≠ **הודעה לג'וינט ישראל** on policy change/cancellation
- **הרחבת CROSS LIABILITY / אחריות צולבת** ≠ **סעיף גופים ממשלתיים**
- **הרחבת אחריות שילוחית** ≠ **סעיף מייצג/פועל מטעמו** (unless that IS the named extension body)
- **טיפול בתביעות (הרחבה)** ≠ **נוהל תביעות כללי** without the same named extension heading

---

## Step A: Internal Mapping (not in output)

Build an internal checklist of every verifiable clause in INPUT:
- Every chapter (property, liability, equipment, mechanical breakdown, appendices…)
- Every amount / limit / deductible / period row in every table
- Every detailed extension with section number
- Every general condition with contractual impact (cancellation, subrogation, jurisdiction, indemnity)
- For every spec chapter with numbered sub-clauses (2.1, 2.2, …): enumerate **every numbered item** — not only rows with ₪ amounts. Each sub-clause must produce a finding in this pass, the missing pass, or the match pass.

## Step B: For each item — find and compare at clause level

**Run Step 0.5 first.** Only if a valid counterpart is confirmed, compare values/conditions/framing.

Ask yourself: "Does REFERENCE contain a specific, measurable value for this exact clause?"
- Yes, and it differs from INPUT → report as difference; use the best available ref_quote (exact or paraphrased)
- Yes, and it matches INPUT (same legal structure AND same content) → skip (match pass handles it)
  NOTE: identical subject matter is NOT a match when legal structures differ — see Step 0.
- No specific value anywhere in REFERENCE → skip (missing pass handles it)
- REFERENCE has the same clause but in different wording, framing, or conditions (even with no numerical change) → **report as difference**, `diff_type = "שינוי ניסוח"`, quote the REFERENCE phrasing as ref_quote
- INPUT and REFERENCE address the same topic but with **different legal structures**: one establishes coverage as the default (affirmative grant) and the other establishes non-coverage as the default (exclusion-with-exception) → **always report as difference**, `diff_type = "שינוי ניסוח"`, `severity = "בינונית"`. This rule applies to any clause type in any policy domain. (Step 0 catches this first — this bullet is a reminder only.)
- **INPUT and REFERENCE have same section title but have different details about it**
- REFERENCE contains the **parent clause** for a topic but is **missing a sub-condition** that INPUT requires (e.g. third-party primacy/excess, schedule reference, time limit, deductible qualifier) → **report as difference**, `diff_type = "שינוי ניסוח"` or `"צמצום כיסוי"`, quote the REFERENCE parent clause as `ref_quote`

**Schedule-deferral rule:** If INPUT states an explicit value but REFERENCE says "נקבע ברשימה" / "to be set in schedule/appendix" → this IS a difference. Quote the schedule-deferral line as ref_quote.

---

## Direction Rules — CRITICAL

Before assigning `diff_type`, determine the **direction** from INPUT's perspective:

**REFERENCE is worse for the insured (reduction):**
- REFERENCE coverage limit < INPUT limit → `diff_type = "צמצום כיסוי"`
- REFERENCE deductible > INPUT deductible → `diff_type = "צמצום כיסוי"` (higher deductible = more out-of-pocket)
- REFERENCE narrows scope, adds conditions, or excludes what INPUT includes
- REFERENCE adds a restrictive qualifier or caveat to a clause that INPUT states unconditionally (e.g., INPUT: "הפוליסה קודמת לכל פוליסה אחרת"; REFERENCE: "הפוליסה קודמת לכל פוליסה אחרת **אלא אם קיים סעיף שעבוד**") → `diff_type = "שינוי ניסוח"`, severity based on contractual impact; ref_quote MUST include the full clause with the caveat. (Example illustrates the PATTERN — the same rule applies to any clause type in any insurance domain.)
- INPUT enumerates multiple specific items and REFERENCE covers only a subset (e.g., INPUT requires: "[כיסוי א', כיסוי ב', כיסוי ג', כיסוי ד']"; REFERENCE covers only "[כיסוי א', כיסוי ב']") → `diff_type = "צמצום כיסוי"`, quote REFERENCE's limited list as ref_quote, severity = "בינונית" or "גבוהה". (Example illustrates the PATTERN — applies to any enumeration of coverage types, parties, or scenarios.)

**Exact amount rule:** Do NOT skip or classify as match when both documents "extend" or "increase" a limit but the **numbers differ** (e.g. INPUT 1,000,000 ₪ vs REFERENCE 400,000 ₪). Always report `שינוי סכום` with the exact gap.

**REFERENCE is better for the insured (addition):**
- REFERENCE coverage limit > INPUT limit → `diff_type = "הוספת כיסוי"`, `severity = "נמוכה"`, `is_material = false`
- REFERENCE deductible < INPUT deductible → `diff_type = "הוספת כיסוי"`, `severity = "נמוכה"`, `is_material = false` (lower deductible = less out-of-pocket)
- REFERENCE broadens scope or adds coverage beyond INPUT

**When REFERENCE is better for the insured → severity is always `נמוכה` and `is_material = false`.**

**No-Absent-Input rule — CRITICAL:**
A `difference` finding is ONLY valid when INPUT contains an actual, explicit clause about this topic.
If the only honest description of INPUT's position is "לא מצוין", "לא מוזכר", "לא מוגדר", "לא נזכר",
or any equivalent phrasing meaning "not specified / not mentioned" — INPUT has NO requirement here.
→ This is a REFERENCE-only clause. Do NOT generate a finding. Ignore it completely.
→ Test: before writing a finding, ask yourself "Can I quote real, explicit text from INPUT for this clause?"
   If the answer is No → skip. Do not fabricate an input_quote describing absence.

---

## Severity Criteria

גבוהה  (high):   gap > 500,000 ₪ (INPUT requires more than REFERENCE provides) | different cancellation/subrogation/indemnity/exclusion clause
בינונית (medium): gap 50,000–500,000 ₪ | wording change with interpretive impact | extension limited vs. required | structural framing difference (affirmative-grant vs. exclusion-with-exception)
נמוכה  (low):    gap < 50,000 ₪ | wording difference with no financial impact | REFERENCE is better than INPUT (addition)

---

## Self-Contradiction Check — CRITICAL

**Self-Contradiction Test — apply before writing every finding:**
Read your `reason` draft and ask: "Does my reason actually describe a gap between INPUT and REFERENCE?"
If your reason describes the two documents as equivalent, identical, or states that any difference is purely cosmetic (section numbering only, different document structure with no substantive change) — you have NO finding. Do not output it. Skip entirely.

This test is absolute: if your own analysis concludes there is no material difference, outputting a `difference` finding is always wrong — regardless of which pass you are in or how many findings you have produced so far.

**IMPORTANT — Self-Contradiction Check does NOT override Step 0:**
If Step 0 identified a structural framing difference (labels differ), the finding MUST be reported even if the subject matter appears semantically equivalent. Structural framing differences are never cosmetic. Do not use the Self-Contradiction Check to suppress a finding that Step 0 mandated.

Before writing each finding:
- If your reason says the values/wording are **identical, equivalent, or "no material difference"** AND Step 0 confirmed labels match → **do NOT report as difference**. It belongs in the match pass. Skip it entirely.
- If your reason describes REFERENCE as **better** than INPUT (higher limit, lower deductible, broader scope) → set `diff_type = "הוספת כיסוי"`, `severity = "נמוכה"`, `is_material = false`.

---

## Pre-Output Purge — run before returning JSON

For every finding already in your output list, answer these questions in order:

1. "Did I assign explicit structure labels ([affirmative-grant] or [exclusion-with-exception]) to both INPUT and REFERENCE before comparing content?"
   - If no → re-run Step 0 now. If labels differ → confirm the finding is reported. If labels match → verify through Step B.

2. "Did strong keyword or terminology overlap between INPUT and REFERENCE cause me to treat the clauses as equivalent without completing the structural check?"
   - If yes → re-run Step 0 with fresh eyes. If labels differ → the finding must be reported.

3. "Does my reason text describe the two clauses as semantically equivalent despite a structural framing difference?"
   - If yes → this is an error. The finding must be reported. Revise the reason to explain the structural difference.

4. "Did I run Step 0.5? Does the REFERENCE heading/function match the INPUT clause identity, not just shared keywords?"
   - If no → remove the finding from output.

5. "If INPUT cites section number X, did I compare to REFERENCE section X (or its named equivalent), not a different section that mentions similar words?"
   - If no → remove the finding from output.

6. "Does my reason import a condition from a REFERENCE clause that is not the INPUT clause's counterpart?"
   - If yes → remove the finding (keyword-bridge trap).

---

## Field Format

**input_quote** — required. Use the structured format:
  ציטוט:
  <הטקסט המדויק מהמסמך: "[פרק] — [סעיף]: [ערך מלא כולל %, מינימום, מקסימום]">
  הקשר:
  <1–3 משפטים קצרים — רק כאשר נדרש: למשל כשהסעיף כולל מספר אופציות (א'/ב'/ג'), ציין את כל האופציות. השמט לחלוטין כשהציטוט מובן מאליו.>
  **If the only text you can supply here is a statement that INPUT does NOT mention this item
  (e.g. "לא מצוין ...", "לא מוגדר ...", "not specified") → skip the entire finding.**
**input_page** — required.
**ref_quote** — MANDATORY. Use the same structured format as input_quote.
  ציטוט:
  <הטקסט המדויק מה-REFERENCE>
  הקשר:
  <1–3 משפטים — רק כשנדרש הקשר נוסף. השמט כשהציטוט מובן מאליו.>
  If you have nothing to quote from REFERENCE → do NOT include this finding (skip it entirely).
  **INVALID ref_quote (do NOT emit — skip or let missing pass handle):** "לא נמצאה התייחסות...", "לא צוין סעיף...", or quoting INPUT section numbers ("של INPUT", "סעיף X של INPUT"). These describe absence, not a REFERENCE clause.
**ref_page** — required.
**ref_file** — required.
**classification** — always "difference".
**diff_type**: צמצום כיסוי | הוספת כיסוי | שינוי סכום | שינוי ניסוח | שינוי אחוז | אחר
**confidence**: high = exact quotes from both, clear numerical comparison; medium = similar clause but not identical; low = inferred
**reason**: עד 60 מילים, **בעברית בלבד** — must confirm counterpart identity first: "ה-INPUT דורש [שם הסעיף]. ה-REFERENCE מכיל סעיף מקביל '[שם]' — ..." then describe the gap and direction [גרוע/טוב יותר למבוטח]. If counterpart names or functions differ materially, do not emit the finding. When the difference is structural, state both structure labels ([affirmative-grant] / [exclusion-with-exception]) and the burden-of-proof impact.
**severity**: per criteria above — not by intuition.
**is_material**: true for medium/high severity; false for low.

---
""" + _JSON_STRING_ESCAPING_RULES + """
## Output rules — CRITICAL

- Output **only** the raw JSON object. No preamble, no explanation, no markdown, no prose before or after.
- Do NOT wrap in ```json fences.
- The very first character of your response must be `{` and the very last must be `}`.
- Do NOT add any text after the closing `}`.

## Output example

{
  "findings": [
    {
      "classification": "difference",
      "section": "ביטוח אחריות צד שלישי - נזק לרכוש בחזקה",
      "input_quote": "ציטוט:\nסעיף 2.3 — נזק לרכוש שהמבוטח בעל זיקה אליו: 4,000,000 ₪ למקרה ולתקופה",
      "input_page": "עמ' 44",
      "ref_quote": "ציטוט:\nסעיף ה(1) — נזק לרכוש שהמבוטח בעל זיקה אליו: 2,000,000 ₪ למקרה ולתקופה",
      "ref_page": "עמ' 1 צד ג",
      "diff_type": "שינוי סכום",
      "severity": "גבוהה",
      "is_material": true,
      "confidence": "high",
      "reason": "ה-INPUT דורש 4,000,000 ₪. ה-REFERENCE מספק 2,000,000 ₪ — פער של 2M ₪, גרוע יותר למבוטח.",
      "note": ""
    },
    {
      "classification": "difference",
      "section": "תנאים כלליים - ביטול הביטוח",
      "input_quote": "ציטוט:\nסעיף 11 — ביטול הביטוח: נוסח הסעיף הוחלף... יתווספו המילים 'על אף האמור לעיל... לפי עילות הביטול המפורטות בחוק חוזה הביטוח, התשמ\\"א-1981.'",
      "input_page": "עמ' 8",
      "ref_quote": "ציטוט:\nסעיף ביטול — ללא התוספת מחוק תשמ\\"א",
      "ref_page": "עמ' 5",
      "diff_type": "שינוי ניסוח",
      "severity": "בינונית",
      "is_material": true,
      "confidence": "high",
      "reason": "ה-INPUT דורש התוספת לפי חוק תשמ\\"א. ה-REFERENCE חסרה.",
      "note": ""
    }
  ]
}

"""
# SYSTEM_PROMPT_DIFFERENCE = """You are an expert Israeli insurance policy analyst.

# ## Task: Find DIFFERENCE items only

# INPUT  = the tender/procurement document — the requirements the client MANDATES. Source of truth.
# REFERENCE = the proposed insurance policy — being tested against INPUT requirements.

# Your ONLY task in this pass: scan INPUT and report clauses where REFERENCE addresses the same INPUT requirement but the **value, scope, conditions, or legal framing** differs — including non-numeric extensions (coverage inclusions, exclusions).

# **The test is at the CLAUSE level, not the TOPIC level:**
# - REFERENCE has a specific value/condition for this clause, and it differs from INPUT → **report as difference**
# - REFERENCE says "נקבע ברשימה" / "to be set in schedule/appendix" while INPUT has a specific value → **report as difference** (the schedule-deferral line is your ref_quote)
# - REFERENCE has no specific value for this clause at all → **skip** (missing pass handles it)
# - REFERENCE matches INPUT exactly (same legal structure AND same content) → **skip** (match pass handles it)
#   NOTE: identical subject matter with different legal structures is NOT a match — report as difference.

# **ref_quote** is required for every difference finding. If the best available quote is approximate or paraphrased, include it and set confidence to "low". Do NOT skip a finding solely because the quote is imperfect.

# ---

# ## Step A: Internal Mapping (not in output)

# Build an internal checklist of every verifiable clause in INPUT:
# - Every chapter (property, liability, equipment, mechanical breakdown, appendices…)
# - Every amount / limit / deductible / period row in every table
# - Every detailed extension with section number
# - Every general condition with contractual impact (cancellation, subrogation, jurisdiction, indemnity)
# - For every spec chapter with numbered sub-clauses (2.1, 2.2, …): enumerate **every numbered item** — not only rows with ₪ amounts. Each sub-clause must produce a finding in this pass, the missing pass, or the match pass.

# ## Step B: For each item — find and compare at clause level

# Ask yourself: "Does REFERENCE contain a specific, measurable value for this exact clause?"
# - Yes, and it differs from INPUT → report as difference; use the best available ref_quote (exact or paraphrased)
# - Yes, and it matches INPUT (same legal structure AND same content) → skip (match pass handles it)
#   NOTE: identical subject matter is NOT a match when legal structures differ — see bullet below.
# - No specific value anywhere in REFERENCE → skip (missing pass handles it)
# - REFERENCE has the same clause but in different wording, framing, or conditions (even with no numerical change) → **report as difference**, `diff_type = "שינוי ניסוח"`, quote the REFERENCE phrasing as ref_quote
# - INPUT and REFERENCE address the same topic but with **different legal structures**: one establishes coverage as the default (affirmative grant — "what is covered") and the other establishes non-coverage as the default (exclusion-with-exception — "what is excluded, with a carve-out") → **always report as difference**, `diff_type = "שינוי ניסוח"`, `severity = "בינונית"`. The insured's legal position — who bears the burden of proof in a dispute — is opposite under these two structures. Do NOT skip this as semantically equivalent. This rule applies to any clause type in any policy domain.
# -  **INPUT and REFERENCE has same section title but have diffrent details about it**
# **Schedule-deferral rule:** If INPUT states an explicit value but REFERENCE says "נקבע ברשימה" / "to be set in schedule/appendix" → this IS a difference. Quote the schedule-deferral line as ref_quote.

# ---

# ## Direction Rules — CRITICAL

# Before assigning `diff_type`, determine the **direction** from INPUT's perspective:

# **REFERENCE is worse for the insured (reduction):**
# - REFERENCE coverage limit < INPUT limit → `diff_type = "צמצום כיסוי"`
# - REFERENCE deductible > INPUT deductible → `diff_type = "צמצום כיסוי"` (higher deductible = more out-of-pocket)
# - REFERENCE narrows scope, adds conditions, or excludes what INPUT includes
# - REFERENCE adds a restrictive qualifier or caveat to a clause that INPUT states unconditionally (e.g., INPUT: "הפוליסה קודמת לכל פוליסה אחרת"; REFERENCE: "הפוליסה קודמת לכל פוליסה אחרת **אלא אם קיים סעיף שעבוד**") → `diff_type = "שינוי ניסוח"`, severity based on contractual impact; ref_quote MUST include the full clause with the caveat. (Example illustrates the PATTERN — the same rule applies to any clause type in any insurance domain.)
# - INPUT enumerates multiple specific items and REFERENCE covers only a subset (e.g., INPUT requires: "[כיסוי א', כיסוי ב', כיסוי ג', כיסוי ד']"; REFERENCE covers only "[כיסוי א', כיסוי ב']") → `diff_type = "צמצום כיסוי"`, quote REFERENCE's limited list as ref_quote, severity = "בינונית" or "גבוהה". (Example illustrates the PATTERN — applies to any enumeration of coverage types, parties, or scenarios.)

# **Exact amount rule:** Do NOT skip or classify as match when both documents "extend" or "increase" a limit but the **numbers differ** (e.g. INPUT 1,000,000 ₪ vs REFERENCE 400,000 ₪). Always report `שינוי סכום` with the exact gap.

# **REFERENCE is better for the insured (addition):**
# - REFERENCE coverage limit > INPUT limit → `diff_type = "הוספת כיסוי"`, `severity = "נמוכה"`, `is_material = false`
# - REFERENCE deductible < INPUT deductible → `diff_type = "הוספת כיסוי"`, `severity = "נמוכה"`, `is_material = false` (lower deductible = less out-of-pocket)
# - REFERENCE broadens scope or adds coverage beyond INPUT

# **When REFERENCE is better for the insured → severity is always `נמוכה` and `is_material = false`.**

# **No-Absent-Input rule — CRITICAL:**
# A `difference` finding is ONLY valid when INPUT contains an actual, explicit clause about this topic.
# If the only honest description of INPUT's position is "לא מצוין", "לא מוזכר", "לא מוגדר", "לא נזכר",
# or any equivalent phrasing meaning "not specified / not mentioned" — INPUT has NO requirement here.
# → This is a REFERENCE-only clause. Do NOT generate a finding. Ignore it completely.
# → Test: before writing a finding, ask yourself "Can I quote real, explicit text from INPUT for this clause?"
#    If the answer is No → skip. Do not fabricate an input_quote describing absence.

# ---

# ## Severity Criteria

# גבוהה  (high):   gap > 500,000 ₪ (INPUT requires more than REFERENCE provides) | different cancellation/subrogation/indemnity/exclusion clause
# בינונית (medium): gap 50,000–500,000 ₪ | wording change with interpretive impact | extension limited vs. required
# נמוכה  (low):    gap < 50,000 ₪ | wording difference with no financial impact | REFERENCE is better than INPUT (addition)

# ---

# ## Self-Contradiction Check — CRITICAL

# **Self-Contradiction Test — apply before writing every finding:**
# Read your `reason` draft and ask: "Does my reason actually describe a gap between INPUT and REFERENCE?"
# If your reason describes the two documents as equivalent, identical, or states that any difference is purely cosmetic (section numbering only, different document structure with no substantive change) — you have NO finding. Do not output it. Skip entirely.

# This test is absolute: if your own analysis concludes there is no material difference, outputting a `difference` finding is always wrong — regardless of which pass you are in or how many findings you have produced so far.

# Before writing each finding:
# - If your reason says the values/wording are **identical, equivalent, or "no material difference"** → **do NOT report this as difference**. It belongs in the match pass. Skip it entirely.
# - If your reason describes REFERENCE as **better** than INPUT (higher limit, lower deductible, broader scope) → set `diff_type = "הוספת כיסוי"`, `severity = "נמוכה"`, `is_material = false`.

# ---

# ## Field Format

# **input_quote** — required. Use the structured format:
#   ציטוט:
#   <הטקסט המדויק מהמסמך: "[פרק] — [סעיף]: [ערך מלא כולל %, מינימום, מקסימום]">
#   הקשר:
#   <1–3 משפטים קצרים — רק כאשר נדרש: למשל כשהסעיף כולל מספר אופציות (א'/ב'/ג'), ציין את כל האופציות. השמט לחלוטין כשהציטוט מובן מאליו.>
#   **If the only text you can supply here is a statement that INPUT does NOT mention this item
#   (e.g. "לא מצוין ...", "לא מוגדר ...", "not specified") → skip the entire finding.**
# **input_page** — required.
# **ref_quote** — MANDATORY. Use the same structured format as input_quote.
#   ציטוט:
#   <הטקסט המדויק מה-REFERENCE>
#   הקשר:
#   <1–3 משפטים — רק כשנדרש הקשר נוסף. השמט כשהציטוט מובן מאליו.>
#   If you have nothing to quote from REFERENCE → do NOT include this finding (skip it entirely).
# **ref_page** — required.
# **ref_file** — required.
# **classification** — always "difference".
# **diff_type**: צמצום כיסוי | הוספת כיסוי | שינוי סכום | שינוי ניסוח | שינוי אחוז | אחר
# **confidence**: high = exact quotes from both, clear numerical comparison; medium = similar clause but not identical; low = inferred
# **reason**: עד 60 מילים, **בעברית בלבד** — "ה-INPUT דורש X, ה-REFERENCE מספק Y — כיוון: [גרוע/טוב יותר למבוטח]".
# **severity**: per criteria above — not by intuition.
# **is_material**: true for medium/high severity; false for low.

# ---

# ## Output rules — CRITICAL

# - Output **only** the raw JSON object. No preamble, no explanation, no markdown, no prose before or after.
# - Do NOT wrap in ```json fences.
# - The very first character of your response must be `{` and the very last must be `}`.
# - Do NOT add any text after the closing `}`.

# ## Output example

# {
#   "findings": [
#     {
#       "classification": "difference",
#       "section": "ביטוח אחריות צד שלישי - נזק לרכוש בחזקה",
#       "input_quote": "ציטוט:\nסעיף 2.3 — נזק לרכוש שהמבוטח בעל זיקה אליו: 4,000,000 ₪ למקרה ולתקופה",
#       "input_page": "עמ' 44",
#       "ref_quote": "ציטוט:\nסעיף ה(1) — נזק לרכוש שהמבוטח בעל זיקה אליו: 2,000,000 ₪ למקרה ולתקופה",
#       "ref_page": "עמ' 1 צד ג",
#       "diff_type": "שינוי סכום",
#       "severity": "גבוהה",
#       "is_material": true,
#       "confidence": "high",
#       "reason": "ה-INPUT דורש 4,000,000 ₪. ה-REFERENCE מספק 2,000,000 ₪ — פער של 2M ₪, גרוע יותר למבוטח.",
#       "note": ""
#     }
#   ]
# }

# """
SYSTEM_PROMPT_MATCH = """You are an expert Israeli insurance policy analyst.

## Task: Find MATCH items only

INPUT  = the tender/procurement document — the requirements the client MANDATES. Source of truth.
REFERENCE = the proposed insurance policy — being tested against INPUT requirements.

Your ONLY task in this pass: report INPUT requirements where REFERENCE **fully satisfies** the requirement — same scope, same meaning, same legal structure, even if wording differs slightly.

Report EVERY INPUT requirement (table row, extension, general condition) where REFERENCE fully satisfies the same substantive meaning — even if:
- Section/chapter numbers differ (INPUT סעיף 2.3 ↔ REFERENCE סעיף 7 is valid)
- Wording or document structure differs
- The clause appears under a different heading or chapter title

Match criteria (all must hold):
- Same legal structure on BOTH sides ([affirmative-grant] or [exclusion-with-exception])
- Same substantive scope and meaning — REFERENCE covers ALL items/parties/scenarios INPUT requires
- Same amounts, limits, deductibles, and conditions (formatting may differ; values must not)

Do NOT require REFERENCE to mirror INPUT section numbering. Locate the counterpart by **topic and meaning**, not by label.

Keep ALL existing disqualification rules (Step 0, rules 1–4, Pre-Output Purge). If any disqualifier applies → skip (difference pass owns it).
Do not stop early — scan every INPUT requirement, but report a match **only** when all disqualifiers pass.

**Section-number rule:** Matching is semantic, not positional. Different section numbers, chapter names, or document order between INPUT and REFERENCE do NOT prevent a match when the underlying requirement and coverage are equivalent.

**CRITICAL: ref_quote is MANDATORY for every finding in this pass.** If you cannot quote the REFERENCE text → skip the item.

---

## Quantity target

Build a checklist of every verifiable INPUT requirement (amounts, limits, conditions, extensions, general terms). For each item, search REFERENCE by **meaning and topic** — not by section number. Report a match with mandatory ref_quote **ONLY when all disqualifiers pass** (Step 0, rules 1–4). When in doubt after the structural check → skip (do not guess match). Do not limit yourself to "key structural" clauses only.

---

## Step 0: Structural Framing Gate — MANDATORY, runs before all other checks

### 0a. Identical Terminology Warning — runs BEFORE Step 0b

**The more similar the terminology between INPUT and REFERENCE, the MORE carefully you must verify structural framing — not less.**
Shared vocabulary across opposite structures is the single most common false-positive pattern in this task. High keyword overlap is a warning signal, not a confidence signal. When you notice strong terminological similarity, treat it as a trigger to slow down and apply the structural check with extra care — not as evidence of a match.

### 0b. Mandatory Structure Labeling

Before comparing any clause content, you MUST assign an explicit structural label to each document's treatment of the topic. You are not permitted to proceed to content comparison until both labels are assigned.

Assign one of these two labels to INPUT and separately to REFERENCE:

- **[affirmative-grant]** — The clause answers "what IS covered." Coverage is the default. In a dispute, the insurer bears the burden of proving an exception applies.
- **[exclusion-with-exception]** — The clause answers "what is NOT covered, with a carve-out." Non-coverage is the default. In a dispute, the insured bears the burden of proving the exception applies.

**If the two labels differ → STOP. This is NOT a match. Do not proceed to content comparison.**

The insured's legal position is opposite under these two structures. Coverage as a default (affirmative grant) is fundamentally different from coverage as a residual exception to an exclusion. Identical subject matter and identical terminology do not make them legally equivalent.
→ Skip; the difference pass handles it as שינוי ניסוח.

### 0c. Linguistic Pattern Recognition for Structure Detection

Certain sentence-opening patterns are definitive signals of exclusion-with-exception structure, regardless of subject matter or how similar the content is to INPUT:

**Exclusion-with-exception signals (Hebrew):**
- "הפוליסה אינה מכסה... למעט..."
- "מוסכם ומוצהר כי אין כיסוי... אלא אם..."
- "הכיסוי אינו חל... למעט במקרה של..."
- "אינו מבוטח... פרט ל..."
- "לא יחריג מקרה ביטוח... בתנאי / אלא אם / ו/או..."
- "לא יפטר את המבטח... למעט..."
- Any opening that declares non-coverage and then carves back a subset
- Any clause that negates exclusion ("לא יחריג", "לא יפטר") then carves back coverage with trailing conditions

**Affirmative-grant signals (Hebrew):**
- "הביטוח חל על..."
- "מכוסה על פי פוליסה זו..."
- "הפוליסה מכסה..."
- "לא יושפע" / "לא ייפגע" / "הכיסוי ימשיך לחול" / "לא יושפע משינויים"
- Any opening that declares coverage or protection as the operative rule without carving back from an exclusion

**Framing distinction — do NOT treat both sides as [affirmative-grant] when keywords overlap:**
- INPUT: "לא יושפע משינויים..." (unconditional protection) → [affirmative-grant]
- REFERENCE: "לא יחריג מקרה ביטוח בשל..." with trailing "בתנאי שנעשה..." / "אלא אם" → [exclusion-with-exception]
High keyword overlap between such pairs is a warning to apply this check — not evidence of equivalence.

If REFERENCE opens with an exclusion-with-exception pattern and INPUT does not → labels differ → NOT a match.
If INPUT opens with an exclusion-with-exception pattern and REFERENCE does not → labels differ → NOT a match.

---

## Four conditions that DISQUALIFY a match — check ALL before reporting

**1. REFERENCE adds ANY restriction, condition, or caveat absent from INPUT:**
Ask yourself: "Does REFERENCE impose any qualifying condition, restriction, limitation, or caveat on this clause that is ABSENT from INPUT?"
Do NOT rely on specific keywords — read the meaning, not the surface form.
Signs of a restriction in REFERENCE that INPUT lacks:
- Any inline condition ("in the event that...", "provided that...", "subject to...", "only when...")
- Any numbered sub-clauses (1.1, 1.2, א, ב, etc.) beneath the REFERENCE clause that do not appear in INPUT
- Any trailing qualification that limits when, how, or to whom the clause applies
If ANY such restriction is present in REFERENCE but absent from INPUT → NOT a match.
The added condition restricts the insured's rights regardless of how minor it seems.
→ Skip; the difference pass handles it as `שינוי ניסוח`.

Example of the PATTERN (applies to any clause type, not only these):
INPUT states a rule unconditionally: "Clause [A] applies."
REFERENCE states the same rule with a condition: "Clause [A] applies, PROVIDED THAT [condition B]."
→ NOT a match — REFERENCE restricts what INPUT grants unconditionally.

Example of the PATTERN (insurance protection clauses):
INPUT: unconditional protection — "לא יושפע משינויים..." / "הכיסוי לא יושפע מתיאור בלתי מדויק..."
REFERENCE: same topic + added qualifier — "בתנאי שנעשה בתום לב", "בתנאי שלא השפיע לרעה על חשיפת המבטח לנזק", "ואינו נחשב לעניין מהותי"
→ NOT a match — the added qualifier restricts the insured's protection even when the subject matter overlaps heavily.

**2. INPUT lists more items/scope than REFERENCE covers:**
If INPUT's clause applies to a broader set (more items, more parties, more scenarios, more coverage types) and REFERENCE only addresses a subset — even if the overlap is substantial — this is NOT a full match. Report a match ONLY when REFERENCE explicitly covers EVERYTHING INPUT requires.

Example of the PATTERN:
INPUT: "[clause] applies to [items A, B, C, D, E]"
REFERENCE: "[clause] applies to [items A, B]"
→ NOT a match. Items C, D, E are uncovered.
(The actual item names will differ per document — the pattern is what matters.)

**3. ref_quote completeness — read to the end of the clause:**
Your ref_quote MUST include the COMPLETE REFERENCE text for the clause — including any trailing conditions, qualifications, or restrictive caveats. Never quote only the shared/common portion. If reading the full clause reveals any difference from INPUT → skip this item and let the difference pass handle it.

**4. Cross-policy domain disqualification:**
If the only REFERENCE clause found is in a different policy domain from the INPUT requirement → NOT a match. Skip; missing pass handles it. (Domain is identified from chapter headings — see SYSTEM_PROMPT_MULTI_REF rules when applicable.)

---

## Field Format

**input_quote** — required. You MUST provide the ENTIRE relevant paragraph or complete section block from the document, not just the single line containing the value. 
  ציטוט:
  <הטקסט המלא והשלם של כל הפסקה/הסעיף מהמסמך - אין לקצר או להשמיט משפטים סמוכים השייכים לאותה פסקה>
  הקשר:
  <מידע נוסף על מיקום או תנאי סף במידת הצורך.>
**input_page** — required.
**ref_quote** — MANDATORY. You MUST provide the ENTIRE relevant paragraph or complete section block from the REFERENCE document that corresponds to the difference.
  ציטוט:
  <הטקסט המלא והשלם של כל הפסקה/הסעיף מה-REFERENCE - אין לקצר או להשמיט משפטים סמוכים השייכים לאותה פסקה>
  If you have nothing to quote → skip this item entirely.
**ref_page** — required.
**ref_file** — required.
**classification** — always "match".
**diff_type** — always "התאמה".
**severity** — always "נמוכה".
**is_material** — always false.
**confidence**: high = exact wording or clear semantic equivalence; medium = similar structure, minor variation
**reason**: עד 40 מילים, **בעברית בלבד** — חייב לכלול: (א) מה התוכן המשותף, (ב) אישור מפורש שהמבנה המשפטי זהה — שני המסמכים מנוסחים כהענקת כיסוי חיובית, או שניהם כחריג-עם-חריג, (ג) תווית המבנה שהוקצתה לכל מסמך: [affirmative-grant] או [exclusion-with-exception].

---

## Pre-Output Purge — run before returning JSON

For every finding already in your output list, answer these questions in order:

1. "Did I assign explicit structure labels ([affirmative-grant] or [exclusion-with-exception]) to both INPUT and REFERENCE before comparing content?"
   - If no → REMOVE this finding.

2. "Do both labels match?"
   - If the labels differ → REMOVE this finding.

3. "Did strong keyword or terminology overlap between INPUT and REFERENCE influence me toward a match before I completed the structural check?"
   - If yes → re-run the structural check now with fresh eyes. If labels differ → REMOVE.

4. "Does my own reason text describe any structural, framing, or legal-position difference?"
   - If yes → REMOVE this finding.
   - Also REMOVE if reason contains gap-description phrases such as: "מוסיף תנאי", "אינו מתנה ב", "ללא תנאים", "מצמצם", "מרחיב לכלול", "שינוי ניסוח", "הבדל מהותי", "השלכות פרשניות" — these describe a difference, not a match.

Only findings where both documents share the same default coverage position (both affirmative grants, or both exclusion-with-exception) may remain.

---
""" + _JSON_STRING_ESCAPING_RULES + """
## Output rules — CRITICAL

- Output **only** the raw JSON object. No preamble, no explanation, no markdown, no prose before or after.
- Do NOT wrap in ```json fences.
- The very first character of your response must be `{` and the very last must be `}`.
- Do NOT add any text after the closing `}`.

## Output example

{
  "findings": [
    {
      "classification": "match",
      "section": "תנאים כלליים - ויתור על זכות תחלוף",
      "input_quote": "ציטוט:\nסעיף 9 — ויתור על זכות התחלוף משוכרים ו/או דיירים של הנ\\"ל",
      "input_page": "עמ' 35",
      "ref_quote": "ציטוט:\nתנאים כלליים סעיף 7 — ויתור על זכות התחלוף כנגד משוכרים ו/או דיירים של הנ\\"ל",
      "ref_page": "עמ' 1",
      "diff_type": "התאמה",
      "severity": "נמוכה",
      "is_material": false,
      "confidence": "high",
      "reason": "התאמה מלאה על סעיף ויתור זכות התחלוף — אותו היקף ואותם הצדדים המכוסים. שני המסמכים: [affirmative-grant] — המבנה המשפטי זהה.",
      "note": ""
    }
  ]
}

## Negative example — DO NOT output (keyword-overlap false positive)

INPUT and REFERENCE share many keywords on the same topic but differ in framing and conditions. **Skip entirely** — the difference pass owns it:

INPUT: "הביטוח לפי פוליסה זו לא יושפע משינויים או תיאור בלתי מדויק של מטרת השימוש..."
REFERENCE: "הביטוח לפי פוליסה זו לא יחריג מקרה ביטוח בשל שינויים... ובתנאי שנעשה בתום לב ו/או אינו נחשב לעניין מהותי"

Why skip (not match):
- Labels differ: INPUT = [affirmative-grant], REFERENCE = [exclusion-with-exception]
- REFERENCE adds qualifying condition ("בתום לב") absent from INPUT
- High keyword overlap is a trap — structural check must run first

Correct action: omit from match-pass output. If reporting in combined pass → `difference`, `שינוי ניסוח`.

"""

SYSTEM_PROMPT_COMBINED_WINDOW = """You are an expert Israeli insurance policy analyst.

## Window context

INPUT is a **slice** of the full tender document. A Hebrew focus preamble at the top marks:
- **Focus chunks** — analyze every requirement here; report all findings for these clauses.
- **Context-only neighbor chunks** — use only for clauses split across chunk boundaries; do not re-report clauses whose focus is entirely outside the focus range.

REFERENCE is the **full** proposed policy — search all of it for counterparts.

## Primary rule: comparison direction

INPUT  = tender/procurement requirements (source of truth).
REFERENCE = proposed policy under test.
Every finding must originate from a requirement in INPUT. Never report REFERENCE-only topics.

## INPUT document layers — all in scope

INPUT may contain several layers. Report findings from **every layer** present in the focus pages:

1. **מכתב פניה / מכרז פרק 1** — cover letter, tender conditions, bid rules, ערבות הגשה, SLA, waivers, binding-document statements.
2. **מפרט / נספח פוליסות** — policy wording, limits, extensions, tables.
3. **תנאים כלליים** — general conditions (subrogation, cancellation, jurisdiction, indemnity).

Do NOT skip layer (1) because it is "administrative" — if it states a **binding insurance or contractual requirement**, compare it to REFERENCE or report **missing**.
Skip only: pure cover art, logos, table-of-contents lines with no clause body, and purely definitional glossary entries with no requirement.

**input_page rule:** Quote from the page where the requirement actually appears. A requirement stated on page 2 of the letter must use `input_page` from that page's `--- עמוד N ---` marker — do not attribute it to a later spec page that merely repeats the same topic.

---

## Structural framing — Step 0 (mandatory before classifying match vs difference)

Assign [affirmative-grant] or [exclusion-with-exception] to INPUT and REFERENCE separately before deciding match vs difference.

**Exclusion-with-exception signals (Hebrew):**
- "הפוליסה אינה מכסה... למעט..."
- "מוסכם ומוצהר כי אין כיסוי... אלא אם..."
- "לא יחריג מקרה ביטוח... בתנאי / אלא אם / ו/או..."
- "לא יפטר את המבטח... למעט..."
- Any clause that negates exclusion ("לא יחריג", "לא יפטר") then carves back coverage with trailing conditions

**Affirmative-grant signals (Hebrew):**
- "הביטוח חל על..." / "מכוסה על פי פוליסה זו..."
- "לא יושפע" / "לא ייפגע" / "הכיסוי ימשיך לחול" / "לא יושפע משינויים"

**Framing distinction — high keyword overlap does NOT mean equivalence:**
- INPUT: "לא יושפע משינויים..." (unconditional protection) → [affirmative-grant]
- REFERENCE: "לא יחריג מקרה ביטוח בשל..." with trailing "בתנאי שנעשה..." → [exclusion-with-exception]
→ always **difference**, never match.

---

## Classification decision tree — follow in exact order

1. Is the topic mentioned at all in REFERENCE (even in different wording)?
   No  → **missing**
   Yes → continue to question 2

2. Are amount/limit/period/scope identical AND wording does not change meaning AND both documents use the **same legal structure** (both [affirmative-grant] or both [exclusion-with-exception])?
   Yes → **match**
   No  → **difference**

**Legal structure rule:** [affirmative-grant] vs [exclusion-with-exception] on the same topic → always **difference**, never match.

**Iron rules:**
- REFERENCE covers topic but with less/more/different value or framing → **difference**, never missing.
- **difference** with empty ref_quote → error; quote REFERENCE or reclassify as missing.
- **match** with empty ref_quote → invalid; reclassify as missing or difference.
- Schedule deferral ("נקבע ברשימה") when INPUT states explicit value → **difference**.
- INPUT explicitly excludes coverage and REFERENCE includes it → **difference**.
- Letter/tender requirement with no policy counterpart in REFERENCE → **missing** (empty ref_quote allowed); do not skip.

---

## Missing — when to use

Report **missing** when INPUT states a requirement AND REFERENCE has **no clause body** addressing it.

This includes מכתב פניה / מכרז items (waiver, SLA, ערבות, binding appendix wording) that impose obligations but appear nowhere in REFERENCE policy text — report **missing** with `reason` explaining the gap; `ref_quote` may be empty.

**Absent (→ missing):**
- No REFERENCE text on the topic
- Chapter title / table-of-contents entry only — no clause body with value/condition
- Topic in a list word only with no limits/conditions
- Deferred to schedule with no actual value anywhere

**NOT missing (→ difference or match):**
- REFERENCE has any clause body (value, condition, exclusion-exception carve-out)
- Reason would say "קיים סעיף", "מופיע בסעיף", "REFERENCE provides…" → must NOT be missing

**Heading-only rule:** "מופיע בכותרת בלבד" / "מופיע ברשימה בלבד" → remains missing.

---

## Difference — when to use

REFERENCE addresses the topic but scope, amount, structure, or conditions differ from INPUT.

Check structural framing before calling a match:
- Assign [affirmative-grant] or [exclusion-with-exception] to INPUT and REFERENCE separately.
- Different labels → **difference** (שינוי ניסוח).
- REFERENCE adds restriction/condition absent from INPUT → **difference**.
- INPUT lists broader scope than REFERENCE → **difference**.

**ref_quote mandatory** for every difference finding.

If REFERENCE has a partial or approximate counterpart, report as **difference** with `confidence: low` rather than skipping — skipped clauses force expensive re-analysis.

---

## Match — when to use

REFERENCE **fully** satisfies INPUT: same structure, same scope, same amounts/limits.

Disqualifiers (→ difference or missing, not skip):
- Different legal structure (affirmative vs exclusion-exception)
- REFERENCE narrower scope than INPUT
- REFERENCE adds conditions INPUT lacks
- Wrong policy domain (cross-domain clause does not satisfy INPUT)
- Cannot quote ref_quote at all → reclassify as **missing**, not skip

---

## Severity

גבוהה: gap > 500,000 ₪ | coverage absent | contractual clause change | required exclusion absent from REFERENCE
בינונית: gap 50,000–500,000 ₪ | wording with interpretive impact
נמוכה: gap < 50,000 ₪ | match always low

---

## Field format

**input_quote** — required; include chapter + clause + exact value.
**input_page** — required. Use the PDF page number from the nearest `--- עמוד N ---` marker above the quoted INPUT text (e.g. `37` or `עמ' 37`). Never use the REFERENCE page number.
**ref_quote** — required for difference and match; may be empty only for missing with no counterpart.
**ref_page**, **ref_file** — required for difference and match.
**diff_type**: שינוי סכום | הוספת כיסוי | צמצום כיסוי | שינוי ניסוח | כיסוי חסר | שינוי אחוז | אחר | התאמה
**reason** — up to 60 words; for difference: "INPUT requires X, REFERENCE provides Y".
**is_material** — false only for low severity match.

---

## Quantity and completeness

Report **ALL** meaningful findings in the **focus chunk range** — no minimum floor, no maximum cap.

**Per-page rule (hard):** For every `--- עמוד N ---` marker inside the focus range, you MUST produce at least one finding whose `input_page` equals N. Never leave a focus page with zero findings unless it contains only cover art or a pure table-of-contents with no clause body.

**Per-section rule (hard):** For every identifiable section heading, table row, extension clause, or general condition in a focus page — classify it explicitly as `missing`, `difference`, or `match` before moving to the next section. Do NOT skip a section because it "looks standard" or "probably matches".

Scan every layer: letter/tender items, policy spec clauses, table rows, and extensions — not just chapter headers.
Do NOT skip a clause because it looks "standard" or "administrative" — verify against REFERENCE.
Do not report cover-only pages or TOC lines with no requirement body.
Do not duplicate the same clause twice in one response (same section + same input_quote on the same input_page).

## Pre-Output Purge — MANDATORY for match findings

For every finding you have already written with `classification: "match"`, answer these questions before finalising output:

1. **Is `diff_type` a substantive difference type?**
   If `diff_type` is anything other than `"התאמה"` (e.g. `"שינוי ניסוח"`, `"שינוי סכום"`, `"צמצום כיסוי"`, `"הוספת כיסוי"`) → change `classification` to `"difference"`. A match finding must always have `diff_type = "התאמה"`.

2. **Does the reason text describe a gap?**
   If `reason` contains any of: `"מוסיף תנאי"`, `"מצמצם"`, `"שינוי ניסוח"`, `"הבדל מהותי"`, `"ובלבד"`, `"בתנאי ש"`, `"קיים הבדל"`, `"יש פער"`, `"השלכות פרשניות"` — these describe a difference, not a match → change `classification` to `"difference"`.

3. **Is `ref_quote` empty?**
   A match with no ref_quote is invalid → change `classification` to `"missing"`.

4. **Does REFERENCE add ANY qualifier absent from INPUT?**
   Even a trailing `"ובלבד..."`, `"בתנאי ש..."`, or `"אלא אם..."` that does not appear in INPUT → NOT a match → `"difference"`, `diff_type = "שינוי ניסוח"`.

**Negative example — trailing qualifier (DO NOT output as match):**

INPUT: "המבוטח יהיה זכאי לבקש והמבטח מתחייב להסכים להארכת הכיסוי... למשך 60 יום נוספים"
REFERENCE: same text + `"ובלבד שהתקופה הכוללת לא תעלה על 18 חודשים"`

→ REFERENCE adds a restrictive condition (18-month cap) absent from INPUT.
→ **NOT a match** → `classification: "difference"`, `diff_type: "שינוי ניסוח"`.
High keyword overlap between INPUT and REFERENCE is NOT evidence of equivalence — it is a warning to check more carefully.

---

## Self-check before output

Before returning JSON, verify:
- Every substantive requirement in the focus pages has at least 3 finding (including מכתב פניה numbered items)
- Every `input_page` matches the `--- עמוד N ---` marker for the quoted INPUT text (letter items → letter page, not a later repeat)
- Every `difference`/`match`: ref_quote + ref_page + ref_file all non-empty
- Letter/tender requirements with no REFERENCE counterpart are reported as **missing**, not omitted
- Every focus page (`--- עמוד N ---`) has at least 3 finding — if any focus page has zero findings, either add a `missing`/`difference` finding now or confirm the page is cover/TOC-only with no clause body
- Every section heading encountered in the focus pages is represented as the `section` field of at least one finding
- Every `match` finding: `diff_type = "התאמה"` and reason does NOT contain gap-description phrases

---
""" + _JSON_STRING_ESCAPING_RULES + """
## Output — JSON only

{
  "findings": [
    {
      "classification": "difference",
      "section": "...",
      "input_quote": "דף 37: הפוליסה כוללת סעיף ויתור על תחלוף משוכרים של הנ\\"ל",
      "input_page": "37",
      "ref_quote": "...",
      "ref_page": "...",
      "ref_file": "...",
      "diff_type": "שינוי סכום",
      "severity": "גבוהה",
      "is_material": true,
      "confidence": "high",
      "reason": "...",
      "note": ""
    }
  ]
}

"""

SYSTEM_PROMPT_COMBINED_CHUNK_RAG = """
## Chunk RAG context (sparse-page retry)

REFERENCE in the user message is **retrieved chunks only** (Pinecone top-K), NOT the full policy.
Search only within the provided reference chunks for counterparts.

**Missing rule:** Report `missing` only when no provided reference chunk addresses the INPUT requirement.
If a related clause appears in any retrieved chunk (even with different framing) → `difference`, not `missing`.

**Match rule:** Do not report `match` unless a retrieved reference chunk fully satisfies the INPUT requirement
(same legal structure, same scope, same amounts/limits). High keyword overlap alone is not sufficient.
"""

# SYSTEM_PROMPT_MATCH = """You are an expert Israeli insurance policy analyst.

# ## Task: Find MATCH items only

# INPUT  = the tender/procurement document — the requirements the client MANDATES. Source of truth.
# REFERENCE = the proposed insurance policy — being tested against INPUT requirements.

# Your ONLY task in this pass: report key structural clauses from INPUT where REFERENCE **fully satisfies** the requirement — same scope, same meaning, same legal structure, even if wording differs slightly.

# Focus on meaningful structural clauses: main policy conditions, general terms, key coverage sections. Do NOT report every table row as a match — only report matches that provide positive signal about policy alignment.

# **CRITICAL: ref_quote is MANDATORY for every finding in this pass.** If you cannot quote the REFERENCE text → skip the item.

# ---

# ## Step 0: Structural Framing Gate — MANDATORY, runs before all other checks

# Before reading the content of any candidate clause, identify the legal structure each document uses for the topic:

# Ask yourself for INPUT: "Is this clause answering 'what is covered by the policy' or 'what is excluded by the policy'?"
# - If it answers "what is covered" → this is an **affirmative grant** (coverage is the default; in a dispute, the insurer bears the burden of proving an exception applies)
# - If it answers "what is excluded, with a carve-out for some cases" → this is an **exclusion-with-exception** (non-coverage is the default; in a dispute, the insured bears the burden of proving the exception applies)

# Ask the same question for REFERENCE.

# If INPUT and REFERENCE give different answers — one is an affirmative grant and the other is an exclusion-with-exception — → **STOP. This is NOT a match, regardless of how similar the subject matter appears.**

# The insured's legal position is opposite under these two structures. Coverage as a default (affirmative grant) is fundamentally different from coverage as a residual exception to an exclusion. Identical subject matter does not make them legally equivalent.
# → Skip; the difference pass handles it as שינוי ניסוח.

# ---

# ## Four conditions that DISQUALIFY a match — check ALL before reporting

# **1. REFERENCE adds ANY restriction, condition, or caveat absent from INPUT:**
# Ask yourself: "Does REFERENCE impose any qualifying condition, restriction, limitation, or caveat on this clause that is ABSENT from INPUT?"
# Do NOT rely on specific keywords — read the meaning, not the surface form.
# Signs of a restriction in REFERENCE that INPUT lacks:
# - Any inline condition ("in the event that...", "provided that...", "subject to...", "only when...")
# - Any numbered sub-clauses (1.1, 1.2, א, ב, etc.) beneath the REFERENCE clause that do not appear in INPUT
# - Any trailing qualification that limits when, how, or to whom the clause applies
# If ANY such restriction is present in REFERENCE but absent from INPUT → NOT a match.
# The added condition restricts the insured's rights regardless of how minor it seems.
# → Skip; the difference pass handles it as `שינוי ניסוח`.

# Example of the PATTERN (applies to any clause type, not only these):
# INPUT states a rule unconditionally: "Clause [A] applies."
# REFERENCE states the same rule with a condition: "Clause [A] applies, PROVIDED THAT [condition B]."
# → NOT a match — REFERENCE restricts what INPUT grants unconditionally.

# **2. INPUT lists more items/scope than REFERENCE covers:**
# If INPUT's clause applies to a broader set (more items, more parties, more scenarios, more coverage types) and REFERENCE only addresses a subset — even if the overlap is substantial — this is NOT a full match. Report a match ONLY when REFERENCE explicitly covers EVERYTHING INPUT requires.

# Example of the PATTERN:
# INPUT: "[clause] applies to [items A, B, C, D, E]"
# REFERENCE: "[clause] applies to [items A, B]"
# → NOT a match. Items C, D, E are uncovered.
# (The actual item names will differ per document — the pattern is what matters.)

# **3. ref_quote completeness — read to the end of the clause:**
# Your ref_quote MUST include the COMPLETE REFERENCE text for the clause — including any trailing conditions, qualifications, or restrictive caveats. Never quote only the shared/common portion. If reading the full clause reveals any difference from INPUT → skip this item and let the difference pass handle it.

# **4. Cross-policy domain disqualification:**
# If the only REFERENCE clause found is in a different policy domain from the INPUT requirement → NOT a match. Skip; missing pass handles it. (Domain is identified from chapter headings — see SYSTEM_PROMPT_MULTI_REF rules when applicable.)

# ---

# ## Field Format

# **input_quote** — required. Use the structured format:
#   ציטוט:
#   <הטקסט המדויק מהמסמך: "[פרק] — [סעיף]: [ערך מלא]">
#   הקשר:
#   <1–3 משפטים — רק כשנדרש הקשר נוסף. השמט כשהציטוט מובן מאליו.>
# **input_page** — required.
# **ref_quote** — MANDATORY. Must quote the matching REFERENCE text. Same structured format as input_quote.
#   ציטוט:
#   <הטקסט המדויק מה-REFERENCE>
#   הקשר:
#   <רק כשנדרש>
#   If you have nothing to quote → skip this item entirely.
# **ref_page** — required.
# **ref_file** — required.
# **classification** — always "match".
# **diff_type** — always "התאמה".
# **severity** — always "נמוכה".
# **is_material** — always false.
# **confidence**: high = exact wording or clear semantic equivalence; medium = similar structure, minor variation
# **reason**: עד 40 מילים, **בעברית בלבד** — חייב לכלול: (א) מה התוכן המשותף, (ב) אישור מפורש שהמבנה המשפטי זהה — שני המסמכים מנוסחים כהענקת כיסוי חיובית, או שניהם כחריג-עם-חריג.

# ---

# ## Pre-Output Purge — run before returning JSON

# For every finding already in your output list, answer this question:
# "Do INPUT and REFERENCE use the same legal structure for this topic — do both answer 'what is covered by the policy' or do both answer 'what is excluded by the policy'?"
# - If the structures differ → REMOVE this finding from the output list.
# - If your own reason text describes any structural, framing, or legal-position difference → REMOVE this finding from the output list.

# Only findings where both documents share the same default coverage position (both affirmative grants, or both exclusion-with-exception) may remain.

# ---

# ## Output rules — CRITICAL

# - Output **only** the raw JSON object. No preamble, no explanation, no markdown, no prose before or after.
# - Do NOT wrap in ```json fences.
# - The very first character of your response must be `{` and the very last must be `}`.
# - Do NOT add any text after the closing `}`.

# ## Output example

# {
#   "findings": [
#     {
#       "classification": "match",
#       "section": "תנאים כלליים - ויתור על זכות תחלוף",
#       "input_quote": "ציטוט:\nסעיף 9 — ויתור על זכות התחלוף כנגד: כל יחידי המבוטחים, עובדיו, בני משפחתם ונושאי משרה",
#       "input_page": "עמ' 35",
#       "ref_quote": "ציטוט:\nתנאים כלליים סעיף 7 — ויתור על זכות התחלוף כנגד כל יחידי המבוטחים, מנהליו, עובדיו ובני משפחתם ונושאי משרה",
#       "ref_page": "עמ' 1",
#       "diff_type": "התאמה",
#       "severity": "נמוכה",
#       "is_material": false,
#       "confidence": "high",
#       "reason": "התאמה מלאה על סעיף ויתור זכות התחלוף — אותו היקף ואותם הצדדים המכוסים. שני המסמכים מנוסחים כהענקת כיסוי חיובית — המבנה המשפטי זהה.",
#       "note": ""
#     }
#   ]
# }

# """

SYSTEM_PROMPT_MULTI_REF = """

## Multiple REFERENCE files (bundle)

- REFERENCE contains multiple policy/spec files, each marked with === REFERENCE FILE: <filename> ===.
- Compare INPUT against **all** reference files together.
- **ref_file** — required field on every finding (except missing findings, where it stays empty "").
- **missing** = the INPUT requirement was not found in **any** REFERENCE file in the bundle (not just one).
- **difference** / **match**: set ref_file to the most relevant file where you found the clause.
- Direction is always INPUT → REFERENCE. Never generate findings that originate from REFERENCE only.

## Cross-Policy Domain Rule — CRITICAL

When satisfying an INPUT requirement, the `ref_file` you cite **must belong to the same insurance domain** as the requirement.

**How to identify the domain:**
- Read the chapter heading, policy title, or section label that contains the INPUT requirement. That heading names the domain (e.g. "ביטוח X", "אחריות Y", "פרק Z").
- A REFERENCE clause satisfies that requirement ONLY if it comes from a REFERENCE file (or chapter) that covers the same domain — as identified by its own chapter/section heading.
- You must infer the domain mapping yourself from what you read in both documents. Do NOT rely on a predefined list of domain names — the domain names will differ across document sets.

**The rule:**
Do NOT use a clause from Domain A to satisfy a requirement from Domain B, even if the clause text sounds relevant. If the correct-domain REFERENCE file lacks the clause → classify as `missing`. Cross-domain matches are invalid.
A clause found in the wrong policy domain does **not** satisfy the INPUT requirement — classify as `missing`, never `match`.
"""


_PASS_TYPE_PROMPTS = {
    "missing":    SYSTEM_PROMPT_MISSING,
    "difference": SYSTEM_PROMPT_DIFFERENCE,
    "match":      SYSTEM_PROMPT_MATCH,
    "combined":   SYSTEM_PROMPT_COMBINED_WINDOW,
}


def fetch_custom_prompt_for_use_case(use_case_id: str) -> str:
    """
    Query phoenix-custom-prompts for the latest custom prompt addendum for
    this use case (reference_batch_id). Returns an empty string if none exists
    or if DynamoDB is unavailable (fail-safe — never block the diff job).
    """
    if not use_case_id:
        return ""
    try:
        from boto3.dynamodb.conditions import Key as DKey
        table = DYNAMODB_RESOURCE.Table(PROMPTS_TABLE)
        resp = table.query(
            KeyConditionExpression=DKey("use_case_id").eq(use_case_id),
            ScanIndexForward=False,
            Limit=1,
        )
        items = resp.get("Items", [])
        if items:
            text = items[0].get("prompt_text", "").strip()
            if text:
                print(f"[custom_prompt] loaded for use_case_id={use_case_id} ({len(text)} chars)")
                return text
    except Exception as exc:  # noqa: BLE001
        print(f"[custom_prompt] WARNING: could not fetch for use_case_id={use_case_id}: {exc}")
    return ""


def _system_prompt_for_compare(
    *,
    multi_ref: bool,
    dual_batch_files: Optional[list[str]] = None,
    pass_type: str = "combined",  # "missing" | "difference" | "match" | "combined"
    custom_prompt_addendum: str = "",
) -> str:
    prompt = _PASS_TYPE_PROMPTS.get(pass_type, SYSTEM_PROMPT)
    if multi_ref:
        prompt += SYSTEM_PROMPT_MULTI_REF
    if dual_batch_files:
        names = ", ".join(dual_batch_files)
        prompt += (
            f"\n## באנדל הנוכחי\n"
            f"קבצים בבאנדל זה בלבד: {names}.\n"
            f"לממצאי missing: לא נמצא באף אחד מקבצים אלה.\n"
        )
    if custom_prompt_addendum:
        prompt += CUSTOM_PROMPT_HEADER + custom_prompt_addendum
    return prompt


def _log_prompt_text_stats(label: str, text: str, cap: int, *, verbose: bool) -> None:
    """Log sent text size and whether truncation marker appears (CloudWatch-friendly)."""
    if not verbose:
        return
    truncated = TRUNCATION_MARKER in text
    print(
        f"  [{label}] sent_chars={len(text)} cap={cap} "
        f"truncated={truncated}"
    )


def _resolve_max_tokens(override: Optional[int] = None) -> int:
    """API override, then MAX_DIFF_OUTPUT_TOKENS env, then DEFAULT_MAX_TOKENS."""
    if override is not None:
        return int(override)
    env_val = os.environ.get("MAX_DIFF_OUTPUT_TOKENS")
    if env_val:
        try:
            return int(env_val)
        except ValueError:
            pass
    return DEFAULT_MAX_TOKENS


def _is_valid_finding_row(row: dict) -> bool:
    if not isinstance(row, dict):
        return False
    cls = row.get("classification")
    if cls not in ("match", "difference", "missing"):
        return False
    return bool(row.get("input_quote") or row.get("reason") or row.get("ref_quote"))


_GERSHAYIM_RE = re.compile(r'([\u0590-\u05FF])"([\u0590-\u05FF])')
_TRAILING_COMMA_RE = re.compile(r',(\s*[}\]])')


_VALID_JSON_ESCAPES = frozenset('"\\/bfnrtu')


def _fix_invalid_json_backslashes(raw: str) -> str:
    """Remove or fix backslashes that are invalid inside JSON string values."""
    out: list[str] = []
    in_string = False
    i = 0
    n = len(raw)
    while i < n:
        c = raw[i]
        if not in_string:
            out.append(c)
            if c == '"':
                in_string = True
            i += 1
            continue
        if c == "\\":
            if i + 1 >= n:
                i += 1
                continue
            nxt = raw[i + 1]
            if nxt in _VALID_JSON_ESCAPES:
                out.append(c)
                out.append(nxt)
                i += 2
                continue
            # Illegal escape (e.g. "\ " or "\(") — drop the backslash, keep the char.
            i += 1
            continue
        out.append(c)
        if c == '"':
            in_string = False
        i += 1
    return "".join(out)


def _escape_hebrew_gershayim(raw: str) -> str:
    """Escape ASCII quotes used as Hebrew gershayim between letters."""
    prev, cur = None, raw
    while cur != prev:
        prev = cur
        cur = _GERSHAYIM_RE.sub(r'\1\\"\2', cur)
    return cur


def _normalize_llm_json_text(raw: str) -> str:
    text = _TRAILING_COMMA_RE.sub(r"\1", raw)
    text = _fix_invalid_json_backslashes(text)
    return _escape_hebrew_gershayim(text)


def _parse_llm_json_object(raw: str) -> tuple[Optional[dict], bool]:
    """Return (parsed dict or None, parse_ok). parse_ok is True only when json.loads succeeds."""
    candidates: list[str] = [raw]
    normalized = _normalize_llm_json_text(raw)
    if normalized != raw:
        candidates.append(normalized)
    obj_match = re.search(r"\{.*\}", raw, re.DOTALL)
    if obj_match:
        block = obj_match.group()
        candidates.append(block)
        norm_block = _normalize_llm_json_text(block)
        if norm_block != block:
            candidates.append(norm_block)
    seen: set[str] = set()
    for text in candidates:
        if text in seen:
            continue
        seen.add(text)
        try:
            data = json.loads(text)
            if isinstance(data, dict):
                return data, True
        except json.JSONDecodeError:
            continue
    return None, False


def _salvage_findings_from_truncated_json(raw: str) -> list[dict]:
    """Parse complete finding objects from a truncated findings JSON array."""
    match = re.search(r'"findings"\s*:\s*\[', raw)
    if not match:
        return []
    array_body = raw[match.end() :]
    decoder = json.JSONDecoder()
    pos = 0
    findings: list[dict] = []
    max_objects = 200
    attempts = 0
    while pos < len(array_body) and attempts < max_objects:
        attempts += 1
        rest = array_body[pos:].lstrip()
        if not rest or rest[0] in "}]":
            break
        if rest[0] != "{":
            pos += len(array_body[pos:]) - len(rest) + 1
            continue
        try:
            obj, end = decoder.raw_decode(rest)
            pos += len(array_body[pos:]) - len(rest) + end
            if isinstance(obj, dict) and _is_valid_finding_row(obj):
                findings.append(obj)
            continue
        except json.JSONDecodeError:
            try:
                normalized_rest = _normalize_llm_json_text(rest)
                if normalized_rest != rest:
                    obj, end = decoder.raw_decode(normalized_rest)
                    pos += len(array_body[pos:]) - len(rest) + end
                    if isinstance(obj, dict) and _is_valid_finding_row(obj):
                        findings.append(obj)
                    continue
            except json.JSONDecodeError:
                pass
            next_brace = rest.find("{", 1)
            if next_brace < 0:
                break
            pos += len(array_body[pos:]) - len(rest) + next_brace
    return findings


def _claude_pass_needs_retry(result: dict) -> tuple[bool, str]:
    """Return (should_retry, log_reason) for a single Claude pass result."""
    if result.get("_api_failed"):
        return False, ""
    raw_len = len(result.get("_raw") or "")
    tok = result.get("output_tokens") or 0
    salvage_only = result.get("salvaged") and not result.get("_llm_repaired")
    _raw_stripped = (result.get("_raw") or "").strip()
    _is_valid_empty = _raw_stripped.startswith("{") and "findings" in _raw_stripped
    _lazy_empty = _is_valid_empty and raw_len < 100 and 0 < tok < 30
    suspiciously_empty = not result["findings"] and (
        (not _is_valid_empty and (raw_len < 100 or (tok > 0 and tok < 100)))
        or _lazy_empty
    )
    malformed_json_empty = (
        not result["findings"]
        and not result.get("parse_ok", True)
        and tok > 300
        and raw_len > 500
    )
    if malformed_json_empty:
        return True, f"malformed JSON with large output ({raw_len} chars / {tok} tokens)"
    if salvage_only:
        return True, "salvage-only recovery (malformed JSON)"
    if suspiciously_empty:
        return True, f"{raw_len} raw chars / {tok} tokens"
    return False, ""


_FALSE_MISSING_EXCEPTIONS = (
    "מופיע בכותרת",
    "מופיע ברשימה בלבד",
    "mentioned only in heading",
    "בכותרת בלבד",
)



def _is_false_missing(reason: str) -> bool:
    """True when reason text indicates REFERENCE has a clause body (not truly missing)."""
    if not reason:
        return False
    for exc in _FALSE_MISSING_EXCEPTIONS:
        if exc in reason:
            return False
    reason_lower = reason.lower()
    if any(p in reason for p in ("קיים סעיף", "קיים כיסוי", "מופיע בסעיף")):
        return True
    if "reference" in reason_lower and "קיים" in reason:
        return True
    return any(
        p in reason_lower
        for p in (
            "reference has a clause",
            "reference covers",
            "reference page",
            "reference provides",
        )
    )


def _strip_quote_label(text: str) -> str:
    """Remove 'ציטוט:' prefix from quote fields before validation."""
    t = (text or "").strip()
    if t.startswith("ציטוט:"):
        t = t[len("ציטוט:"):].strip()
    return t


_ABSENT_REF_RE = re.compile(
    r"^לא נמצא|^לא נמצאה|^לא נמצאו|^לא קיים|^לא קיימת|^לא קיימים|"
    r"^לא צוין|^not found|^no equivalent|^no matching",
    re.IGNORECASE,
)

_INPUT_CITATION_IN_REF_RE = re.compile(
    r"של INPUT|ב-INPUT|from INPUT|INPUT מתייחס",
    re.IGNORECASE,
)

_FALSE_DIFFERENCE_REF_PHRASES = (
    "לא נמצאה התייחסות",
    "לא צוין סעיף",
    "לא קיימת הרחבה",
    "אין התייחסות",
    "not found",
    "no matching",
)

_FALSE_DIFFERENCE_REASON_PHRASES = (
    "אינו כולל",
    "לא נמצא",
    "נעדר לחלוטין",
    "לא נמצאה הרחבה",
    "אין סעיף",
    "completely absent",
    "not found in reference",
)


def _has_substantive_ref_quote(ref_text: str) -> bool:
    """True when ref_quote looks like an actual policy clause, not a header/absence."""
    t = (ref_text or "").strip()
    if len(t) < 25:
        return False
    if _INPUT_CITATION_IN_REF_RE.search(t):
        return False
    if _ABSENT_REF_RE.search(t):
        return False
    if any(p in t for p in _FALSE_DIFFERENCE_REF_PHRASES):
        return False
    return True


def _is_false_difference(finding: dict) -> bool:
    """True when a difference finding has no actual REFERENCE clause quoted."""
    ref_text = _strip_quote_label(
        finding.get("ref_quote") or finding.get("reference_text") or ""
    )
    reason = (finding.get("reason") or "").strip()
    reason_lower = reason.lower()
    substantive_ref = _has_substantive_ref_quote(ref_text)

    if ref_text:
        if _INPUT_CITATION_IN_REF_RE.search(ref_text):
            return True
        if _ABSENT_REF_RE.search(ref_text):
            return True
        if any(p in ref_text for p in _FALSE_DIFFERENCE_REF_PHRASES):
            return True

    # Reason-only phrases (e.g. "אינו כולל") apply only when ref_quote is absent or weak.
    if not substantive_ref:
        if reason and any(p in reason for p in _FALSE_DIFFERENCE_REASON_PHRASES):
            return True
        if reason_lower and any(
            p in reason_lower
            for p in ("completely absent", "not found in reference")
        ):
            return True

    return False


def _merge_three_pass_findings(
    missing: list[dict],
    difference: list[dict],
    match: list[dict],
    *,
    verbose: bool = True,
) -> tuple[list[dict], list[dict]]:
    """
    Merge findings from the three specialized passes.
    Priority: difference > missing > match.
    Lower-priority findings that overlap with a higher-priority finding are suppressed.

    Returns (merged, suppressed) where suppressed contains all findings that were
    dropped during the merge with an added "_suppressed_reason" field.
    """
    suppressed: list[dict] = []

    # Highest priority: keep validated difference findings only
    validated_diff: list[dict] = []
    for f in difference:
        if _is_false_difference(f):
            if verbose:
                print(
                    f"  [false-difference guard] suppressed — no REFERENCE clause quoted: "
                    f"{(f.get('section') or '')[:60]}"
                )
            suppressed.append(
                {**f, "_suppressed_reason": "false difference — no REFERENCE clause quoted"}
            )
        else:
            validated_diff.append(f)
    merged: list[dict] = list(validated_diff)

    # Add missing findings that don't overlap with any validated difference finding
    for f in missing:
        reason = (f.get("reason") or "").strip()
        if _is_false_missing(reason):
            if verbose:
                print(
                    f"  [false-missing guard] suppressed — reason mentions REFERENCE clause: "
                    f"{reason[:80]}"
                )
            suppressed.append(
                {**f, "_suppressed_reason": "false missing — reason mentions REFERENCE clause"}
            )
            continue
        if not any(_dedupe_findings_overlap(f, d) for d in validated_diff):
            merged.append(f)
        else:
            if verbose:
                print(
                    f"  [merge-3pass] suppressed missing (overlap with difference): "
                    f"{(f.get('section') or '')[:60]}"
                )
            suppressed.append({**f, "_suppressed_reason": "overlap with difference"})

    # Add match findings that don't overlap with anything already accepted.
    # Use _dedupe_findings_overlap (quote/exact-section only, no word-set overlap)
    # so that findings in the same chapter but different sub-clauses are not
    # incorrectly suppressed.
    for f in match:
        overlapping = [m for m in merged if _dedupe_findings_overlap(f, m)]
        if not overlapping:
            merged.append(f)
        else:
            overlap_cls = overlapping[0].get("classification", "unknown")
            if verbose:
                print(
                    f"  [merge-3pass] suppressed match (overlap with {overlap_cls}): "
                    f"{(f.get('section') or '')[:60]}"
                )
            suppressed.append({**f, "_suppressed_reason": f"overlap with {overlap_cls}"})

    if verbose:
        n_diff  = sum(1 for f in merged if f.get("classification") == "difference")
        n_miss  = sum(1 for f in merged if f.get("classification") == "missing")
        n_match = sum(1 for f in merged if f.get("classification") == "match")
        print(
            f"  [merge-3pass] difference={n_diff} missing={n_miss} match={n_match} "
            f"total={len(merged)} suppressed={len(suppressed)} "
            f"(raw: diff={len(difference)} miss={len(missing)} match_raw={len(match)})"
        )

    return merged, suppressed


def _split_findings_by_classification(findings: list[dict]) -> dict[str, list[dict]]:
    """Bucket findings by classification for backward-compat _pass_findings output."""
    buckets: dict[str, list[dict]] = {"missing": [], "difference": [], "match": []}
    for f in findings:
        cls = _normalize_classification(f.get("classification", ""))
        if cls in buckets:
            buckets[cls].append(f)
    return buckets


def _apply_output_guards(
    findings: list[dict],
    *,
    verbose: bool = True,
) -> tuple[list[dict], list[dict]]:
    """Filter false-positive difference/missing rows from combined-pass output."""
    kept: list[dict] = []
    suppressed: list[dict] = []
    for f in findings:
        cls = _normalize_classification(f.get("classification", ""))
        if cls == "difference" and _is_false_difference(f):
            if verbose:
                print(
                    f"  [false-difference guard] suppressed — no REFERENCE clause quoted: "
                    f"{(f.get('section') or '')[:60]}"
                )
            suppressed.append({
                **f,
                "_suppressed_reason": "false difference — no REFERENCE clause quoted",
            })
            continue
        reason = (f.get("reason") or "").strip()
        if cls == "missing" and _is_false_missing(reason):
            if verbose:
                print(
                    f"  [false-missing guard] suppressed — reason mentions REFERENCE clause: "
                    f"{reason[:80]}"
                )
            suppressed.append({
                **f,
                "_suppressed_reason": "false missing — reason mentions REFERENCE clause",
            })
            continue
        kept.append(f)

    if verbose:
        print(
            f"  [output-guards] kept={len(kept)} ({_findings_count_summary(kept)}) "
            f"suppressed={len(suppressed)}"
        )
    return kept, suppressed


# ---------------------------------------------------------------------------
# Completeness-check prompt & function (4th LLM pass, runs after merge)
# ---------------------------------------------------------------------------

SYSTEM_PROMPT_COMPLETENESS = """You are an expert Israeli insurance policy analyst performing a final completeness review.

## Context

You will receive:
1. INPUT — the client's tender/procurement document with insurance requirements.
2. REFERENCE — the proposed insurance policy.
3. EXISTING_FINDINGS — a JSON list of findings already identified (differences, missing items, matches).

## Your task

Scan every chapter and clause in INPUT. For each clause, check whether it already appears in EXISTING_FINDINGS (by section name or topic).

Report **only** clauses that are completely absent from EXISTING_FINDINGS (no finding with overlapping section/topic) AND fall into one of these buckets:

**Bucket 1 — missing from REFERENCE:**
- Genuinely missing from REFERENCE (no specific, actionable coverage provided)
- → `classification: "missing"` (use missing-pass field format below)

**Bucket 2 — present in REFERENCE but unrouted:**
- REFERENCE contains a related clause (same topic, exclusion-exception pattern, or parallel extension) but with different framing or conditions
- → `classification: "difference"` with full `ref_quote`, `ref_page`, `ref_file` (use difference-pass field format)

**Trigger for bucket 2:** If an INPUT numbered sub-clause has **zero findings** in EXISTING_FINDINGS and REFERENCE contains a related clause → report as `difference`, not skip.

**Bucket 3 — misrouted missing in EXISTING_FINDINGS:**
- Scan EXISTING_FINDINGS for rows with `classification: "missing"` whose `reason` states REFERENCE **has** a clause body (e.g. `"קיים סעיף"`, `"קיים כיסוי"`, `"מופיע בסעיף"`, `"ב-REFERENCE קיים"`) but a sub-condition or framing differs from INPUT
- Treat these as **misrouted** — emit a corrected `difference` finding with full `ref_quote`, `ref_page`, `ref_file` (even though a `missing` row exists for the same topic)
- Typical pattern: reason says `"קיים סעיף … אך ללא תנאי …"` or `"נעדר תנאי"` while also acknowledging REFERENCE coverage text

**Cross-check before output:** Verify every numbered sub-clause in each INPUT spec chapter appears in at least one finding (EXISTING_FINDINGS or your new output). If not, investigate and add.

**Do NOT re-report** findings that are already correctly classified in EXISTING_FINDINGS. This pass fills gaps and corrects misrouted `missing` rows only.

## Severity Criteria (same as main pass)

גבוהה  (high):   clause completely absent from REFERENCE | no value provided where INPUT requires one | INPUT explicitly requires an exclusion clause that is entirely absent from REFERENCE (any exclusion type)
בינונית (medium): extension absent or only vaguely mentioned | condition deferred without a value | framing difference with interpretive impact
נמוכה  (low):    minor clause absent with no financial impact

## Field Format

**For `missing` findings** (bucket 1):
**classification** — "missing"
**section** — chapter + sub-clause name (e.g. "ביטוח אש מורחב - שיקום קרקע")
**input_quote** — "[chapter] — [clause]: [full value]"
**input_page** — required
**ref_quote** — always ""
**ref_page** — always ""
**ref_file** — always ""
**diff_type** — always "כיסוי חסר"
**severity** — per criteria above
**is_material** — true for medium/high; false for low
**confidence** — high | medium | low
**reason** — up to 60 words

**For `difference` findings** (bucket 2):
**classification** — "difference"
**section**, **input_quote**, **input_page** — required
**ref_quote**, **ref_page**, **ref_file** — required (MANDATORY)
**diff_type** — שינוי ניסוח | צמצום כיסוי | אחר (as appropriate)
**severity** — per criteria above
**is_material** — true for medium/high; false for low
**confidence** — high | medium | low
**reason** — up to 60 words

## Output rules

- Output ONLY the raw JSON object. No preamble, no markdown, no prose.
- First character must be `{`, last must be `}`.

{
  "findings": [ ... ]
}

"""


def _call_claude_completeness_check(
    client: AnthropicBedrock,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    model: str,
    merged_findings: list[dict],
    max_tokens: int = 16000,
    verbose: bool = True,
) -> list[dict]:
    """
    4th LLM pass: completeness check.
    Receives the already-merged findings list and asks Claude to report INPUT
    clauses absent from EXISTING_FINDINGS — either genuinely missing from
    REFERENCE (missing) or present with different framing (difference).
    Returns a (possibly empty) list of additional findings.
    """
    if not merged_findings and not input_text:
        return []

    # Build a compact summary of existing findings for context
    existing_summary_lines = []
    for f in merged_findings:
        section = (f.get("section") or "")[:80]
        cls = f.get("classification", "")
        line = f"- [{cls}] {section}"
        if _normalize_classification(cls) == "missing":
            reason = (f.get("reason") or "")[:120]
            if reason:
                line += f" | reason: {reason}"
        existing_summary_lines.append(line)
    existing_summary = "\n".join(existing_summary_lines) or "(none)"

    user_message = (
        f"## INPUT (מסמך המכרז)\n{input_text}\n\n"
        f"## EXISTING_FINDINGS (already identified — do NOT re-report these)\n"
        f"{existing_summary}"
    )
    ref_block = f"## REFERENCE ({ref_filename})\n{ref_text}"

    try:
        with client.messages.stream(
            model=model,
            max_tokens=max_tokens,
            system=SYSTEM_PROMPT_COMPLETENESS,
            messages=[{"role": "user", "content": [
                {"type": "text", "text": ref_block, "cache_control": {"type": "ephemeral"}},
                {"type": "text", "text": user_message},
            ]}],
        ) as stream:
            response = stream.get_final_message()
        usage = getattr(response, "usage", None)
        input_tokens          = getattr(usage, "input_tokens",               None) if usage else None
        output_tokens         = getattr(usage, "output_tokens",              None) if usage else None
        cache_creation_tokens = getattr(usage, "cache_creation_input_tokens", None) if usage else None
        cache_read_tokens     = getattr(usage, "cache_read_input_tokens",     None) if usage else None

        raw = _extract_text_from_response(response)
        fence_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
        if fence_match:
            raw = fence_match.group(1).strip()

        data, parse_ok = _parse_llm_json_object(raw)
        findings = _parse_findings_list(data) if data is not None else []
        if verbose:
            cache_status = (
                f"cache_creation={cache_creation_tokens} cache_read={cache_read_tokens}"
                if (cache_creation_tokens or cache_read_tokens)
                else "cache=off"
            )
            print(
                f"  [completeness-check] ref={ref_filename} "
                f"input={input_tokens} output={output_tokens} {cache_status} "
                f"existing={len(merged_findings)} "
                f"new_findings={len(findings)} ({_findings_count_summary(findings)})"
            )
            if not findings:
                raw_preview = raw[:2000]
                truncated_note = "…(truncated)" if len(raw) > 2000 else ""
                print(f"  [completeness-check] RAW response:\n{raw_preview}{truncated_note}")
        return findings

    except Exception as exc:
        if verbose:
            print(f"  [WARN] completeness-check failed for {ref_filename}: {exc}")
        return []


def _call_claude_three_pass(
    client: AnthropicBedrock,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    model: str,
    max_tokens: int = DEFAULT_MAX_TOKENS,
    verbose: bool = True,
    multi_ref: bool = False,
    dual_batch_files: Optional[list[str]] = None,
    ref_preamble: str = "",
    custom_prompt_addendum: str = "",
    debug_job_id: Optional[str] = None,
    debug_s3_bucket: Optional[str] = None,
    debug_s3_prefix: Optional[str] = None,
) -> dict:
    """
    Run three parallel specialized Claude calls (missing / difference / match) and merge.
    The three calls are independent and execute concurrently via ThreadPoolExecutor.
    Returns the same dict shape as _call_claude().
    """
    pass_types = ("missing", "difference", "match")

    def _run_pass(pass_type: str) -> tuple[str, dict]:
        sp = _system_prompt_for_compare(
            multi_ref=multi_ref,
            dual_batch_files=dual_batch_files,
            pass_type=pass_type,
            custom_prompt_addendum=custom_prompt_addendum,
        )
        #print('len of input_text is : ' ,len(input_text))
        #print('input_text is : ' ,input_text)
        result = _call_claude(
            client,
            input_text=input_text,
            ref_text=ref_text,
            ref_filename=ref_filename,
            model=model,
            max_tokens=max_tokens,
            verbose=verbose,
            system_prompt=sp,
            ref_preamble=ref_preamble,
            pass_label=pass_type,
            debug_job_id=debug_job_id,
            debug_s3_bucket=debug_s3_bucket,
            debug_s3_prefix=debug_s3_prefix,
        )
        # Retry once if:
        # (a) the response is suspiciously empty — 0 findings AND the raw text does NOT look
        #     like a valid empty sentinel (starts with "{" and contains "findings"), AND
        #     raw text < 100 chars or output_tokens < 100. This catches truncated/malformed
        #     responses like `{}` while skipping legitimate `{"findings":[]}` results, OR
        # (b) salvage was the only recovery method (LLM repair unavailable/failed) —
        #     meaning partial findings were returned and a fresh call may produce clean JSON, OR
        # (c) large malformed JSON blob that failed parse and yielded 0 findings.
        needs_retry, retry_reason = _claude_pass_needs_retry(result)
        if needs_retry:
            if verbose:
                print(f"      [{pass_type}] [retry] retrying once ({retry_reason})")
            result = _call_claude(
                client,
                input_text=input_text,
                ref_text=ref_text,
                ref_filename=ref_filename,
                model=model,
                max_tokens=max_tokens,
                verbose=verbose,
                system_prompt=sp,
                ref_preamble=ref_preamble,
                pass_label=f"{pass_type}:retry",
                debug_job_id=debug_job_id,
                debug_s3_bucket=debug_s3_bucket,
                debug_s3_prefix=debug_s3_prefix,
            )
        return pass_type, result

    results: dict[str, dict] = {}
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {pool.submit(_run_pass, pt): pt for pt in pass_types}
        for fut in as_completed(futures):
            pass_type, result = fut.result()
            results[pass_type] = result

    merged, suppressed = _merge_three_pass_findings(
        missing=results["missing"]["findings"],
        difference=results["difference"]["findings"],
        match=results["match"]["findings"],
        verbose=verbose,
    )

    # 4th pass: completeness check — find clauses entirely absent from merged findings
    completeness_findings = _call_claude_completeness_check(
        client,
        input_text=input_text,
        ref_text=ref_text,
        ref_filename=ref_filename,
        model=model,
        merged_findings=merged,
        verbose=verbose,
    )
    if completeness_findings:
        if verbose:
            print(
                f"  [completeness-check] adding {len(completeness_findings)} new finding(s) "
                f"to merged list"
            )
        merged = merged + completeness_findings

    return {
        "findings": merged,
        "output_truncated": any(r["output_truncated"] for r in results.values()),
        "salvaged":         any(r["salvaged"]          for r in results.values()),
        "stop_reason":      results["difference"]["stop_reason"],
        # _raw is intentionally not forwarded — it is internal to _call_claude_three_pass
        "_pass_findings": {
            "missing":    results["missing"]["findings"],
            "difference": results["difference"]["findings"],
            "match":      results["match"]["findings"],
        },
        "_suppressed": suppressed,
    }


def _run_combined_pass(
    client: AnthropicBedrock,
    *,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    model: str,
    max_tokens: int,
    verbose: bool,
    multi_ref: bool,
    dual_batch_files: Optional[list[str]],
    ref_preamble: str,
    custom_prompt_addendum: str,
    pass_label: str,
    debug_job_id: Optional[str],
    debug_s3_bucket: Optional[str],
    debug_s3_prefix: Optional[str],
) -> dict:
    """Run one combined classify pass with empty/salvage retry."""
    empty_result = {
        "findings": [],
        "output_truncated": False,
        "salvaged": False,
        "_llm_repaired": False,
        "stop_reason": None,
        "_raw": "",
        "output_tokens": 0,
    }
    if not (input_text or "").strip():
        return empty_result

    sp = _system_prompt_for_compare(
        multi_ref=multi_ref,
        dual_batch_files=dual_batch_files,
        pass_type="combined",
        custom_prompt_addendum=custom_prompt_addendum,
    )
    result = _call_claude(
        client,
        input_text=input_text,
        ref_text=ref_text,
        ref_filename=ref_filename,
        model=model,
        max_tokens=max_tokens,
        verbose=verbose,
        system_prompt=sp,
        ref_preamble=ref_preamble,
        pass_label=pass_label,
        debug_job_id=debug_job_id,
        debug_s3_bucket=debug_s3_bucket,
        debug_s3_prefix=debug_s3_prefix,
    )

    needs_retry, retry_reason = _claude_pass_needs_retry(result)
    if needs_retry:
        if verbose:
            print(f"      [{pass_label}] [retry] retrying once ({retry_reason})")
        result = _call_claude(
            client,
            input_text=input_text,
            ref_text=ref_text,
            ref_filename=ref_filename,
            model=model,
            max_tokens=max_tokens,
            verbose=verbose,
            system_prompt=sp,
            ref_preamble=ref_preamble,
            pass_label=f"{pass_label}:retry",
            debug_job_id=debug_job_id,
            debug_s3_bucket=debug_s3_bucket,
            debug_s3_prefix=debug_s3_prefix,
        )
    return result


def _call_nova_combined(
    client: AnthropicBedrock,
    *,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    max_tokens: int,
    verbose: bool,
    system_prompt: Optional[str],
    ref_preamble: str,
    pass_label: str,
    debug_job_id: Optional[str],
    debug_s3_bucket: Optional[str],
    debug_s3_prefix: Optional[str],
    user_msg: Optional[str] = None,
) -> dict:
    """Sparse-page retry: Nova Pro via Bedrock Converse (qa_verify_findings pattern)."""
    lbl = f"[{pass_label}] " if pass_label else ""
    if user_msg is not None:
        final_user_msg = user_msg
        input_chars = len(input_text or "")
        ref_chars = len(ref_text or "")
    else:
        ref_block = f"## REFERENCE ({ref_filename})\n"
        if ref_preamble:
            ref_block += ref_preamble.strip() + "\n"
        ref_block += ref_text
        input_block = f"## INPUT (מסמך המכרז)\n{input_text}"
        final_user_msg = f"{ref_block}\n\n{input_block}"
        input_chars = len(input_text)
        ref_chars = len(ref_text)

    nova_max = min(max_tokens, SPARSE_PAGE_RETRY_MAX_TOKENS, NOVA_MAX_OUTPUT_TOKENS)
    try:
        raw, resp = _nova_converse(system_prompt or SYSTEM_PROMPT, final_user_msg, nova_max)
    except Exception as exc:
        if verbose:
            if "Read timeout" in str(exc) or "ReadTimeoutError" in type(exc).__name__:
                print(
                    f"      {lbl}[WARN] Nova timed out after {NOVA_READ_TIMEOUT_SEC}s: {exc}"
                )
            else:
                print(f"      {lbl}[WARN] Nova call failed: {exc}")
        return {
            "findings": [],
            "output_truncated": False,
            "salvaged": False,
            "_llm_repaired": False,
            "parse_ok": False,
            "_api_failed": True,
            "stop_reason": None,
            "_raw": "",
            "output_tokens": 0,
        }

    usage = resp.get("usage") or {}
    input_tokens = usage.get("inputTokens")
    output_tokens = usage.get("outputTokens")
    stop_reason = resp.get("stopReason")
    output_truncated = stop_reason == "max_tokens"

    if verbose:
        print(
            f"      {lbl}[tokens] model={SPARSE_PAGE_RETRY_MODEL} ref={ref_filename} "
            f"input={input_tokens} output={output_tokens}"
        )

    findings, parse_ok, salvaged, llm_repaired, raw = _parse_and_recover_findings(
        client,
        raw,
        output_truncated=output_truncated,
        output_tokens=output_tokens,
        verbose=verbose,
        lbl=lbl,
        ref_filename=ref_filename,
        pass_label=pass_label,
        debug_job_id=debug_job_id,
        debug_s3_bucket=debug_s3_bucket,
        debug_s3_prefix=debug_s3_prefix,
    )

    if verbose:
        print(
            f"      {lbl}Nova ref={ref_filename} stop_reason={stop_reason} "
            f"input_chars={input_chars} ref_chars={ref_chars} "
            f"user_msg_chars={len(final_user_msg)} "
            f"output_tokens={output_tokens} {_findings_count_summary(findings)}"
        )
        if not findings or salvaged or output_truncated or not parse_ok:
            raw_preview = raw[:2000]
            truncated_note = "…(truncated)" if len(raw) > 2000 else ""
            print(f"      {lbl}RAW response (first 2000 chars):\n{raw_preview}{truncated_note}")

    if verbose and not findings:
        print(
            f"      {lbl}[DEBUG zero-findings] ref={ref_filename} "
            f"output_tokens={output_tokens} parse_ok={parse_ok} raw_len={len(raw)}"
        )
        print(
            f"      {lbl}[DEBUG zero-findings] INPUT snippet (first 400 chars):\n"
            f"{input_text[:400]}"
        )
        print(
            f"      {lbl}[DEBUG zero-findings] REF snippet (first 400 chars):\n"
            f"{ref_text[:400]}"
        )

    if output_truncated:
        print(
            f"[WARN] {lbl}Output truncated (max_tokens={nova_max}) for {ref_filename}"
        )

    return {
        "findings": findings,
        "output_truncated": output_truncated,
        "salvaged": salvaged,
        "_llm_repaired": llm_repaired,
        "parse_ok": parse_ok,
        "stop_reason": stop_reason,
        "_raw": raw,
        "output_tokens": output_tokens,
    }


def _run_nova_combined_pass(
    client: AnthropicBedrock,
    *,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    max_tokens: int,
    verbose: bool,
    multi_ref: bool,
    dual_batch_files: Optional[list[str]],
    ref_preamble: str,
    custom_prompt_addendum: str,
    pass_label: str,
    debug_job_id: Optional[str],
    debug_s3_bucket: Optional[str],
    debug_s3_prefix: Optional[str],
    user_msg: Optional[str] = None,
    chunk_rag: bool = False,
) -> dict:
    """Nova combined pass for sparse-page retry (same prompt + retry rules as Sonnet)."""
    empty_result = {
        "findings": [],
        "output_truncated": False,
        "salvaged": False,
        "_llm_repaired": False,
        "parse_ok": True,
        "stop_reason": None,
        "_raw": "",
        "output_tokens": 0,
    }
    if user_msg is None and not (input_text or "").strip():
        return empty_result
    if user_msg is not None and not user_msg.strip():
        return empty_result

    rag_addendum = SYSTEM_PROMPT_COMBINED_CHUNK_RAG if chunk_rag else ""
    merged_addendum = (custom_prompt_addendum or "") + rag_addendum

    sp = _system_prompt_for_compare(
        multi_ref=multi_ref,
        dual_batch_files=dual_batch_files,
        pass_type="combined",
        custom_prompt_addendum=merged_addendum,
    )
    result = _call_nova_combined(
        client,
        input_text=input_text,
        ref_text=ref_text,
        ref_filename=ref_filename,
        max_tokens=max_tokens,
        verbose=verbose,
        system_prompt=sp,
        ref_preamble=ref_preamble,
        pass_label=pass_label,
        debug_job_id=debug_job_id,
        debug_s3_bucket=debug_s3_bucket,
        debug_s3_prefix=debug_s3_prefix,
        user_msg=user_msg,
    )

    needs_retry, retry_reason = _claude_pass_needs_retry(result)
    if needs_retry:
        if verbose:
            print(f"      [{pass_label}] [retry] retrying once ({retry_reason})")
        result = _call_nova_combined(
            client,
            input_text=input_text,
            ref_text=ref_text,
            ref_filename=ref_filename,
            max_tokens=max_tokens,
            verbose=verbose,
            system_prompt=sp,
            ref_preamble=ref_preamble,
            pass_label=f"{pass_label}:retry",
            debug_job_id=debug_job_id,
            debug_s3_bucket=debug_s3_bucket,
            debug_s3_prefix=debug_s3_prefix,
            user_msg=user_msg,
        )
    return result


def _call_claude_three_pass_split_input(
    client: AnthropicBedrock,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    model: str,
    max_tokens: int = DEFAULT_MAX_TOKENS,
    verbose: bool = True,
    multi_ref: bool = False,
    dual_batch_files: Optional[list[str]] = None,
    ref_preamble: str = "",
    custom_prompt_addendum: str = "",
    input_chunks: Optional[list[dict]] = None,
    input_max_chars: int = DEFAULT_INPUT_MAX_CHARS,
    reference_batch_id: str = "Phoenix_reference1",
    debug_job_id: Optional[str] = None,
    debug_s3_bucket: Optional[str] = None,
    debug_s3_prefix: Optional[str] = None,
) -> dict:
    """
    Split INPUT into INPUT_SPLIT_PARTS overlapping windows and run one combined
    classify pass per window (missing / difference / match in a single JSON).

    When input_chunks is provided, windows are built directly from the chunk list
    (prev + group + next). Otherwise falls back to page-marker splitting on input_text.

    Total LLM calls: INPUT_SPLIT_LLM_CALLS (9 windows × 1 combined pass).
    First non-empty window runs sequentially to warm the Bedrock REFERENCE prompt
    cache; remaining windows run in parallel so they can cache_read the REF block.
    Completeness check receives the full original input_text.
    """
    windows = _build_input_overlap_windows(
        input_text, input_chunks, max_chars=input_max_chars
    )
    n_windows = len(windows)

    if verbose:
        non_empty = sum(1 for w in windows if (w.get("text") or "").strip())
        src = "chunks" if input_chunks else "pages"
        sizes = [len(w.get("text") or "") for w in windows]
        focus_info = [
            f"{w.get('focus_start')}-{w.get('focus_end')}"
            if w.get("focus_start") is not None else "—"
            for w in windows
        ]
        print(
            f"  [split-input] starting {INPUT_SPLIT_LLM_CALLS} LLM calls (combined pass) "
            f"ref={ref_filename} ref_chars={len(ref_text)}"
        )
        print(
            f"  [split-input] source={src} windows={non_empty}/{n_windows} "
            f"chars=[{', '.join(str(s) for s in sizes)}] "
            f"focus_chunks=[{', '.join(focus_info)}]"
        )

    def _run_one(win_idx: int) -> tuple[int, dict]:
        win = windows[win_idx]
        part = win.get("text") or ""
        focus_lbl = _window_focus_label(win)
        if not part.strip():
            if verbose:
                print(
                    f"      [split-input] skip combined:ninth{win_idx} "
                    f"({focus_lbl}, empty window)"
                )
            return win_idx, {
                "findings": [],
                "output_truncated": False,
                "salvaged": False,
                "_llm_repaired": False,
                "stop_reason": None,
                "_raw": "",
                "output_tokens": 0,
            }

        pass_label = f"combined:ninth{win_idx}"
        result = _run_combined_pass(
            client,
            input_text=part,
            ref_text=ref_text,
            ref_filename=ref_filename,
            model=model,
            max_tokens=max_tokens,
            verbose=verbose,
            multi_ref=multi_ref,
            dual_batch_files=dual_batch_files,
            ref_preamble=ref_preamble,
            custom_prompt_addendum=custom_prompt_addendum,
            pass_label=pass_label,
            debug_job_id=debug_job_id,
            debug_s3_bucket=debug_s3_bucket,
            debug_s3_prefix=debug_s3_prefix,
        )

        tagged = _tag_findings_with_focus(result.get("findings") or [], win)
        if verbose:
            flags: list[str] = []
            if result.get("output_truncated"):
                flags.append("truncated")
            if result.get("salvaged"):
                flags.append("salvaged")
            if result.get("_llm_repaired"):
                flags.append("llm_repaired")
            flag_str = f" [{', '.join(flags)}]" if flags else ""
            print(
                f"      [split-input] {pass_label} {focus_lbl} "
                f"input_chars={len(part)} "
                f"{_findings_count_summary(tagged)}{flag_str}"
            )
            # per-page findings count for this window
            focus_pages_set: set[int] = set()
            for f in tagged:
                p = _parse_input_page(f.get("input_page"))
                if p is not None:
                    focus_pages_set.add(p)
                else:
                    fallback = _parse_page_from_quote(
                        f.get("input_quote") or f.get("input_text") or ""
                    )
                    if fallback is not None:
                        focus_pages_set.add(fallback)
            if focus_pages_set:
                page_counts = _count_findings_by_page(tagged, focus_pages_set)
                page_summary = "  ".join(
                    f"p{p}:{page_counts[p]}" for p in sorted(page_counts)
                )
                print(f"      [split-input] {pass_label} per-page: {page_summary}")
        return win_idx, {**result, "findings": tagged}

    raw_results: dict[int, dict] = {}
    completed_jobs = 0

    if verbose:
        print(
            f"  [split-input] cache warmup: combined:ninth0 first (sequential) "
            f"to write cache, then {n_windows - 1} window(s) in parallel (cache_read)"
        )

    # Run window 0 alone first so it writes the prompt+reference cache
    idx0, result0 = _run_one(0)
    raw_results[idx0] = result0
    completed_jobs += 1
    if verbose:
        n = len(result0.get("findings") or [])
        print(
            f"  [split-input] progress {completed_jobs}/{n_windows} "
            f"(combined:ninth{idx0} finished, findings={n})"
        )

    # Remaining windows in parallel — they cache_read the prefix written above
    if n_windows > 1:
        with ThreadPoolExecutor(max_workers=INPUT_SPLIT_MAX_WORKERS) as pool:
            futures = {pool.submit(_run_one, i): i for i in range(1, n_windows)}
            for fut in as_completed(futures):
                idx, result = fut.result()
                raw_results[idx] = result
                completed_jobs += 1
                if verbose:
                    n = len(result.get("findings") or [])
                    print(
                        f"  [split-input] progress {completed_jobs}/{n_windows} "
                        f"(combined:ninth{idx} finished, findings={n})"
                    )

    all_findings: list[dict] = []
    output_truncated = False
    salvaged_any = False
    window_findings_docs: list[dict] = []
    if verbose:
        print("  [split-input] per-window findings (before dedupe):")
    for i in range(n_windows):
        win = windows[i]
        focus_lbl = _window_focus_label(win)
        findings_i = raw_results[i].get("findings") or []
        r = raw_results[i]
        window_findings_docs.append({
            "window_index": i,
            "window_label": f"ninth{i}",
            "focus_chunks": focus_lbl,
            "focus_start": win.get("focus_start"),
            "focus_end": win.get("focus_end"),
            "findings_count": len(findings_i),
            "summary": _findings_count_summary(findings_i),
            "output_truncated": bool(r.get("output_truncated")),
            "salvaged": bool(r.get("salvaged")),
            "findings": findings_i,
        })
        if verbose:
            print(
                f"      ninth{i} ({focus_lbl}): "
                f"{len(findings_i)} ({_findings_count_summary(findings_i)})"
            )
        all_findings.extend(findings_i)
        if r.get("output_truncated"):
            output_truncated = True
        if r.get("salvaged"):
            salvaged_any = True

    if verbose:
        print(
            f"  [split-input] total before dedupe: {len(all_findings)} "
            f"({_findings_count_summary(all_findings)})"
        )

    before_dedupe = len(all_findings)
    all_findings, dedupe_dropped = _dedupe_findings(all_findings, verbose=verbose)
    if verbose and before_dedupe != len(all_findings):
        print(
            f"  [split-input] after window dedupe: {before_dedupe} -> {len(all_findings)} "
            f"({_findings_count_summary(all_findings)}) dropped={len(dedupe_dropped)}"
        )

    page_retry_docs: list[dict] = []
    sparse_pages_retried: list[int] = []
    extra_llm_calls = 0

    if input_chunks:
        pages_by_num = _group_chunks_by_page(input_chunks)
        page_counts = _count_findings_by_page(all_findings, pages_by_num.keys())
        sparse_pages = sorted(p for p, n in page_counts.items() if n == 0)

        if sparse_pages:
            if verbose:
                sparse_summary = ", ".join(
                    f"p{p}={page_counts[p]}" for p in sparse_pages[:20]
                )
                suffix = "…" if len(sparse_pages) > 20 else ""
                print(
                    f"  [sparse-page-retry] {len(sparse_pages)} pages with "
                    f"0 findings: {sparse_summary}{suffix}"
                )
                print(
                    f"  [sparse-page-retry] model={SPARSE_PAGE_RETRY_MODEL} "
                    f"region={SPARSE_PAGE_RETRY_REGION} "
                    f"pinecone_batch={reference_batch_id} "
                    f"top_k={SPARSE_CHUNK_RETRY_TOP_K_REF} "
                    f"max_workers={SPARSE_PAGE_RETRY_MAX_WORKERS}"
                )

            retry_jobs: list[tuple[int, int, dict, int]] = []
            for page_num in sparse_pages:
                page_chunks = pages_by_num.get(page_num, [])
                for chunk_idx, chunk in enumerate(page_chunks):
                    retry_jobs.append((page_num, chunk_idx, chunk, len(page_chunks)))

            def _retry_chunk(
                job: tuple[int, int, dict, int],
            ) -> tuple[int, int, dict, dict]:
                page_num, chunk_idx, chunk, chunks_on_page = job
                part, chunk_meta = _build_sparse_chunk_retry_subset(
                    page_num,
                    chunk_idx,
                    pages_by_num.get(page_num, []),
                    pages_by_num,
                    max_chars=input_max_chars,
                )
                ref_matches = _query_ref_chunks_for_input_chunk(
                    chunk,
                    reference_batch_id,
                    SPARSE_CHUNK_RETRY_TOP_K_REF,
                    verbose=verbose,
                )
                use_fallback = not ref_matches
                if use_fallback and verbose:
                    print(
                        f"  [sparse-page-retry] page {page_num} chunk "
                        f"{chunk_idx + 1}/{chunks_on_page}: 0 Pinecone hits "
                        f"-> full ref_text fallback"
                    )
                user_msg = _build_nova_chunk_user_message(
                    part,
                    ref_matches,
                    ref_filename=ref_filename,
                    ref_text=ref_text,
                    ref_preamble=ref_preamble,
                    use_full_ref_fallback=use_fallback,
                )
                chunk_idx_val = chunk_meta.get("chunk_index", chunk_idx)
                if verbose:
                    print(
                        f"  [sparse-page-retry] page {page_num} chunk "
                        f"{chunk_idx + 1}/{chunks_on_page} (idx={chunk_idx_val}): "
                        f"Nova {SPARSE_PAGE_RETRY_MODEL} "
                        f"pinecone_hits={len(ref_matches)} "
                        f"fallback={use_fallback} "
                        f"chars={chunk_meta.get('char_count', len(part))}"
                    )
                result = _run_nova_combined_pass(
                    client,
                    input_text=part,
                    ref_text=ref_text,
                    ref_filename=ref_filename,
                    max_tokens=max_tokens,
                    verbose=verbose,
                    multi_ref=multi_ref,
                    dual_batch_files=dual_batch_files,
                    ref_preamble=ref_preamble,
                    custom_prompt_addendum=custom_prompt_addendum,
                    pass_label=(
                        f"combined:page_retry:{page_num}:chunk{chunk_idx_val}"
                    ),
                    debug_job_id=debug_job_id,
                    debug_s3_bucket=debug_s3_bucket,
                    debug_s3_prefix=debug_s3_prefix,
                    user_msg=user_msg,
                    chunk_rag=not use_fallback,
                )
                tagged = [
                    {
                        **f,
                        "_retry_page": page_num,
                        "_retry_chunk_index": chunk_idx_val,
                        "_pinecone_ref_count": len(ref_matches),
                    }
                    for f in (result.get("findings") or [])
                ]
                if verbose:
                    print(
                        f"  [sparse-page-retry] page {page_num} chunk "
                        f"{chunk_idx + 1}/{chunks_on_page}: "
                        f"+{len(tagged)} findings "
                        f"({_findings_count_summary(tagged)})"
                    )
                chunk_doc = {
                    **chunk_meta,
                    "pinecone_hits": len(ref_matches),
                    "used_full_ref_fallback": use_fallback,
                    "findings_count": len(tagged),
                    "summary": _findings_count_summary(tagged),
                    "output_truncated": bool(result.get("output_truncated")),
                    "salvaged": bool(result.get("salvaged")),
                    "findings": tagged,
                }
                return page_num, chunk_idx, {**result, "findings": tagged}, chunk_doc

            page_agg: dict[int, dict] = {}

            with ThreadPoolExecutor(max_workers=SPARSE_PAGE_RETRY_MAX_WORKERS) as pool:
                futures = {pool.submit(_retry_chunk, job): job for job in retry_jobs}
                for fut in as_completed(futures):
                    page_num, _chunk_idx, result, chunk_doc = fut.result()
                    extra_llm_calls += 1
                    agg = page_agg.setdefault(
                        page_num,
                        {
                            "page_number": page_num,
                            "model": SPARSE_PAGE_RETRY_MODEL,
                            "findings_before_retry": page_counts[page_num],
                            "chunk_count": 0,
                            "chunk_retries": [],
                            "llm_calls": 0,
                            "findings": [],
                            "output_truncated": False,
                            "salvaged": False,
                        },
                    )
                    agg["chunk_retries"].append(chunk_doc)
                    agg["llm_calls"] += 1
                    agg["chunk_count"] = len(agg["chunk_retries"])
                    findings_i = result.get("findings") or []
                    agg["findings"].extend(findings_i)
                    if result.get("output_truncated"):
                        agg["output_truncated"] = True
                    if result.get("salvaged"):
                        agg["salvaged"] = True

            for page_num in sorted(page_agg):
                agg = page_agg[page_num]
                sparse_pages_retried.append(page_num)
                findings_i = agg["findings"]
                page_retry_docs.append({
                    "page_number": page_num,
                    "model": SPARSE_PAGE_RETRY_MODEL,
                    "findings_before_retry": agg["findings_before_retry"],
                    "chunk_count": agg["chunk_count"],
                    "llm_calls": agg["llm_calls"],
                    "chunk_retries": agg["chunk_retries"],
                    "findings_count": len(findings_i),
                    "summary": _findings_count_summary(findings_i),
                    "output_truncated": agg["output_truncated"],
                    "salvaged": agg["salvaged"],
                    "findings": findings_i,
                })
                all_findings.extend(findings_i)
                if agg["output_truncated"]:
                    output_truncated = True
                if agg["salvaged"]:
                    salvaged_any = True
                if verbose:
                    print(
                        f"  [sparse-page-retry] page {page_num} done: "
                        f"{agg['llm_calls']} chunk(s), "
                        f"+{len(findings_i)} findings "
                        f"({_findings_count_summary(findings_i)})"
                    )

            if extra_llm_calls:
                pre_final_dedupe = len(all_findings)
                all_findings, dedupe_dropped_retry = _dedupe_findings(
                    all_findings, verbose=verbose
                )
                dedupe_dropped.extend(dedupe_dropped_retry)
                if verbose:
                    print(
                        f"  [sparse-page-retry] done: {extra_llm_calls} chunk call(s) "
                        f"across {len(sparse_pages_retried)} page(s), "
                        f"{pre_final_dedupe} -> {len(all_findings)} after dedupe "
                        f"({_findings_count_summary(all_findings)})"
                    )
                before_dedupe = pre_final_dedupe
        elif verbose:
            print(
                "  [sparse-page-retry] none needed (no pages with 0 findings)"
            )

    merged, suppressed = _apply_output_guards(all_findings, verbose=verbose)

    # Completeness check disabled — 4th LLM pass on full INPUT to catch clauses missed
    # across all windows. Re-enable if windowed scans leave gaps.
    # completeness_findings = _call_claude_completeness_check(
    #     client,
    #     input_text=input_text,
    #     ref_text=ref_text,
    #     ref_filename=ref_filename,
    #     model=model,
    #     merged_findings=merged,
    #     verbose=verbose,
    # )
    # if completeness_findings:
    #     if verbose:
    #         print(
    #             f"  [completeness-check] adding {len(completeness_findings)} new finding(s) "
    #             f"({_findings_count_summary(completeness_findings)}) to merged list"
    #         )
    #     merged = merged + completeness_findings

    stop_reason = raw_results.get(0, {}).get("stop_reason")

    if verbose:
        print(
            f"  [split-input] FINAL ref={ref_filename} "
            f"merged={len(merged)} ({_findings_count_summary(merged)}) "
            f"suppressed={len(suppressed)} "
            f"truncated={output_truncated} salvaged={salvaged_any} "
            f"stop_reason={stop_reason}"
        )

    pass_findings_all = _split_findings_by_classification(merged)

    return {
        "findings": merged,
        "output_truncated": output_truncated,
        "salvaged": salvaged_any,
        "stop_reason": stop_reason,
        "_pass_findings": pass_findings_all,
        "_suppressed": suppressed,
        "_dedupe_dropped": dedupe_dropped,
        "_dedupe_before_count": before_dedupe,
        "_dedupe_after_count": len(all_findings),
        "_window_findings": window_findings_docs,
        "_page_retry_findings": page_retry_docs,
        "_sparse_pages_retried": sparse_pages_retried,
        "_extra_llm_calls": extra_llm_calls,
    }


def _repair_json_with_llm(
    client: AnthropicBedrock,
    broken_text: str,
    verbose: bool = True,
) -> Optional[str]:
    """
    Ask a cheap model to extract the valid JSON object from a noisy LLM response.
    Returns the repaired JSON string (ready for json.loads), or None on failure.
    Only called as a last resort when all deterministic parse attempts have failed.
    """
    system = (
        "You are a JSON extraction assistant. "
        "Your only job is to extract a valid JSON object from text that may contain prose."
    )
    user = (
        'The text below should contain a JSON object with a "findings" key. '
        "Extract and return ONLY that JSON object — no explanation, no markdown, no prose. "
        "The first character of your response must be { and the last must be }.\n\n"
        "Repair rules:\n"
        '- Escape every double-quote inside string values as \\" (Hebrew gershayim: הנ\\"ל, ע\\"י, דו\\"ח, תשמ\\"א, מצ\\"ב).\n'
        "- Do NOT add extra backslashes except before \" for gershayim; fix invalid \\\\ sequences.\n"
        "- Remove illegal escapes (e.g. backslash before space or parenthesis).\n"
        "- Use JSON literals true, false, null — not Python True/False/None.\n"
        '- Return only {"findings": [...]}.\n\n'
        f"Text:\n{broken_text}"
    )
    try:
        resp = client.messages.create(
            model=_JSON_REPAIR_MODEL,
            max_tokens=8192,
            system=system,
            messages=[{"role": "user", "content": user}],
        )
        repaired = resp.content[0].text.strip()
        repaired_data, parse_ok = _parse_llm_json_object(repaired)
        if not parse_ok:
            raise ValueError("repaired JSON still invalid")
        return repaired
    except Exception as exc:
        if verbose:
            print(f"      [WARN] LLM JSON repair failed: {exc}")
        return None


def _get_nova_bedrock_client():
    """Lazy boto3 bedrock-runtime client for sparse-page Nova retries."""
    global _NOVA_BEDROCK_CLIENT
    if _NOVA_BEDROCK_CLIENT is None:
        _NOVA_BEDROCK_CLIENT = boto3.client(
            "bedrock-runtime",
            region_name=SPARSE_PAGE_RETRY_REGION,
            config=Config(
                read_timeout=NOVA_READ_TIMEOUT_SEC,
                connect_timeout=NOVA_CONNECT_TIMEOUT_SEC,
                retries={"max_attempts": 1},
            ),
        )
    return _NOVA_BEDROCK_CLIENT


def _nova_converse(
    system_prompt: str,
    user_msg: str,
    max_tokens: int,
) -> tuple[str, dict]:
    """
    Call Nova via Bedrock Converse API — same pattern as qa_verify_findings._converse.
    Returns (combined_text, full_response_dict).
    """
    resp = _get_nova_bedrock_client().converse(
        modelId=SPARSE_PAGE_RETRY_MODEL,
        system=[{"text": system_prompt}],
        messages=[{"role": "user", "content": [{"text": user_msg}]}],
        inferenceConfig={"maxTokens": max_tokens, "temperature": 0},
    )
    raw = ""
    for block in resp.get("output", {}).get("message", {}).get("content", []):
        if isinstance(block, dict) and "text" in block:
            raw += block["text"]
    return raw.strip(), resp


def _extract_json_from_llm_text(raw: str) -> str:
    fence_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if fence_match:
        return fence_match.group(1).strip()
    return raw


def _parse_and_recover_findings(
    client: AnthropicBedrock,
    raw: str,
    *,
    output_truncated: bool,
    output_tokens: Optional[int],
    verbose: bool,
    lbl: str,
    ref_filename: str,
    pass_label: str,
    debug_job_id: Optional[str],
    debug_s3_bucket: Optional[str],
    debug_s3_prefix: Optional[str],
) -> tuple[list[dict], bool, bool, bool, str]:
    """Parse LLM JSON and run repair/salvage. Returns findings + metadata + stripped raw."""
    raw = _extract_json_from_llm_text(raw)
    findings: list[dict] = []
    data, parse_ok = _parse_llm_json_object(raw)
    if data is not None:
        findings = _parse_findings_list(data)

    salvaged = False
    llm_repaired = False

    if not findings and (output_truncated or not parse_ok):
        if not parse_ok and output_tokens and output_tokens > 300:
            repaired_raw = _repair_json_with_llm(client, raw, verbose)
            if repaired_raw:
                repaired_data, _ = _parse_llm_json_object(repaired_raw)
                if repaired_data is not None:
                    repaired_findings = _parse_findings_list(repaired_data)
                    if repaired_findings:
                        findings = repaired_findings
                        salvaged = True
                        llm_repaired = True
                        if verbose:
                            print(
                                f"      {lbl}[WARN] Recovered {len(findings)} findings via LLM JSON repair "
                                f"({ref_filename})"
                            )

        if not findings:
            salvaged_rows = _salvage_findings_from_truncated_json(raw)
            if salvaged_rows:
                findings = salvaged_rows
                salvaged = True
                if verbose:
                    reason = (
                        "truncated output" if output_truncated else "malformed/incomplete JSON"
                    )
                    print(
                        f"      {lbl}[WARN] Recovered {len(findings)} findings via salvage "
                        f"({reason}) ({ref_filename})"
                    )

    if not parse_ok or salvaged or output_truncated:
        _save_broken_llm_json_to_s3(
            raw,
            pass_label=pass_label,
            ref_filename=ref_filename,
            job_id=debug_job_id,
            s3_bucket=debug_s3_bucket,
            s3_prefix=debug_s3_prefix,
        )

    return findings, parse_ok, salvaged, llm_repaired, raw


def _call_claude(
    client: AnthropicBedrock,
    input_text: str,
    ref_text: str,
    ref_filename: str,
    model: str,
    max_tokens: int = DEFAULT_MAX_TOKENS,
    verbose: bool = True,
    system_prompt: Optional[str] = None,
    ref_preamble: str = "",
    pass_label: str = "",
    debug_job_id: Optional[str] = None,
    debug_s3_bucket: Optional[str] = None,
    debug_s3_prefix: Optional[str] = None,
) -> dict:
    """Call Claude on Bedrock; return findings plus truncation metadata."""
    lbl = f"[{pass_label}] " if pass_label else ""

    ref_block = f"## REFERENCE ({ref_filename})\n"
    if ref_preamble:
        ref_block += ref_preamble.strip() + "\n"
    ref_block += ref_text
    input_block = f"## INPUT (מסמך המכרז)\n{input_text}"

    if verbose:
        print(f"      {lbl}start")
    t0 = time.perf_counter()
    with client.messages.stream(
        model=model,
        max_tokens=max_tokens,
        system=system_prompt or SYSTEM_PROMPT,
        messages=[{"role": "user", "content": [
            {"type": "text", "text": ref_block, "cache_control": {"type": "ephemeral"}},
            {"type": "text", "text": input_block},
        ]}],
    ) as stream:
        response = stream.get_final_message()
    if verbose:
        elapsed = time.perf_counter() - t0
        print(f"      {lbl}end elapsed_sec={elapsed:.1f}")
    stop_reason = getattr(response, "stop_reason", None)
    usage = getattr(response, "usage", None)
    input_tokens               = getattr(usage, "input_tokens",               None) if usage else None
    output_tokens              = getattr(usage, "output_tokens",              None) if usage else None
    cache_creation_tokens      = getattr(usage, "cache_creation_input_tokens", None) if usage else None
    cache_read_tokens          = getattr(usage, "cache_read_input_tokens",     None) if usage else None

    if verbose:
        cache_status = (
            f"cache_creation={cache_creation_tokens} cache_read={cache_read_tokens}"
            if (cache_creation_tokens or cache_read_tokens)
            else "cache=off"
        )
        print(
            f"      {lbl}[tokens] ref={ref_filename} "
            f"input={input_tokens} output={output_tokens} {cache_status}"
        )

    raw = _extract_text_from_response(response)

    findings, parse_ok, salvaged, llm_repaired, raw = _parse_and_recover_findings(
        client,
        raw,
        output_truncated=stop_reason == "max_tokens",
        output_tokens=output_tokens,
        verbose=verbose,
        lbl=lbl,
        ref_filename=ref_filename,
        pass_label=pass_label,
        debug_job_id=debug_job_id,
        debug_s3_bucket=debug_s3_bucket,
        debug_s3_prefix=debug_s3_prefix,
    )
    output_truncated = stop_reason == "max_tokens"

    if verbose:
        print(
            f"      {lbl}Claude ref={ref_filename} stop_reason={stop_reason} "
            f"input_chars={len(input_text)} ref_chars={len(ref_text)} "
            f"output_tokens={output_tokens} {_findings_count_summary(findings)}"
        )
        if not findings or salvaged or output_truncated or not parse_ok:
            raw_preview = raw[:2000]
            truncated_note = "…(truncated)" if len(raw) > 2000 else ""
            print(f"      {lbl}RAW response (first 2000 chars):\n{raw_preview}{truncated_note}")

    # Diagnostic: print full context whenever a pass returns zero findings so the cause
    # can be diagnosed from logs (input empty, ref empty, Claude refused, etc.).
    if verbose and not findings:
        print(
            f"      {lbl}[DEBUG zero-findings] ref={ref_filename} "
            f"output_tokens={output_tokens} parse_ok={parse_ok} raw_len={len(raw)}"
        )
        print(
            f"      {lbl}[DEBUG zero-findings] INPUT snippet (first 400 chars):\n"
            f"{input_text[:400]}"
        )
        print(
            f"      {lbl}[DEBUG zero-findings] REF snippet (first 400 chars):\n"
            f"{ref_text[:400]}"
        )

    if output_truncated:
        print(
            f"[WARN] {lbl}Output truncated (max_tokens={max_tokens}) for {ref_filename}; "
            "consider raising max_tokens on /diff"
        )

    return {
        "findings": findings,
        "output_truncated": output_truncated,
        "salvaged": salvaged,
        "_llm_repaired": llm_repaired,
        "parse_ok": parse_ok,
        "stop_reason": stop_reason,
        "_raw": raw,   # internal: used by retry logic in _call_claude_three_pass
        "output_tokens": output_tokens,
    }


def _parse_findings_list(data: dict) -> list[dict]:
    """Accept findings (new) or differences (legacy) from model JSON."""
    if not isinstance(data, dict):
        return []
    rows = data.get("findings") or data.get("differences") or []
    return rows if isinstance(rows, list) else []


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HIGH_FILL   = PatternFill("solid", start_color="FFB3B3")
_MED_FILL    = PatternFill("solid", start_color="FFE0A0")
_LOW_FILL    = PatternFill("solid", start_color="D6EAF8")
_WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
_ALT_FILL    = PatternFill("solid", start_color="F5F5F5")

_THIN  = Side(style="thin", color="AAAAAA")
_BRD   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_SEV_FILL = {"גבוהה": _HIGH_FILL, "בינונית": _MED_FILL, "נמוכה": _LOW_FILL}
_SEV_COL  = {"גבוהה": "CC0000",   "בינונית": "7D4600",   "נמוכה": "1A5276"}


def _hdr_cell(ws, row: int, col: int, val: str) -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _BRD


def _data_cell(
    ws,
    row: int,
    col: int,
    val: str,
    fill: PatternFill,
    bold: bool = False,
    color: str = "000000",
    center: bool = False,
) -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, color=color, size=9)
    c.fill = fill
    c.alignment = Alignment(
        horizontal="center" if center else "right",
        vertical="top",
        wrap_text=True,
    )
    c.border = _BRD


_COLUMNS = [
    ("#",                                 4),
    ("מסמך רפרנס",                       28),
    ("פרק / סעיף",                       26),
    ("ציטוט מ-Input",                    52),
    ("עמוד ב-Input",                     12),
    ("ציטוט מרפרנס",                     52),
    ("עמוד ברפרנס",                      12),
    ("סוג הבדל",                         18),
    ("חומרה",                            10),
    ("הערות",                            40),
]


def _build_excel(
    all_diffs: list[dict],   # each dict has "ref_file" + fields from Claude
    output_path: str | Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "הבדלים"
    ws.sheet_view.rightToLeft = True

    # Headers
    for ci, (label, _) in enumerate(_COLUMNS, 1):
        _hdr_cell(ws, 1, ci, label)
    ws.row_dimensions[1].height = 36

    # Data rows
    for i, d in enumerate(all_diffs):
        row  = i + 2
        sev  = d.get("severity", "נמוכה")
        sev_fill = _SEV_FILL.get(sev, _LOW_FILL)
        row_fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        ws.row_dimensions[row].height = 90

        _data_cell(ws, row, 1, str(i + 1),                       row_fill, bold=True, center=True)
        _data_cell(ws, row, 2, d.get("ref_file", ""),             row_fill)
        _data_cell(ws, row, 3, d.get("section", ""),              row_fill)
        _data_cell(ws, row, 4, d.get("input_quote", ""),          row_fill)
        _data_cell(ws, row, 5, d.get("input_page", ""),           row_fill, center=True)
        _data_cell(ws, row, 6, d.get("ref_quote", ""),            row_fill)
        _data_cell(ws, row, 7, d.get("ref_page", ""),             row_fill, center=True)
        _data_cell(ws, row, 8, d.get("diff_type", ""),            row_fill)
        _data_cell(ws, row, 9, sev, sev_fill,
                   bold=True, color=_SEV_COL.get(sev, "000000"), center=True)
        _data_cell(ws, row, 10, d.get("note", ""),                row_fill)

    # Column widths
    for ci, (_, width) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = "A2"

    # Legend sheet
    ls = wb.create_sheet("מקרא")
    ls.sheet_view.rightToLeft = True
    ls["A1"] = "מקרא חומרת הבדל"
    ls["A1"].font = Font(name="Arial", bold=True, size=13)

    legend = [
        ("גבוהה",   _HIGH_FILL, "CC0000", "הבדל מהותי – שינוי כספי / כיסויי משמעותי"),
        ("בינונית", _MED_FILL,  "7D4600", "הבדל משמעותי – ניסוח בעל השלכות פרשניות"),
        ("נמוכה",   _LOW_FILL,  "1A5276", "הבדל קל – שינוי שולי / כיסוי לא רלוונטי"),
    ]
    for ri, (label, fill, tc, desc) in enumerate(legend, 3):
        c = ls.cell(row=ri, column=1, value=label)
        c.font = Font(name="Arial", bold=True, color=tc, size=11)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        c.border = _BRD
        d2 = ls.cell(row=ri, column=2, value=desc)
        d2.font = Font(name="Arial", size=10)
        d2.alignment = Alignment(wrap_text=True)
    ls.column_dimensions["A"].width = 15
    ls.column_dimensions["B"].width = 55

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def analyze_insurance_docs(
    input_files: Sequence[str | Path],
    reference_files: Sequence[str | Path],
    output_xlsx: str | Path = "insurance_gaps.xlsx",
    model: str = "eu.anthropic.claude-sonnet-4-6",
    aws_region: str = "us-east-1",
    aws_access_key: str | None = None,
    aws_secret_key: str | None = None,
    aws_session_token: str | None = None,
    input_max_chars: int = DEFAULT_INPUT_MAX_CHARS,
    ref_max_chars: int = DEFAULT_REF_MAX_CHARS,
    max_tokens: int = DEFAULT_MAX_TOKENS,
    reference_strategy: str = "auto",
    ref_bundle_single_max_chars: int = REF_BUNDLE_SINGLE_MAX_CHARS,
    ref_bundle_batch_max_chars: int = REF_BUNDLE_BATCH_MAX_CHARS,
    verbose: bool = True,
    combined_input_override: Optional[str] = None,
    ref_measurements_override: Optional[list] = None,
    input_chunks_override: Optional[list[dict]] = None,
    custom_prompt_addendum: str = "",
    reference_batch_id: str = "Phoenix_reference1",
    debug_job_id: Optional[str] = None,
    debug_s3_bucket: Optional[str] = None,
    debug_s3_prefix: Optional[str] = None,
) -> dict:
    """
    Compare insurance spec PDFs and write a gap-analysis Excel workbook.

    Uses AWS Bedrock — no Anthropic API key required. AWS credentials are
    resolved from the standard chain (env vars, ~/.aws/credentials, IAM role)
    unless explicitly provided via aws_access_key / aws_secret_key.

    Parameters
    ----------
    input_files       : One or more input files (.pdf, .doc, .docx).
                        Their text is concatenated into a single input document.
    reference_files   : One or more reference files (.pdf, .doc, .docx).
                        Strategy auto: one bundled call (small multi-ref),
                        two batched calls, or per-file loop if too large.
    reference_strategy: auto | single | dual | per_file
    ref_bundle_single_max_chars : Max total ref chars for single bundled call.
    ref_bundle_batch_max_chars  : Max ref chars per batch in dual mode.
    output_xlsx       : Destination path for the Excel workbook.
    model             : Bedrock model ID. Defaults to "eu.anthropic.claude-sonnet-4-6".
                        Other options:
                          "global.anthropic.claude-sonnet-4-6"  (global endpoint)
                          "us.anthropic.claude-sonnet-4-6"      (US regional)
                          "eu.anthropic.claude-sonnet-4-6"      (EU regional)
    aws_region        : AWS region for Bedrock (default: "us-east-1").
    aws_access_key    : Optional. If omitted, uses the default AWS credential chain.
    aws_secret_key    : Optional. Required if aws_access_key is provided.
    aws_session_token : Optional. For temporary/STS credentials.
    input_max_chars   : Character cap for the concatenated input text sent to Claude.
    ref_max_chars     : Character cap for each reference document sent to Claude.
    max_tokens        : Bedrock max output tokens per reference comparison (default 64000).
    verbose           : Print progress messages.
    combined_input_override      : Pre-built input text string; skips PDF extraction (step 1).
    ref_measurements_override    : Pre-built reference measurements list; skips PDF extraction (step 2).
    input_chunks_override        : Pre-extracted input chunks list; used for 9-way overlap splitting.

    Returns
    -------
    dict with output_path, differences, reference_strategy, llm_calls.
    """
    # Build client — only pass explicit credentials if provided
    client_kwargs: dict = {"aws_region": aws_region}
    if aws_access_key:
        client_kwargs["aws_access_key"] = aws_access_key
    if aws_secret_key:
        client_kwargs["aws_secret_key"] = aws_secret_key
    if aws_session_token:
        client_kwargs["aws_session_token"] = aws_session_token

    client = AnthropicBedrock(**client_kwargs)

    # ------------------------------------------------------------------
    # 1. Build combined input text
    # ------------------------------------------------------------------
    if combined_input_override is not None:
        combined_input = combined_input_override
        if verbose:
            print(
                f"[1/3] Using pre-extracted input text from chunks: "
                f"sent_chars={len(combined_input)} cap={input_max_chars}"
            )
    else:
        if verbose:
            print(f"[1/3] Extracting text from {len(input_files)} input file(s)...")

        combined_input_parts = []
        extracted_input_chars = 0
        for fpath in input_files:
            pages = _extract_document_pages(fpath)
            fname = Path(fpath).name
            extracted_input_chars += sum(len(t) for _, t in pages)
            combined_input_parts.append(f"\n\n=== קובץ input: {fname} ===\n")
            combined_input_parts.append(_pages_to_str(pages, max_chars=input_max_chars))

        combined_input = "".join(combined_input_parts)[:input_max_chars]
        if verbose:
            print(
                f"  [INPUT] extracted_chars≈{extracted_input_chars} "
                f"sent_chars={len(combined_input)} cap={input_max_chars}"
            )
        _log_prompt_text_stats("INPUT", combined_input, input_max_chars, verbose=verbose)

    # ------------------------------------------------------------------
    # 2. Compare references against combined input (bundled / dual / per-file)
    # ------------------------------------------------------------------
    if ref_measurements_override is not None:
        measurements = ref_measurements_override
    else:
        measurements = _measure_reference_corpus(reference_files)
    total_ref_chars = sum(m["chars"] for m in measurements)
    total_ref_pages = sum(m["pages"] for m in measurements)
    strategy = _choose_reference_strategy(
        measurements,
        reference_strategy,
        single_max=ref_bundle_single_max_chars,
        batch_max=ref_bundle_batch_max_chars,
    )

    if verbose:
        source_label = "pre-extracted chunks" if ref_measurements_override is not None else "PDF files"
        print(
            f"[2/3] reference corpus ({source_label}): files={len(measurements)} "
            f"total_chars≈{total_ref_chars} total_pages≈{total_ref_pages} "
            f"strategy={strategy}"
        )
        for m in measurements:
            print(f"      {m['name']}: chars≈{m['chars']} pages={m['pages']}")

    all_diffs: list[dict] = []
    all_pass_findings: dict[str, list[dict]] = {"missing": [], "difference": [], "match": []}
    all_suppressed: list[dict] = []
    all_dedupe_dropped: list[dict] = []
    all_window_findings: list[dict] = []
    all_page_retry_findings: list[dict] = []
    dedupe_before_count = 0
    dedupe_after_count = 0
    llm_calls = 0
    output_truncated = False
    salvaged_any = False

    def _consume_call(call_result: dict, *, label: str = "") -> list[dict]:
        nonlocal output_truncated, salvaged_any, dedupe_before_count, dedupe_after_count, llm_calls
        llm_calls += int(call_result.get("_extra_llm_calls") or 0)
        if call_result.get("output_truncated"):
            output_truncated = True
        if call_result.get("salvaged"):
            salvaged_any = True
        # Accumulate raw per-pass findings before merge
        for pass_name in ("missing", "difference", "match"):
            all_pass_findings[pass_name].extend(
                call_result.get("_pass_findings", {}).get(pass_name, [])
            )
        # Accumulate suppressed findings
        all_suppressed.extend(call_result.get("_suppressed", []))
        all_dedupe_dropped.extend(call_result.get("_dedupe_dropped", []))
        for win_doc in call_result.get("_window_findings") or []:
            all_window_findings.append({**win_doc, "call_label": label})
        for page_doc in call_result.get("_page_retry_findings") or []:
            all_page_retry_findings.append({**page_doc, "call_label": label})
        if call_result.get("_dedupe_before_count") is not None:
            dedupe_before_count += int(call_result["_dedupe_before_count"])
        if call_result.get("_dedupe_after_count") is not None:
            dedupe_after_count += int(call_result["_dedupe_after_count"])
        findings = call_result.get("findings") or []
        if verbose and label:
            extra = int(call_result.get("_extra_llm_calls") or 0)
            extra_note = f" page_retries={extra}" if extra else ""
            print(
                f"  [compare] {label} returned {len(findings)} merged finding(s) "
                f"({_findings_count_summary(findings)}) "
                f"suppressed={len(call_result.get('_suppressed') or [])} "
                f"dedupe_dropped={len(call_result.get('_dedupe_dropped') or [])}"
                f"{extra_note}"
            )
        return findings

    if strategy == "single":
        bundle, included = _build_reference_bundle(
            measurements, ref_bundle_single_max_chars, verbose=verbose
        )
        multi_ref = len(measurements) > 1
        ref_label = measurements[0]["name"] if len(measurements) == 1 else "all_references"
        if verbose:
            print(
                f"  [single] bundled {len(included)}/{len(measurements)} files, "
                f"sent_chars={len(bundle)}"
            )
        _log_prompt_text_stats("REFERENCE_BUNDLE", bundle, ref_bundle_single_max_chars, verbose=verbose)

        call_result = _call_claude_three_pass_split_input(
                client,
                input_text=combined_input,
                ref_text=bundle,
                ref_filename=ref_label,
                model=model,
                max_tokens=max_tokens,
                verbose=verbose,
                multi_ref=multi_ref,
                custom_prompt_addendum=custom_prompt_addendum,
                input_chunks=input_chunks_override,
                input_max_chars=input_max_chars,
                reference_batch_id=reference_batch_id,
                debug_job_id=debug_job_id,
                debug_s3_bucket=debug_s3_bucket,
                debug_s3_prefix=debug_s3_prefix,
            )
        llm_calls += INPUT_SPLIT_LLM_CALLS
        diffs = _consume_call(call_result, label=f"single/{ref_label}")
        _ensure_ref_file_on_findings(diffs, measurements[0]["name"] if len(measurements) == 1 else "")
        all_diffs.extend(diffs)

    elif strategy == "dual":
        batch_a, batch_b = _split_reference_batches(measurements, ref_bundle_batch_max_chars)
        if verbose:
            print(
                f"  [dual] batch_a={[m['name'] for m in batch_a]} "
                f"batch_b={[m['name'] for m in batch_b]}"
            )

        findings_a: list[dict] = []
        findings_b: list[dict] = []

        for batch, label, acc in (
            (batch_a, "batch_a", findings_a),
            (batch_b, "batch_b", findings_b),
        ):
            if not batch:
                continue
            bundle, included = _build_reference_bundle(
                batch, ref_bundle_batch_max_chars, verbose=verbose
            )
            names = [m["name"] for m in batch]
            _log_prompt_text_stats(f"REFERENCE_{label}", bundle, ref_bundle_batch_max_chars, verbose=verbose)
            call_result = _call_claude_three_pass_split_input(
                    client,
                    input_text=combined_input,
                    ref_text=bundle,
                    ref_filename=f"references_{label}",
                    model=model,
                    max_tokens=max_tokens,
                    verbose=verbose,
                    multi_ref=True,
                    dual_batch_files=names,
                    custom_prompt_addendum=custom_prompt_addendum,
                    input_chunks=input_chunks_override,
                    input_max_chars=input_max_chars,
                    reference_batch_id=reference_batch_id,
                    debug_job_id=debug_job_id,
                    debug_s3_bucket=debug_s3_bucket,
                    debug_s3_prefix=debug_s3_prefix,
                )
            llm_calls += INPUT_SPLIT_LLM_CALLS
            diffs = _consume_call(call_result, label=f"dual/{label}")
            _ensure_ref_file_on_findings(diffs)
            acc.extend(diffs)

        all_diffs = _merge_batched_findings(findings_a, findings_b, verbose=verbose)

    else:
        if verbose:
            print(f"[2/3] per_file fallback: {len(measurements)} Claude call(s)...")
        for m in measurements:
            ref_fname = m["name"]
            if verbose:
                print(f"      → {ref_fname}")
            ref_text = _pages_to_str(m["pages_data"], max_chars=ref_max_chars)
            if verbose:
                print(
                    f"  [REFERENCE {ref_fname}] extracted_chars≈{m['chars']} "
                    f"sent_chars={len(ref_text)} cap={ref_max_chars}"
                )
            _log_prompt_text_stats(
                f"REFERENCE:{ref_fname}", ref_text, ref_max_chars, verbose=verbose
            )
            call_result = _call_claude_three_pass_split_input(
                    client,
                    input_text=combined_input,
                    ref_text=ref_text,
                    ref_filename=ref_fname,
                    model=model,
                    max_tokens=max_tokens,
                    verbose=verbose,
                    multi_ref=False,
                    custom_prompt_addendum=custom_prompt_addendum,
                    input_chunks=input_chunks_override,
                    input_max_chars=input_max_chars,
                    reference_batch_id=reference_batch_id,
                    debug_job_id=debug_job_id,
                    debug_s3_bucket=debug_s3_bucket,
                    debug_s3_prefix=debug_s3_prefix,
                )
            llm_calls += INPUT_SPLIT_LLM_CALLS
            diffs = _consume_call(call_result, label=f"per_file/{ref_fname}")
            for d in diffs:
                d["ref_file"] = ref_fname
            if verbose:
                print(f"         {len(diffs)} ממצאים זוהו")
            all_diffs.extend(diffs)

    if verbose:
        print(f"  [2/3] done strategy={strategy} llm_calls={llm_calls} findings={len(all_diffs)} "
              f"({_findings_count_summary(all_diffs)}) suppressed={len(all_suppressed)}")

    # Sort: severity descending (גבוהה → בינונית → נמוכה)
    _sev_order = {"גבוהה": 0, "בינונית": 1, "נמוכה": 2}
    all_diffs.sort(key=lambda d: _sev_order.get(d.get("severity", "נמוכה"), 3))

    # ------------------------------------------------------------------
    # 3. Write Excel
    # ------------------------------------------------------------------
    if verbose:
        print(f"[3/3] Writing Excel → {output_xlsx}  ({len(all_diffs)} rows total)")

    output_path = Path(output_xlsx)
    _build_excel(all_diffs, output_path)

    if verbose:
        print(f"Done. {output_path.resolve()}")

    warning = None
    if output_truncated and not all_diffs:
        warning = "output_truncated_zero_findings"
    elif output_truncated:
        warning = "output_truncated_partial" if salvaged_any else "output_truncated"

    return {
        "output_path": output_path,
        "differences": all_diffs,
        "pass_findings": all_pass_findings,
        "suppressed_findings": all_suppressed,
        "dedupe_dropped_findings": all_dedupe_dropped,
        "dedupe_before_count": dedupe_before_count,
        "dedupe_after_count": dedupe_after_count,
        "window_findings": all_window_findings,
        "page_retry_findings": all_page_retry_findings,
        "reference_strategy": strategy,
        "llm_calls": llm_calls,
        "output_truncated": output_truncated,
        "salvaged": salvaged_any,
        "warning": warning,
    }


# ---------------------------------------------------------------------------
# Findings JSON (Phoenix UI / chat compatibility)
# ---------------------------------------------------------------------------

def _default_bedrock_model() -> str:
    try:
        cfg = configparser.ConfigParser()
        cfg.read("config2.ini")
        return (
            cfg["bedrock"].get("model_id_sonnet_4_6")
            or "eu.anthropic.claude-sonnet-4-6"
        )
    except Exception:
        return "eu.anthropic.claude-sonnet-4-6"


def _he_to_en_severity(sev: str) -> str:
    return {"גבוהה": "high", "בינונית": "medium", "נמוכה": "low"}.get(sev, "low")


def _map_diff_type(diff_type_he: str) -> str:
    mapping = {
        "התאמה": "none",
        "שינוי סכום": "reduction",
        "הוספת כיסוי": "addition",
        "צמצום כיסוי": "reduction",
        "שינוי ניסוח": "meaning_change",
        "כיסוי חסר": "reduction",
        "שינוי אחוז": "reduction",
    }
    for k, v in mapping.items():
        if k in (diff_type_he or ""):
            return v
    return "meaning_change"


def build_findings_json(
    all_diffs: list[dict],
    job_id: str,
    input_batch_id: str,
    reference_batch_id: str,
) -> dict:
    """Convert insurance_diff rows into findings.json schema for the Phoenix UI."""
    findings = []
    for idx, d in enumerate(all_diffs, 1):
        sev_he = d.get("severity_he") or d.get("severity", "נמוכה")
        if sev_he in ("high", "medium", "low"):
            sev_he = {"high": "גבוהה", "medium": "בינונית", "low": "נמוכה"}.get(sev_he, "נמוכה")
        sev_en = d.get("severity") if d.get("severity") in ("high", "medium", "low") else _he_to_en_severity(str(sev_he))
        classification = _normalize_classification(d.get("classification", "difference"))
        is_material = d.get("is_material")
        if is_material is None:
            is_material = classification != "match" and sev_en == "high"
        conf = (d.get("confidence") or "medium").strip().lower()
        if conf not in ("high", "medium", "low"):
            conf = "medium"
        reason = (d.get("reason") or d.get("note") or "").strip()
        note = (d.get("note") or "").strip()
        input_page = d.get("input_page") or None
        ref_page = d.get("ref_page") or None
        input_text = d.get("input_quote", "")
        reference_text = d.get("ref_quote", "") or ""

        # BUG 6: non-match rows with no input page should carry an explicit marker
        if classification != "match" and not input_page:
            input_page = "N/A"

        # False-difference guard: ref_quote is absence meta-text or cites INPUT — reclassify as missing
        if classification == "difference" and _is_false_difference(d):
            print(
                f"  [false-difference guard] reclassified finding as missing "
                f"(no REFERENCE clause quoted): {(d.get('section') or '')[:60]}"
            )
            classification = "missing"
            reference_text = ""
            ref_page = None
            d = {**d, "ref_quote": "", "ref_page": "", "ref_file": "", "diff_type": "כיסוי חסר"}

        # BUG 5: missing rows with no reference text/page should carry explicit markers
        if classification == "missing" and not reference_text:
            if _is_false_missing(reason):
                reference_text = "לא צוטט מפוליסת הייחוס"
            else:
                reference_text = "לא נמצא בפוליסת הייחוס"
            ref_page = "N/A"

        # Self-contradiction guard: if a "difference" finding's own reason says there is no
        # difference, auto-reclassify as match. This catches LLM slippage where the model
        # correctly reasons no gap exists but still outputs classification="difference".
        _IDENTITY_PHRASES = (
            "אין הבדל", "אין פער", "זהה לחלוטין", "שניהם קובעים", "שני המסמכים קובעים",
            "identical", "no difference", "same in both", "no material difference",
            "אין הבדל מהותי", "זהים לחלוטין", "ללא הבדל",
        )
        if classification == "difference" and any(p in reason for p in _IDENTITY_PHRASES):
            print(
                f"  [self-contradiction guard] reclassified finding as match "
                f"(reason indicates no gap): {reason[:80]}"
            )
            classification = "match"
            sev_en = "low"
            sev_he = "נמוכה"
            is_material = False

        # Reverse self-contradiction guard: if a "match" finding's own reason or diff_type
        # signals a gap, reclassify as difference.  Three independent triggers:
        #   (a) diff_type is a substantive difference type — the LLM already knew it was a gap
        #   (b) reason contains phrases from the match Pre-Output Purge list
        #   (c) match with no ref_quote — prompts say reclassify as missing
        _RAW_DIFF_TYPE = (d.get("diff_type") or "").strip()
        _MATCH_GAP_DIFF_TYPES = {
            "שינוי ניסוח", "שינוי סכום", "צמצום כיסוי", "הוספת כיסוי",
            "כיסוי חסר", "שינוי אחוז", "אחר",
        }
        _GAP_PHRASES = (
            "מוסיף תנאי", "מוסיפה תנאי", "שינוי ניסוח", "מצמצם", "מצמצמת",
            "ללא תנאים", "הבדל מהותי", "השלכות פרשניות", "אינו מתנה",
            "מגביל", "מגבילה", "הגבלה", "ובלבד", "בתנאי ש", "אלא אם",
            "נוסף תנאי", "תנאי נוסף", "קיים הבדל", "יש הבדל", "יש פער",
        )
        if classification == "match":
            _gap_from_diff_type = _RAW_DIFF_TYPE in _MATCH_GAP_DIFF_TYPES
            _gap_from_reason = any(p in reason for p in _GAP_PHRASES)
            _no_ref_quote = not _has_substantive_ref_quote(
                _strip_quote_label(reference_text)
            )

            if _gap_from_diff_type or _gap_from_reason:
                _trigger = "diff_type" if _gap_from_diff_type else "reason"
                print(
                    f"  [false-match guard] reclassified finding as difference "
                    f"(trigger={_trigger}, diff_type={_RAW_DIFF_TYPE!r}): "
                    f"{(d.get('section') or '')[:60]}"
                )
                classification = "difference"
                # Preserve LLM's diff_type when it's substantive; fall back to שינוי ניסוח
                if not _RAW_DIFF_TYPE or _RAW_DIFF_TYPE == "התאמה":
                    d = {**d, "diff_type": "שינוי ניסוח"}
                # Severity: keep existing if already medium/high; set to medium otherwise
                if sev_en not in ("medium", "high"):
                    sev_en = "medium"
                    sev_he = "בינונית"
                is_material = sev_en in ("medium", "high")
            elif _no_ref_quote:
                print(
                    f"  [false-match guard] reclassified finding as missing "
                    f"(match with no ref_quote): {(d.get('section') or '')[:60]}"
                )
                classification = "missing"
                reference_text = "לא נמצא בפוליסת הייחוס"
                ref_page = "N/A"

        # Enforce match invariants: severity must be low and is_material must be false
        if classification == "match":
            sev_en = "low"
            sev_he = "נמוכה"
            is_material = False

        # BUG 4: match rows must never carry a substantive difference_type
        diff_type_mapped = _map_diff_type(d.get("diff_type", ""))
        if classification == "match":
            diff_type_mapped = "none"

        findings.append(
            {
                "finding_num": idx,
                "classification": classification,
                "is_material": bool(is_material),
                "severity": sev_en,
                "confidence": conf,
                "difference_type": diff_type_mapped,
                "input_text": input_text,
                # BUG 1: expose input quote as context so the HTML "הקשר קלט" column is populated
                "input_text_context": input_text,
                "reference_text": reference_text,
                "reason": reason,
                "note": note,
                "input_page": input_page,
                "ref_page": ref_page,
                "ref_file": d.get("ref_file", ""),
                "section": d.get("section", ""),
                "reference_file_name": d.get("ref_file", ""),
                "reference_page": ref_page,
                "reference_location": d.get("section", ""),
                "diff_type_he": d.get("diff_type", ""),
                "severity_he": sev_he,
                "source": "insurance_diff",
                "findings_source": "main_pass",
                "input_location": (
                    {
                        "chunk_index": d.get("_focus_chunk_start"),
                        "sub_chunk_index": d.get("_focus_chunk_end"),
                    }
                    if d.get("_focus_chunk_start") is not None
                    else {"chunk_index": None, "sub_chunk_index": None}
                ),
            }
        )

    diff_count = sum(1 for f in findings if f["classification"] == "difference")
    missing_count = sum(1 for f in findings if f["classification"] == "missing")
    match_count = sum(1 for f in findings if f["classification"] == "match")
    high_count = sum(1 for f in findings if f["severity"] == "high")

    return {
        "job_id": job_id,
        "input_batch_id": input_batch_id,
        "reference_batch_id": reference_batch_id,
        "total_findings": len(findings),
        "findings": findings,
        "summary": {
            # UI tiles use singular keys; keep plural for create_diff_table compat.
            "match": match_count,
            "difference": diff_count,
            "missing": missing_count,
            "matches": match_count,
            "differences": diff_count,
            "material": sum(1 for f in findings if f.get("is_material")),
            "high_severity": high_count,
        },
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }


# ---------------------------------------------------------------------------
# S3 helpers (Phoenix Lambda job)
# ---------------------------------------------------------------------------

def _download_s3_uri(uri: str, local_dir: str) -> str:
    rest = uri[5:]
    bucket, key = rest.split("/", 1)
    filename = key.split("/")[-1]
    local_path = os.path.join(local_dir, filename)
    S3_CLIENT.download_file(bucket, key, local_path)
    return local_path


# ---------------------------------------------------------------------------
# Chunk pre-loading helpers
# ---------------------------------------------------------------------------

def _load_s3_chunks(bucket: str, key: str, verbose: bool = True) -> Optional[list[dict]]:
    """Load a chunks JSON list from S3. Returns None on any error or missing key."""
    try:
        obj = S3_CLIENT.get_object(Bucket=bucket, Key=key)
        data = json.loads(obj["Body"].read().decode("utf-8"))
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            # Some chunk files wrap the list under a key
            for v in data.values():
                if isinstance(v, list):
                    return v
        return None
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in ("404", "NoSuchKey", "NotFound"):
            return None
        if verbose:
            print(f"[chunks] S3 error loading s3://{bucket}/{key}: {exc}")
        return None
    except Exception as exc:
        if verbose:
            print(f"[chunks] error loading s3://{bucket}/{key}: {exc}")
        return None


def _total_chunks_chars(chunks: list[dict]) -> int:
    """Return total character count across all chunk text fields."""
    return sum(len(c.get("text") or "") for c in chunks)


def _clean_chunk_text(text: str) -> str:
    """
    Remove formatting noise from PDF-extracted chunk text without altering real content.

    Cleans:
    - Markdown bold/italic markers (**text** / *text* / ***text***) — keeps inner text
    - Runs of 2+ bare stars with no content between them
    - Signature placeholder lines (3+ consecutive underscores)
    - Markdown heading prefixes (# / ## / etc. at line start)
    - Escaped double-quotes (\") → real "
    - Trailing whitespace on each line
    - 3+ consecutive blank lines collapsed to 2
    """
    # escaped quotes
    text = text.replace('\\"', '"')
    # markdown bold/italic: ***x***, **x**, *x* — keep inner text (non-greedy, dot matches newline excluded)
    text = re.sub(r'\*{1,3}(.+?)\*{1,3}', r'\1', text)
    # remaining bare star runs (≥2) — redaction masks / decorative
    text = re.sub(r'\*{2,}', '', text)
    # signature placeholder underscores (≥3)
    text = re.sub(r'_{3,}', '', text)
    # markdown heading markers at line start
    text = re.sub(r'^#+\s*', '', text, flags=re.MULTILINE)
    # trailing whitespace per line
    text = re.sub(r'[ \t]+$', '', text, flags=re.MULTILINE)
    # collapse 3+ consecutive blank lines to 2
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _chunks_to_pages(chunks: list[dict]) -> list[tuple[int, str]]:
    """Convert a chunk list to [(page_number, text), ...] for use with _pages_to_str."""
    pages: list[tuple[int, str]] = []
    for c in chunks:
        page_num = c.get("page_number") or c.get("page") or 1
        try:
            page_num = int(page_num)
        except (TypeError, ValueError):
            page_num = 1
        text = _clean_chunk_text(c.get("text") or "")
        if text:
            pages.append((page_num, text))
    return pages or [(1, "")]


def _find_ref_corpus_chunks_key(
    bucket: str,
    s3_prefix: str,
    reference_batch_id: str,
    verbose: bool = True,
) -> Optional[str]:
    """
    List S3 under {s3_prefix}/ref-corpus-{reference_batch_id}-* to find
    a reference_chunks.json file. Returns the S3 key or None.
    """
    search_prefix = f"{s3_prefix}/{REF_CORPUS_SUBDIR}-{reference_batch_id}-"
    try:
        paginator = S3_CLIENT.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=bucket, Prefix=search_prefix):
            for obj in page.get("Contents", []):
                key = obj["Key"]
                if key.endswith("/reference_chunks.json") or key.endswith("reference_chunks.json"):
                    return key
    except Exception as exc:
        if verbose:
            print(f"[chunks] S3 list error for prefix {search_prefix}: {exc}")
    return None


def _chunks_to_measurement(chunks: list[dict], name: str) -> dict:
    """
    Wrap a chunk list as a single measurements-style dict compatible with
    _build_reference_bundle / _choose_reference_strategy.
    """
    pages = _chunks_to_pages(chunks)
    chars = sum(len(t) for _, t in pages)
    return {
        "path": Path(name),
        "name": name,
        "chars": chars,
        "pages": len(pages),
        "pages_data": pages,
    }


def _chunks_to_measurements_by_file(
    chunks: list[dict],
    fallback_name: str,
) -> list[dict]:
    """
    Group chunks by file_name and return one measurements-style dict per unique file.
    Falls back to a single measurement named fallback_name when no file_name is present.
    This ensures _build_reference_bundle emits per-file headers so Claude can attribute
    each finding to the correct reference file.
    """
    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    for c in chunks:
        fname = (c.get("file_name") or c.get("source") or "").strip()
        groups[fname or fallback_name].append(c)
    measurements = []
    for fname, file_chunks in groups.items():
        pages = _chunks_to_pages(file_chunks)
        chars = sum(len(t) for _, t in pages)
        measurements.append({
            "path": Path(fname),
            "name": fname,
            "chars": chars,
            "pages": len(pages),
            "pages_data": pages,
        })
    return measurements


def _list_doc_uris(bucket: str, prefix: str) -> list[str]:
    uris: list[str] = []
    paginator = S3_CLIENT.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            ext = key.lower().rsplit(".", 1)[-1] if "." in key else ""
            if ext in DOC_EXTENSIONS:
                uris.append(f"s3://{bucket}/{key}")
    return sorted(uris)


def _resolve_input_s3_uris(
    bucket: str,
    job_id: str,
    input_s3_uri: Optional[str] = None,
) -> list[str]:
    if input_s3_uri:
        return [input_s3_uri]
    prefix = f"{INPUT_S3_PREFIX}/{job_id}/"
    uris = _list_doc_uris(bucket, prefix)
    if uris:
        return uris
    # Fallback: the chunker lambda saves input_s3_uris into chunk_status.json.
    # This allows debug re-runs of the diff step when the original PDF is no
    # longer present under doc-comparison-input/{job_id}/.
    try:
        import pdf_to_chunks as _ptc
        status = _ptc.load_chunk_status(bucket, DEFAULT_S3_PREFIX, job_id)
        if status and status.get("input_s3_uris"):
            saved = [u for u in status["input_s3_uris"] if u]
            if saved:
                print(
                    f"[insurance_diff] input_s3_uris resolved from chunk_status.json: {saved}"
                )
                return saved
    except Exception as _exc:
        print(f"[insurance_diff] chunk_status.json fallback failed: {_exc}")
    raise ValueError(
        f"No input document found under s3://{bucket}/{prefix}. "
        "Provide input_s3_uri in the /diff request."
    )


def _resolve_reference_s3_uris(bucket: str, reference_batch_id: str) -> list[str]:
    prefix = f"{REFERENCE_S3_PREFIX}/{reference_batch_id}/"
    uris = _list_doc_uris(bucket, prefix)
    if not uris:
        raise ValueError(
            f"No reference documents under s3://{bucket}/{prefix}. "
            "Upload reference files in the Reference tab first."
        )
    return uris


def _sanitize_debug_filename_part(s: str) -> str:
    cleaned = re.sub(r"[:/\\]+", "_", (s or "").strip()).replace(" ", "_")
    return cleaned.strip("_") or "unknown"


def _broken_json_debug_key(
    job_prefix: str, job_id: str, pass_label: str, ref_filename: str
) -> str:
    pass_part = _sanitize_debug_filename_part(pass_label or "unknown_pass")
    ref_part = _sanitize_debug_filename_part(Path(ref_filename).stem if ref_filename else "unknown_ref")
    prefix = job_prefix.rstrip("/")
    return f"{prefix}/{job_id}/debug/broken_json_{pass_part}_{ref_part}.json"


def _save_text_s3(text: str, bucket: str, key: str, *, content_type: str) -> None:
    S3_CLIENT.put_object(
        Bucket=bucket,
        Key=key,
        Body=text.encode("utf-8"),
        ContentType=content_type,
    )
    print(f"  Saved → s3://{bucket}/{key}")


def _save_broken_llm_json_to_s3(
    raw: str,
    *,
    pass_label: str,
    ref_filename: str,
    job_id: Optional[str],
    s3_bucket: Optional[str],
    s3_prefix: Optional[str],
) -> None:
    """Persist malformed LLM JSON for offline debugging (S3 only)."""
    if not raw or not job_id or not s3_bucket or not s3_prefix:
        return
    key = _broken_json_debug_key(s3_prefix, job_id, pass_label, ref_filename)
    try:
        _save_text_s3(
            raw,
            s3_bucket,
            key,
            content_type="application/json; charset=utf-8",
        )
    except Exception as exc:
        print(f"      [WARN] Failed to save broken JSON debug to S3: {exc}")


def _save_json_s3(obj: dict, bucket: str, key: str) -> None:
    body = json.dumps(obj, ensure_ascii=False, indent=2).encode("utf-8")
    S3_CLIENT.put_object(Bucket=bucket, Key=key, Body=body, ContentType="application/json")
    print(f"  Saved → s3://{bucket}/{key}")


def _save_excel_s3(local_path: Path, bucket: str, key: str) -> None:
    with open(local_path, "rb") as fh:
        S3_CLIENT.put_object(
            Bucket=bucket,
            Key=key,
            Body=fh.read(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    print(f"  Saved → s3://{bucket}/{key}")


def _findings_key(s3_prefix: str, job_id: str) -> str:
    return f"{s3_prefix}/{job_id}/findings.json"


def _diff_status_key(s3_prefix: str, job_id: str) -> str:
    return f"{s3_prefix}/{job_id}/diff_status.json"


def save_diff_status(
    s3_bucket: str,
    s3_prefix: str,
    job_id: str,
    state: str,
    message: str = "",
    job_s3_prefix: Optional[str] = None,
    **extra,
) -> None:
    """Track async /diff job state: queued | running | succeeded | failed."""
    effective_prefix = job_s3_prefix if job_s3_prefix is not None else s3_prefix
    doc = {
        "job_id": job_id,
        "state": state,
        "message": message,
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        **extra,
    }
    _save_json_s3(doc, s3_bucket, _diff_status_key(effective_prefix, job_id))


def _s3_object_exists(bucket: str, key: str) -> bool:
    try:
        S3_CLIENT.head_object(Bucket=bucket, Key=key)
        return True
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in ("404", "NoSuchKey", "NotFound"):
            return False
        raise


def run_insurance_diff_job(
    *,
    job_id: str,
    s3_bucket: str,
    reference_batch_id: str = "Phoenix_reference1",
    input_batch_id: Optional[str] = None,
    input_s3_uri: Optional[str] = None,
    input_s3_uris: Optional[list[str]] = None,
    reference_s3_uris: Optional[list[str]] = None,
    s3_prefix: str = DEFAULT_S3_PREFIX,
    job_s3_prefix: Optional[str] = None,
    output_dir: Optional[str | Path] = None,
    aws_region: str = DEFAULT_AWS_REGION,
    model: Optional[str] = None,
    input_max_chars: int = DEFAULT_INPUT_MAX_CHARS,
    ref_max_chars: int = DEFAULT_REF_MAX_CHARS,
    max_tokens: int = DEFAULT_MAX_TOKENS,
    reference_strategy: str = "auto",
    ref_bundle_single_max_chars: int = REF_BUNDLE_SINGLE_MAX_CHARS,
    ref_bundle_batch_max_chars: int = REF_BUNDLE_BATCH_MAX_CHARS,
    verbose: bool = True,
    enable_conflict_resolution: bool = False,
) -> dict:
    """
    Phoenix Step 4: download input + reference from S3, run full-document
    insurance_diff, upload findings.json + gaps.xlsx.

    Returns the same dict shape as legacy create_diff_table.create_diff_table().
    """
    if not job_id:
        raise ValueError("job_id is required")

    model = model or _default_bedrock_model()
    max_tokens = _resolve_max_tokens(max_tokens)
    input_batch_id = input_batch_id or job_id

    if input_s3_uris:
        input_uris = input_s3_uris
    else:
        input_uris = _resolve_input_s3_uris(s3_bucket, input_batch_id, input_s3_uri)

    # Check for pre-extracted chunks before requiring raw reference PDFs.
    # Hoist here so the variable is visible for logging below.
    _ref_chunks_key: Optional[str] = (
        _find_ref_corpus_chunks_key(s3_bucket, s3_prefix, reference_batch_id, verbose=False)
        if (s3_bucket and not reference_s3_uris) else None
    )

    if reference_s3_uris:
        ref_uris = reference_s3_uris
    else:
        # If chunks exist they will be used inside _run_insurance_diff_job_inner;
        # raw PDFs are only needed as a fallback when chunks are absent.
        if _ref_chunks_key is not None:
            ref_uris = []  # chunks found — raw PDFs not needed
        else:
            ref_uris = _resolve_reference_s3_uris(s3_bucket, reference_batch_id)

    if verbose:
        print(f"\n[insurance_diff job] job_id={job_id}")
        if input_batch_id != job_id:
            print(f"  input_batch_id:   {input_batch_id} (read-only source)")
        print(f"  input files:      {len(input_uris)}")
        if ref_uris:
            print(f"  reference files:  {len(ref_uris)}")
        else:
            chunks_note = (
                f" (pre-extracted chunks: {_ref_chunks_key})"
                if _ref_chunks_key
                else " (NONE — will fail)"
            )
            print(f"  reference files:  {chunks_note}")
        print(f"  reference_batch:  {reference_batch_id}")
        print(f"  model:            {model}")
        print(
            f"  limits:           input_max_chars={input_max_chars} "
            f"ref_max_chars={ref_max_chars} max_tokens={max_tokens} "
            f"reference_strategy={reference_strategy}"
        )

    if s3_bucket:
        save_diff_status(
            s3_bucket, s3_prefix, job_id, "running",
            job_s3_prefix=job_s3_prefix,
            message="Downloading files and calling Claude",
            input_uris=input_uris,
            reference_uris=ref_uris,
            reference_strategy=reference_strategy,
        )

    try:
        return _run_insurance_diff_job_inner(
            job_id=job_id,
            input_batch_id=input_batch_id,
            reference_batch_id=reference_batch_id,
            input_uris=input_uris,
            ref_uris=ref_uris,
            s3_bucket=s3_bucket,
            s3_prefix=s3_prefix,
            job_s3_prefix=job_s3_prefix,
            output_dir=output_dir,
            model=model,
            aws_region=aws_region,
            input_max_chars=input_max_chars,
            ref_max_chars=ref_max_chars,
            max_tokens=max_tokens,
            reference_strategy=reference_strategy,
            ref_bundle_single_max_chars=ref_bundle_single_max_chars,
            ref_bundle_batch_max_chars=ref_bundle_batch_max_chars,
            verbose=verbose,
            enable_conflict_resolution=enable_conflict_resolution,
        )
    except Exception as exc:
        if s3_bucket:
            save_diff_status(
                s3_bucket, s3_prefix, job_id, "failed",
                job_s3_prefix=job_s3_prefix,
                message=str(exc),
            )
        raise


def _run_insurance_diff_job_inner(
    *,
    job_id: str,
    input_batch_id: str,
    reference_batch_id: str,
    input_uris: list[str],
    ref_uris: list[str],
    s3_bucket: Optional[str],
    s3_prefix: str,
    job_s3_prefix: Optional[str] = None,
    output_dir: Optional[str | Path],
    model: str,
    aws_region: str,
    input_max_chars: int,
    ref_max_chars: int,
    max_tokens: int,
    reference_strategy: str,
    ref_bundle_single_max_chars: int,
    ref_bundle_batch_max_chars: int,
    verbose: bool,
    enable_conflict_resolution: bool = False,
) -> dict:
    # job_prefix is used for all job-specific artifact paths;
    # s3_prefix continues to be used for reference corpus lookups.
    job_prefix = job_s3_prefix if job_s3_prefix is not None else s3_prefix
    # ------------------------------------------------------------------
    # Load custom prompt addendum from DynamoDB (fail-safe)
    # ------------------------------------------------------------------
    custom_prompt_addendum = fetch_custom_prompt_for_use_case(reference_batch_id)

    # ------------------------------------------------------------------
    # Try loading pre-extracted chunks from S3 before downloading PDFs
    # ------------------------------------------------------------------
    combined_input_override: Optional[str] = None
    input_chunks_override: Optional[list[dict]] = None
    ref_measurements_override: Optional[list] = None

    if s3_bucket:
        # Input chunks — read from input_batch_id (source job in debug resume)
        input_chunks_key = f"{job_prefix}/{input_batch_id}/input_chunks.json"
        input_chunks_uri = f"s3://{s3_bucket}/{input_chunks_key}"
        print(f"[chunks] trying input: {input_chunks_uri}")
        input_chunks = _load_s3_chunks(s3_bucket, input_chunks_key, verbose)

        # Fallback: chunks may be stored at the flat base prefix when the job
        # was chunked without a use-case-scoped prefix.
        if input_chunks is None and job_prefix != s3_prefix:
            flat_key = f"{s3_prefix}/{input_batch_id}/input_chunks.json"
            flat_uri = f"s3://{s3_bucket}/{flat_key}"
            print(f"[chunks] input not found at scoped path, trying flat fallback: {flat_uri}")
            input_chunks = _load_s3_chunks(s3_bucket, flat_key, verbose)
            if input_chunks is not None:
                input_chunks_uri = flat_uri

        if input_chunks is None:
            print(f"[chunks] input NOT found or failed to load: {input_chunks_uri}")
        else:
            total_chars = _total_chunks_chars(input_chunks)
            if total_chars > CHUNKS_MIN_TEXT_CHARS:
                pages = _chunks_to_pages(input_chunks)
                combined_input_override = _pages_to_str(pages, max_chars=input_max_chars)
                input_chunks_override = input_chunks
                print(
                    f"[chunks] input OK  uri={input_chunks_uri}  "
                    f"chunks={len(input_chunks)}  chars={total_chars}"
                )
            else:
                print(
                    f"[chunks] input INVALID (total_chars={total_chars} < {CHUNKS_MIN_TEXT_CHARS}): "
                    f"{input_chunks_uri}"
                )

        # Reference chunks
        ref_key = _find_ref_corpus_chunks_key(s3_bucket, s3_prefix, reference_batch_id, verbose)
        if ref_key is None:
            print(
                f"[chunks] reference NOT found for batch_id={reference_batch_id} under "
                f"s3://{s3_bucket}/{s3_prefix}/{REF_CORPUS_SUBDIR}-{reference_batch_id}-*"
            )
        else:
            ref_chunks_uri = f"s3://{s3_bucket}/{ref_key}"
            print(f"[chunks] trying reference: {ref_chunks_uri}")
            ref_chunks = _load_s3_chunks(s3_bucket, ref_key, verbose)
            if ref_chunks is None:
                print(f"[chunks] reference NOT found or failed to load: {ref_chunks_uri}")
            else:
                total_chars = _total_chunks_chars(ref_chunks)
                if total_chars > CHUNKS_MIN_TEXT_CHARS:
                    ref_measurements_override = _chunks_to_measurements_by_file(
                        ref_chunks, reference_batch_id
                    )
                    print(
                        f"[chunks] reference OK  uri={ref_chunks_uri}  "
                        f"chunks={len(ref_chunks)}  files={len(ref_measurements_override)}  "
                        f"chars={total_chars}"
                    )
                else:
                    print(
                        f"[chunks] reference INVALID (total_chars={total_chars} < {CHUNKS_MIN_TEXT_CHARS}): "
                        f"{ref_chunks_uri}"
                    )

    if combined_input_override is not None and ref_measurements_override is not None:
        print("[chunks] BOTH input+reference loaded from chunks — skipping PDF download")
    else:
        missing = []
        if combined_input_override is None:   missing.append("input")
        if ref_measurements_override is None: missing.append("reference")
        print(
            f"[chunks] PARTIAL FALLBACK to PDF for: {', '.join(missing)} "
            f"(input_ok={combined_input_override is not None}, "
            f"ref_ok={ref_measurements_override is not None})"
        )

    # Excel must live outside TemporaryDirectory — the dir is deleted before S3 upload.
    tmp_excel = Path(f"/tmp/{job_id}_gaps.xlsx")

    with tempfile.TemporaryDirectory() as tmp_dir:
        input_local = (
            [_download_s3_uri(u, tmp_dir) for u in input_uris]
            if combined_input_override is None else []
        )
        ref_local = (
            [_download_s3_uri(u, tmp_dir) for u in ref_uris]
            if ref_measurements_override is None else []
        )
        result = analyze_insurance_docs(
            input_files=input_local,
            reference_files=ref_local,
            output_xlsx=tmp_excel,
            model=model,
            aws_region=aws_region,
            input_max_chars=input_max_chars,
            ref_max_chars=ref_max_chars,
            max_tokens=max_tokens,
            reference_strategy=reference_strategy,
            ref_bundle_single_max_chars=ref_bundle_single_max_chars,
            ref_bundle_batch_max_chars=ref_bundle_batch_max_chars,
            verbose=verbose,
            combined_input_override=combined_input_override,
            ref_measurements_override=ref_measurements_override,
            input_chunks_override=input_chunks_override,
            reference_batch_id=reference_batch_id,
            debug_job_id=job_id,
            debug_s3_bucket=s3_bucket,
            debug_s3_prefix=job_prefix,
        )

    all_diffs = result["differences"]
    strategy_used = result.get("reference_strategy", reference_strategy)
    llm_calls = result.get("llm_calls", 0)
    output_truncated = bool(result.get("output_truncated"))
    salvaged = bool(result.get("salvaged"))
    warning = result.get("warning")
    pass_findings = result.get("pass_findings", {})
    suppressed_findings = result.get("suppressed_findings", [])
    dedupe_dropped_findings = result.get("dedupe_dropped_findings", [])
    dedupe_before_count = result.get("dedupe_before_count", 0)
    dedupe_after_count = result.get("dedupe_after_count", len(all_diffs))
    window_findings = result.get("window_findings", [])
    page_retry_findings = result.get("page_retry_findings", [])
    findings_json = build_findings_json(
        all_diffs, job_id, input_batch_id, reference_batch_id
    )

    findings_key = _findings_key(job_prefix, job_id)
    excel_key = f"{job_prefix}/{job_id}/gaps.xlsx"

    if s3_bucket:
        _save_json_s3(findings_json, s3_bucket, findings_key)
        findings_path = f"s3://{s3_bucket}/{findings_key}"
        _save_excel_s3(tmp_excel, s3_bucket, excel_key)
        excel_path = f"s3://{s3_bucket}/{excel_key}"
        tmp_excel.unlink(missing_ok=True)
        # Save raw per-pass LLM outputs for debugging
        for pass_name in ("missing", "difference", "match"):
            raw_pass = pass_findings.get(pass_name, [])
            pass_key = f"{job_prefix}/{job_id}/findings_pass_{pass_name}.json"
            _save_json_s3(
                {
                    "job_id": job_id,
                    "pass": pass_name,
                    "findings_count": len(raw_pass),
                    "findings": raw_pass,
                },
                s3_bucket,
                pass_key,
            )
        # Save suppressed findings for QA recovery and debugging
        suppressed_key = f"{job_prefix}/{job_id}/suppressed_findings.json"
        _save_json_s3(
            {
                "job_id": job_id,
                "suppressed_count": len(suppressed_findings),
                "findings": suppressed_findings,
            },
            s3_bucket,
            suppressed_key,
        )
        # Save overlap-dedupe drops for debugging (e.g. 164 -> 70 window merge)
        dedupe_dropped_key = f"{job_prefix}/{job_id}/deduped_dropped_findings.json"
        _save_json_s3(
            {
                "job_id": job_id,
                "before_dedupe": dedupe_before_count,
                "after_dedupe": dedupe_after_count,
                "dropped_count": len(dedupe_dropped_findings),
                "findings": dedupe_dropped_findings,
            },
            s3_bucket,
            dedupe_dropped_key,
        )
        # Save raw per-window LLM outputs (ninth0..ninth8) before cross-window dedupe
        call_labels = {w.get("call_label", "") for w in window_findings}
        multi_call = len(call_labels - {"", "default"}) > 1
        for win_doc in window_findings:
            idx = win_doc.get("window_index", 0)
            call_label = win_doc.get("call_label") or ""
            if multi_call and call_label:
                safe = re.sub(r"[^\w.-]+", "_", call_label).strip("_")
                win_key = f"{job_prefix}/{job_id}/findings_window_{safe}_ninth{idx}.json"
            else:
                win_key = f"{job_prefix}/{job_id}/findings_window_ninth{idx}.json"
            _save_json_s3(
                {
                    "job_id": job_id,
                    "window_index": idx,
                    "window_label": win_doc.get("window_label", f"ninth{idx}"),
                    "focus_chunks": win_doc.get("focus_chunks"),
                    "focus_start": win_doc.get("focus_start"),
                    "focus_end": win_doc.get("focus_end"),
                    "call_label": call_label or None,
                    "findings_count": win_doc.get("findings_count", len(win_doc.get("findings") or [])),
                    "summary": win_doc.get("summary"),
                    "output_truncated": win_doc.get("output_truncated"),
                    "salvaged": win_doc.get("salvaged"),
                    "findings": win_doc.get("findings") or [],
                },
                s3_bucket,
                win_key,
            )
        if window_findings:
            _save_json_s3(
                {
                    "job_id": job_id,
                    "window_count": len(window_findings),
                    "windows": [
                        {
                            "window_index": w.get("window_index"),
                            "window_label": w.get("window_label"),
                            "focus_chunks": w.get("focus_chunks"),
                            "call_label": w.get("call_label"),
                            "findings_count": w.get("findings_count"),
                            "summary": w.get("summary"),
                            "s3_key": (
                                f"{job_prefix}/{job_id}/findings_window_ninth{w.get('window_index', 0)}.json"
                                if not multi_call
                                else None
                            ),
                        }
                        for w in window_findings
                    ],
                },
                s3_bucket,
                f"{job_prefix}/{job_id}/findings_windows_index.json",
            )
        call_labels_retry = {p.get("call_label", "") for p in page_retry_findings}
        multi_call_retry = len(call_labels_retry - {"", "default"}) > 1
        for page_doc in page_retry_findings:
            page_num = page_doc.get("page_number", 0)
            call_label = page_doc.get("call_label") or ""
            if multi_call_retry and call_label:
                safe = re.sub(r"[^\w.-]+", "_", call_label).strip("_")
                page_key = (
                    f"{job_prefix}/{job_id}/findings_page_retry_{safe}_page{page_num}.json"
                )
            else:
                page_key = f"{job_prefix}/{job_id}/findings_page_retry_{page_num}.json"
            _save_json_s3(
                {
                    "job_id": job_id,
                    "page_number": page_num,
                    "model": page_doc.get("model") or SPARSE_PAGE_RETRY_MODEL,
                    "findings_before_retry": page_doc.get("findings_before_retry"),
                    "chunk_count": page_doc.get("chunk_count"),
                    "llm_calls": page_doc.get("llm_calls"),
                    "chunk_retries": page_doc.get("chunk_retries", []),
                    "call_label": call_label or None,
                    "findings_count": page_doc.get("findings_count", len(page_doc.get("findings") or [])),
                    "summary": page_doc.get("summary"),
                    "output_truncated": page_doc.get("output_truncated"),
                    "salvaged": page_doc.get("salvaged"),
                    "findings": page_doc.get("findings") or [],
                },
                s3_bucket,
                page_key,
            )
        if page_retry_findings:
            total_chunk_calls = sum(
                int(p.get("llm_calls") or p.get("chunk_count") or 0)
                for p in page_retry_findings
            )
            _save_json_s3(
                {
                    "job_id": job_id,
                    "retry_zero_findings_only": True,
                    "min_findings_for_retry": 0,
                    "sparse_page_retry_model": SPARSE_PAGE_RETRY_MODEL,
                    "sparse_chunk_retry_top_k": SPARSE_CHUNK_RETRY_TOP_K_REF,
                    "pages_retried": len(page_retry_findings),
                    "total_chunk_calls": total_chunk_calls,
                    "pages": [
                        {
                            "page_number": p.get("page_number"),
                            "model": p.get("model"),
                            "findings_before_retry": p.get("findings_before_retry"),
                            "findings_count": p.get("findings_count"),
                            "chunk_count": p.get("chunk_count"),
                            "llm_calls": p.get("llm_calls"),
                            "summary": p.get("summary"),
                            "call_label": p.get("call_label"),
                        }
                        for p in page_retry_findings
                    ],
                },
                s3_bucket,
                f"{job_prefix}/{job_id}/findings_page_retry_summary.json",
            )
    else:
        out = Path(output_dir) if output_dir else Path(".")
        out.mkdir(parents=True, exist_ok=True)
        findings_path = str(out / f"{job_id}_findings.json")
        excel_path = str(out / f"{job_id}_gaps.xlsx")
        with open(findings_path, "w", encoding="utf-8") as fh:
            json.dump(findings_json, fh, ensure_ascii=False, indent=2)
        _build_excel(all_diffs, excel_path)

    status_msg = f"{len(all_diffs)} findings ({strategy_used}, {llm_calls} LLM call(s))"
    if output_truncated:
        status_msg += "; output was truncated at max_tokens"
        if salvaged:
            status_msg += " — recovered partial JSON"
        elif not all_diffs:
            status_msg += " — no parseable findings"

    result = {
        "job_id": job_id,
        "findings_path": findings_path,
        "excel_path": excel_path,
        "diff_count": len(all_diffs),
        "summary": findings_json["summary"],
        "reference_strategy": strategy_used,
        "llm_calls": llm_calls,
        "output_truncated": output_truncated,
        "salvaged": salvaged,
        "warning": warning,
    }

    if s3_bucket:
        save_diff_status(
            s3_bucket, s3_prefix, job_id, "succeeded",
            job_s3_prefix=job_s3_prefix,
            message=status_msg,
            diff_count=len(all_diffs),
            reference_strategy=strategy_used,
            llm_calls=llm_calls,
            output_truncated=output_truncated,
            salvaged=salvaged,
            warning=warning,
            enable_conflict_resolution=enable_conflict_resolution,
        )

    return result


def _findings_resolved_key(s3_prefix: str, job_id: str) -> str:
    return f"{s3_prefix}/{job_id}/findings_resolved.json"


def resolve_latest_findings(
    job_id: str,
    s3_bucket: str,
    s3_prefix: str = DEFAULT_S3_PREFIX,
    job_s3_prefix: Optional[str] = None,
) -> dict:
    """
    Return the newest findings artifact for a job.

    Priority: findings_resolved.json → findings.json
    (matches diff_job_status / Phoenix UI fetchFindings order).
    """
    if not job_id:
        raise ValueError("job_id is required")

    job_prefix = job_s3_prefix if job_s3_prefix is not None else s3_prefix
    resolved_key = _findings_resolved_key(job_prefix, job_id)
    raw_key = _findings_key(job_prefix, job_id)

    for artifact, key in (
        ("findings_resolved", resolved_key),
        ("findings", raw_key),
    ):
        if not _s3_object_exists(s3_bucket, key):
            continue
        obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=key)
        findings_doc = json.loads(obj["Body"].read().decode("utf-8"))
        return {
            "artifact": artifact,
            "s3_key": key,
            "s3_uri": f"s3://{s3_bucket}/{key}",
            "findings_doc": findings_doc,
            "findings_count": findings_doc.get(
                "total_findings", len(findings_doc.get("findings", []))
            ),
            "summary": findings_doc.get("summary", {}),
        }

    raise FileNotFoundError(
        f"No findings.json or findings_resolved.json for job_id={job_id} "
        f"under s3://{s3_bucket}/{job_prefix}/"
    )


def merge_page_retry_into_resolved(
    job_id: str,
    s3_bucket: str,
    s3_prefix: str = DEFAULT_S3_PREFIX,
    job_s3_prefix: Optional[str] = None,
    dry_run: bool = False,
) -> dict:
    """Merge all findings_page_retry_*.json files for a job into findings_resolved.json.

    Reads the existing ``findings_resolved.json`` (falling back to ``findings.json``),
    back-fills ``findings_source='main_pass'`` on every original finding, then reads
    every ``findings_page_retry_<N>.json`` file in the same S3 prefix, normalises its
    raw LLM findings through ``build_findings_json``, tags them
    ``findings_source='page_retry_<page_num>'``, deduplicates the combined list, and
    overwrites ``findings_resolved.json`` in S3.

    Parameters
    ----------
    job_id:        Phoenix job identifier.
    s3_bucket:     S3 bucket name.
    s3_prefix:     Top-level S3 prefix (default: ``doc-comparison``).
    job_s3_prefix: Override prefix used for this specific job (replaces s3_prefix
                   when constructing the job key path).
    dry_run:       If ``True``, compute and return stats without writing to S3.

    Returns
    -------
    dict with keys:
        merged_count   – total findings in the overwritten file
        original_count – findings that were in resolved/findings before merge
        retry_added    – net new findings from page retries (after dedup)
        pages_merged   – page numbers whose retry findings survived dedup
        pages_skipped  – retry files that contained 0 findings
        dedup_dropped  – duplicate findings removed during merge
        dry_run        – mirrors the ``dry_run`` parameter
    """
    if not job_id:
        raise ValueError("job_id is required")

    job_prefix = job_s3_prefix if job_s3_prefix is not None else s3_prefix

    # ── 1. Load the existing resolved/findings document ───────────────────────
    resolved_key = _findings_resolved_key(job_prefix, job_id)
    raw_key = _findings_key(job_prefix, job_id)
    write_key = resolved_key  # we always write back to findings_resolved.json

    base_doc: dict = {}
    for key in (resolved_key, raw_key):
        if _s3_object_exists(s3_bucket, key):
            obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=key)
            base_doc = json.loads(obj["Body"].read().decode("utf-8"))
            break
    if not base_doc:
        raise FileNotFoundError(
            f"No findings.json or findings_resolved.json for job_id={job_id} "
            f"under s3://{s3_bucket}/{job_prefix}/"
        )

    input_batch_id = base_doc.get("input_batch_id", job_id)
    reference_batch_id = base_doc.get("reference_batch_id", "")

    # ── 2. Back-fill findings_source on all original findings ─────────────────
    original_findings: list[dict] = base_doc.get("findings") or []
    for f in original_findings:
        f.setdefault("findings_source", "main_pass")
    original_count = len(original_findings)

    # ── 3. Discover all findings_page_retry_*.json keys for this job ──────────
    retry_prefix = f"{job_prefix}/{job_id}/findings_page_retry_"
    paginator = S3_CLIENT.get_paginator("list_objects_v2")
    retry_keys: list[str] = []
    for page in paginator.paginate(Bucket=s3_bucket, Prefix=retry_prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            # exclude the summary file and any per-call-label variants
            # (we only want the canonical findings_page_retry_<number>.json)
            basename = key.split("/")[-1]
            if basename == "findings_page_retry_summary.json":
                continue
            retry_keys.append(key)
    retry_keys.sort()

    if not retry_keys:
        return {
            "merged_count": original_count,
            "original_count": original_count,
            "retry_added": 0,
            "pages_merged": [],
            "pages_skipped": [],
            "dedup_dropped": 0,
            "dry_run": dry_run,
            "message": "No findings_page_retry_*.json files found for this job.",
        }

    # ── 4. Read and normalise each retry file ────────────────────────────────
    all_retry_findings: list[dict] = []
    pages_with_findings: list[int] = []
    pages_skipped: list[int] = []

    for key in retry_keys:
        obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=key)
        retry_doc = json.loads(obj["Body"].read().decode("utf-8"))
        page_num = retry_doc.get("page_number", 0)
        raw_findings: list[dict] = retry_doc.get("findings") or []

        if not raw_findings:
            pages_skipped.append(page_num)
            continue

        # Normalise raw LLM dicts into the standard findings schema
        normalised_doc = build_findings_json(
            raw_findings, job_id, input_batch_id, reference_batch_id
        )
        for f in normalised_doc.get("findings", []):
            f["findings_source"] = f"page_retry_{page_num}"

        all_retry_findings.extend(normalised_doc.get("findings", []))
        pages_with_findings.append(page_num)

    # ── 5. Merge + deduplicate ────────────────────────────────────────────────
    combined = original_findings + all_retry_findings
    deduped, dropped = _dedupe_findings(combined, verbose=True)

    # ── 6. Re-number and recompute summary ───────────────────────────────────
    for new_idx, f in enumerate(deduped, 1):
        f["finding_num"] = new_idx

    diff_count = sum(1 for f in deduped if f.get("classification") == "difference")
    missing_count = sum(1 for f in deduped if f.get("classification") == "missing")
    match_count = sum(1 for f in deduped if f.get("classification") == "match")
    high_count = sum(1 for f in deduped if f.get("severity") == "high")
    material_count = sum(1 for f in deduped if f.get("is_material"))

    merged_doc = {
        **base_doc,
        "total_findings": len(deduped),
        "findings": deduped,
        "summary": {
            "match": match_count,
            "difference": diff_count,
            "missing": missing_count,
            "matches": match_count,
            "differences": diff_count,
            "material": material_count,
            "high_severity": high_count,
        },
        "merged_page_retries": pages_with_findings,
        "merged_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }

    # ── 7. Write back (unless dry_run) ────────────────────────────────────────
    if not dry_run:
        _save_json_s3(merged_doc, s3_bucket, write_key)

    retry_added = len(deduped) - original_count
    return {
        "merged_count": len(deduped),
        "original_count": original_count,
        "retry_added": retry_added,
        "pages_merged": pages_with_findings,
        "pages_skipped": pages_skipped,
        "dedup_dropped": len(dropped),
        "dry_run": dry_run,
        "s3_key": write_key,
        "s3_uri": f"s3://{s3_bucket}/{write_key}",
    }


def diff_job_status(
    job_id: str,
    s3_bucket: str,
    s3_prefix: str = DEFAULT_S3_PREFIX,
    reference_batch_id: str = "Phoenix_reference1",
    job_s3_prefix: Optional[str] = None,
) -> dict:
    """Poll whether findings.json exists (for async /diff)."""
    job_prefix = job_s3_prefix if job_s3_prefix is not None else s3_prefix
    key = _findings_key(job_prefix, job_id)
    status_key = _diff_status_key(job_prefix, job_id)
    out: dict = {
        "job_id": job_id,
        "reference_batch_id": reference_batch_id,
        "ready": False,
        "status": "pending",
    }

    state = "pending"
    status_doc: dict = {}
    if _s3_object_exists(s3_bucket, status_key):
        obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=status_key)
        status_doc = json.loads(obj["Body"].read().decode("utf-8"))
        state = status_doc.get("state", "pending")
        out["status"] = state
        out["message"] = status_doc.get("message", "")
        if status_doc.get("output_truncated") is not None:
            out["output_truncated"] = bool(status_doc["output_truncated"])
        if status_doc.get("salvaged") is not None:
            out["salvaged"] = bool(status_doc["salvaged"])
        if status_doc.get("warning"):
            out["warning"] = status_doc["warning"]
        if status_doc.get("enable_conflict_resolution") is not None:
            out["enable_conflict_resolution"] = bool(status_doc["enable_conflict_resolution"])
        if state == "failed":
            out["error"] = status_doc.get("message", "Comparison failed")
            return out

    if state in ("queued", "running"):
        out["ready"] = False
        out["message"] = out.get("message") or "Comparison in progress"
        return out

    # Conflict resolution running after diff
    if state == "running_conflict_resolution":
        out["ready"] = False
        out["status"] = "running_conflict_resolution"
        out["message"] = out.get("message") or "Conflict resolution in progress..."
        return out

    resolved_key = _findings_resolved_key(job_prefix, job_id)

    if state == "succeeded":
        enable_conflict = bool(status_doc.get("enable_conflict_resolution"))

        if enable_conflict:
            if not _s3_object_exists(s3_bucket, resolved_key):
                out["ready"] = False
                out["status"] = "running_conflict_resolution"
                out["message"] = "Waiting for conflict resolution to start..."
                return out
            obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=resolved_key)
            findings_doc = json.loads(obj["Body"].read().decode("utf-8"))
            out["ready"] = True
            out["findings_path"] = f"s3://{s3_bucket}/{resolved_key}"
            out["excel_path"] = f"s3://{s3_bucket}/{job_prefix}/{job_id}/gaps.xlsx"
            out["diff_count"] = findings_doc.get("total_findings", len(findings_doc.get("findings", [])))
            out["summary"] = findings_doc.get("summary", {})
            out["conflict_resolved"] = True
            if "output_truncated" not in out:
                out["output_truncated"] = False
            return out

        if _s3_object_exists(s3_bucket, key):
            obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=key)
            findings_doc = json.loads(obj["Body"].read().decode("utf-8"))
            out["ready"] = True
            out["findings_path"] = f"s3://{s3_bucket}/{key}"
            out["excel_path"] = f"s3://{s3_bucket}/{job_prefix}/{job_id}/gaps.xlsx"
            out["diff_count"] = findings_doc.get("total_findings", len(findings_doc.get("findings", [])))
            out["summary"] = findings_doc.get("summary", {})
            if "output_truncated" not in out:
                out["output_truncated"] = False
            return out

        out["message"] = "Job finished; waiting for findings.json on S3"
        return out

    # Legacy jobs without diff_status.json: prefer resolved, then raw
    if _s3_object_exists(s3_bucket, resolved_key):
        obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=resolved_key)
        findings_doc = json.loads(obj["Body"].read().decode("utf-8"))
        out["ready"] = True
        out["status"] = "succeeded"
        out["findings_path"] = f"s3://{s3_bucket}/{resolved_key}"
        out["excel_path"] = f"s3://{s3_bucket}/{job_prefix}/{job_id}/gaps.xlsx"
        out["diff_count"] = findings_doc.get("total_findings", len(findings_doc.get("findings", [])))
        out["summary"] = findings_doc.get("summary", {})
        out["conflict_resolved"] = True
        return out

    if _s3_object_exists(s3_bucket, key):
        obj = S3_CLIENT.get_object(Bucket=s3_bucket, Key=key)
        findings_doc = json.loads(obj["Body"].read().decode("utf-8"))
        out["ready"] = True
        out["status"] = "succeeded"
        out["findings_path"] = f"s3://{s3_bucket}/{key}"
        out["excel_path"] = f"s3://{s3_bucket}/{job_prefix}/{job_id}/gaps.xlsx"
        out["diff_count"] = findings_doc.get("total_findings", len(findings_doc.get("findings", [])))
        out["summary"] = findings_doc.get("summary", {})
        return out

    if state == "pending":
        out["message"] = (
            "No diff job status found. The background Lambda may not have started "
            "(check IAM: lambda:InvokeFunction on lambda-create-diff)."
        )
    return out


# ---------------------------------------------------------------------------
# CLI convenience
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Compare insurance PDFs and produce a gap-analysis Excel file (via AWS Bedrock)."
    )
    parser.add_argument(
        "--aws-region",
        default=os.getenv("AWS_REGION", "us-east-1"),
        help="AWS region for Bedrock (default: us-east-1 or $AWS_REGION)",
    )
    parser.add_argument(
        "--aws-access-key",
        default=None,
        help="AWS access key ID (optional — uses default credential chain if omitted)",
    )
    parser.add_argument(
        "--aws-secret-key",
        default=None,
        help="AWS secret access key (optional — required if --aws-access-key is set)",
    )
    parser.add_argument(
        "--aws-session-token",
        default=None,
        help="AWS session token for temporary credentials (optional)",
    )
    parser.add_argument("--input",     nargs="+", required=True,  metavar="PDF",
                        help="Input / tender PDF file(s)")
    parser.add_argument("--reference", nargs="+", required=True,  metavar="PDF",
                        help="Reference spec PDF file(s)")
    parser.add_argument("--output",    default="insurance_gaps.xlsx", metavar="XLSX")
    parser.add_argument("--model",     default="eu.anthropic.claude-sonnet-4-6",
                        help="Bedrock model ID (default: eu.anthropic.claude-sonnet-4-6)")
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=DEFAULT_MAX_TOKENS,
        help=f"Bedrock max output tokens per reference call (default: {DEFAULT_MAX_TOKENS})",
    )
    parser.add_argument(
        "--input-max-chars",
        type=int,
        default=DEFAULT_INPUT_MAX_CHARS,
        help=f"Max INPUT chars sent to Claude (default: {DEFAULT_INPUT_MAX_CHARS})",
    )
    parser.add_argument(
        "--ref-max-chars",
        type=int,
        default=DEFAULT_REF_MAX_CHARS,
        help=f"Max REFERENCE chars per file (default: {DEFAULT_REF_MAX_CHARS})",
    )
    parser.add_argument(
        "--reference-strategy",
        default="auto",
        choices=("auto", "single", "dual", "per_file"),
        help="Reference compare strategy (default: auto)",
    )
    parser.add_argument("--quiet",     action="store_true")
    args = parser.parse_args()

    analyze_insurance_docs(
        input_files=args.input,
        reference_files=args.reference,
        output_xlsx=args.output,
        model=args.model,
        aws_region=args.aws_region,
        aws_access_key=args.aws_access_key,
        aws_secret_key=args.aws_secret_key,
        aws_session_token=args.aws_session_token,
        input_max_chars=args.input_max_chars,
        ref_max_chars=args.ref_max_chars,
        max_tokens=args.max_tokens,
        reference_strategy=args.reference_strategy,
        verbose=not args.quiet,
    )